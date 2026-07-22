"""
build_state_inputs.py — Sample 200 rows per state (NV, KY, FL, KS) from
Desktop/Providers_with NPI_Board.xlsx and write one input file per state
into PSV_DEV/ with a 'PSV Tab' sheet (same layout as existing Input.xlsx).

Usage:
    python build_state_inputs.py
"""

import random
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

SRC = Path(r"C:\Users\n661685\Desktop\Providers_with NPI_Board.xlsx")
OUT_DIR = Path(r"C:\Users\n661685\PSV_DEV")
STATES = ["NV", "KY", "FL", "KS"]
PER_STATE = 200
SEED = 42

C_LIC_STATE = 9   # License State column index (0-based)
C_NPI_NO = 14     # NPI_NO column index (0-based)

random.seed(SEED)

print(f"Reading {SRC} ...")
wb_src = openpyxl.load_workbook(str(SRC), read_only=True, data_only=True)
ws_src = wb_src["Sheet1"]
all_rows = list(ws_src.iter_rows(min_row=1, values_only=True))
wb_src.close()

header = list(all_rows[0])
data_rows = all_rows[1:]
print(f"  Total data rows: {len(data_rows)}")

# Group by state, require NPI_NO
buckets = defaultdict(list)
missing_npi = defaultdict(int)
for row in data_rows:
    state = str(row[C_LIC_STATE] or "").strip().upper()
    if state not in STATES:
        continue
    npi = row[C_NPI_NO]
    if not npi or str(npi).strip() in ("", "None", "nan"):
        missing_npi[state] += 1
        continue
    buckets[state].append(row)

print("\nAvailable rows with NPI by state:")
for s in STATES:
    print(f"  {s}: {len(buckets[s])} available (skipped {missing_npi[s]} missing NPI)")

# Sample PER_STATE randomly
output_files = {}
for s in STATES:
    pool = buckets[s]
    sample = random.sample(pool, min(PER_STATE, len(pool)))

    out_path = OUT_DIR / f"Input_200_{s}.xlsx"
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "PSV Tab"

    ws_out.append(header)
    for cell in ws_out[1]:
        cell.font = Font(bold=True)

    for row in sample:
        ws_out.append(list(row))

    wb_out.save(str(out_path))
    output_files[s] = out_path
    print(f"  Wrote {len(sample)} rows -> {out_path.name}")

print("\nAll input files created:")
for s, p in output_files.items():
    print(f"  {s}: {p}")
print("\nReady to run parallel PSV verification.")

