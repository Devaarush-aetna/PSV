"""
prep_jun_sample.py — Build a 100-row-per-state sample from Desktop/license_details_Jun.xlsx
for FL, KS, KY, and NV, then print the run_psv.py command.

Usage:
    python prep_jun_sample.py

Output:
    Input_Jun_FLKSKYNV.xlsx  (PSV Tab sheet, same column layout as existing inputs)

Run command (printed at end):
    python run_psv.py --input Input_Jun_FLKSKYNV.xlsx --states FL KS KY NV --no-ai
"""

import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

SRC = Path(r"C:\Users\n661685\Desktop\license_details_Jun.xlsx")
OUT = Path(r"C:\Users\n661685\PSV_DEV\Input_Jun_FLKSKYNV.xlsx")
SHEET_SRC = "Sheet1"
SHEET_OUT = "PSV Tab"
STATES = ["FL", "KS", "KY", "NV"]
PER_STATE = 100

# Column indices in the source file (0-based, matching psv_test.py constants)
# 0: First Name | 1: Middle Name | 2: Last Name | 3: EPDB PIN
# 4: Provider Type | 5: Source Code | 6: Status Eff Date | 7: Maintained By
# 8: Netid's | 9: License State | 10: LIC_TYPE_NM | 11: License ID
# 12: LIC_EXPRTN_DT | 13: LIC_PRDEXPN_DT | 14: NPI_NO
C_LIC_STATE = 9
C_NPI_NO = 14

print(f"Reading {SRC} …")
wb_src = openpyxl.load_workbook(str(SRC), read_only=True, data_only=True)
if SHEET_SRC not in wb_src.sheetnames:
    sys.exit(f"Sheet '{SHEET_SRC}' not found; available: {wb_src.sheetnames}")

ws_src = wb_src[SHEET_SRC]
all_rows = list(ws_src.iter_rows(min_row=1, values_only=True))
wb_src.close()

header = all_rows[0]
data_rows = all_rows[1:]
print(f"  Total data rows: {len(data_rows)}")
print(f"  Headers: {header}")

# Group rows by state; require NPI_NO present
buckets = defaultdict(list)
missing_npi = defaultdict(int)
for row in data_rows:
    state = str(row[C_LIC_STATE] or "").strip().upper()
    if state not in STATES:
        continue
    npi = row[C_NPI_NO]
    if not npi or str(npi).strip() in ("", "None", "nan"):
        missing_npi[state] += 1
        continue
    buckets[state].append(row)

print("\nAvailable rows with NPI_NO by state:")
for s in STATES:
    print(f"  {s}: {len(buckets[s])} (skipped {missing_npi[s]} rows missing NPI)")

# Take up to PER_STATE rows per state
sampled = {}
for s in STATES:
    sampled[s] = buckets[s][:PER_STATE]
    print(f"  Sampled {s}: {len(sampled[s])} rows")

# Write output Excel with 'PSV Tab' sheet
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = SHEET_OUT

# Write header row
ws_out.append(list(header))
for cell in ws_out[1]:
    cell.font = Font(bold=True)

total = 0
for s in STATES:
    for row in sampled[s]:
        ws_out.append(list(row))
        total += 1

wb_out.save(str(OUT))
print(f"\nWrote {total} rows → {OUT}")

print("\n" + "=" * 60)
print("Run command:")
print(f"  python run_psv.py --input {OUT} --states FL KS KY NV --no-ai")
print("=" * 60)
