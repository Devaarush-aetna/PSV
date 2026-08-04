"""Combine Standard Excel files from multiple PSV runs into one file.

Usage:
    python combine_standard.py --run-ids MA_Part1_v2 MA_Part2_v2 MA_Part3_v2
    python combine_standard.py --run-ids MA_Part1_v2 MA_Part2_v2 MA_Part3_v2 --out MA_Combined.xlsx

Reads Standard/*.xlsx from each run folder under PSV_DEV/Output/YYYYMM/<run-id>/
and writes a single combined Excel.
"""
from __future__ import annotations

import argparse
import glob
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

PSV_DEV = Path(__file__).parents[3]
OUTPUT_BASE = PSV_DEV / "Output"

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")


def find_standard_xlsx(run_id: str) -> Path:
    """Locate the Standard Excel for a given run_id."""
    pattern = str(OUTPUT_BASE / "??????" / run_id / "Standard" / "Standard_*.xlsx")
    matches = sorted(glob.glob(pattern))
    if not matches:
        raise FileNotFoundError(
            f"No Standard/*.xlsx found for run_id={run_id!r}\n"
            f"  Searched: {pattern}"
        )
    if len(matches) > 1:
        # Pick newest
        matches.sort(key=lambda p: Path(p).stat().st_mtime, reverse=True)
        print(f"  [warn] {run_id}: multiple Standard files found, using newest: {matches[0]}")
    return Path(matches[0])


def load_rows(path: Path) -> tuple[list[str], list[list]]:
    """Return (headers, data_rows) from an xlsx file."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h) for h in rows[0]]
    return headers, [list(r) for r in rows[1:]]


def combine(run_ids: list[str], out_path: Path) -> None:
    all_headers: list[str] = []
    all_data: list[tuple[list, list[str]]] = []  # (row_values, headers_for_this_file)

    for run_id in run_ids:
        try:
            src = find_standard_xlsx(run_id)
        except FileNotFoundError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(1)

        headers, rows = load_rows(src)
        if not headers:
            print(f"  [warn] {run_id}: Standard file is empty, skipping")
            continue

        print(f"  {run_id}: {len(rows)} rows  ({src.name})")

        if not all_headers:
            all_headers = headers
        else:
            # Warn on column mismatch but continue
            if headers != all_headers:
                extra = set(headers) - set(all_headers)
                missing = set(all_headers) - set(headers)
                if extra:
                    print(f"  [warn] {run_id}: extra columns not in first file: {sorted(extra)}")
                if missing:
                    print(f"  [warn] {run_id}: missing columns vs first file: {sorted(missing)}")

        for row in rows:
            all_data.append((row, headers))

    if not all_data:
        print("No rows to write — aborting.", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PSV Results"

    # Header row
    ws.append(all_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows — align each row to the master header
    for row_vals, row_headers in all_data:
        row_map = dict(zip(row_headers, row_vals))
        aligned = [row_map.get(h, "") for h in all_headers]
        ws.append(aligned)

        status = row_map.get("status", "")
        fill = GREEN if status == "Pass" else RED
        row_idx = ws.max_row
        for col in range(1, len(all_headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    wb.save(str(out_path))
    print(f"\nCombined {len(all_data)} rows → {out_path}")


def cli() -> None:
    p = argparse.ArgumentParser(description="Combine Standard Excels from multiple PSV runs")
    p.add_argument("--run-ids", nargs="+", required=True, metavar="RUN_ID",
                   help="run_ids to combine, in order (e.g. MA_Part1_v2 MA_Part2_v2 MA_Part3_v2)")
    p.add_argument("--out", default=None,
                   help="Output file path (default: PSV_DEV/MA_Combined_YYYYMMDD_HHMM.xlsx)")
    args = p.parse_args()

    if args.out:
        out_path = Path(args.out)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = PSV_DEV / f"MA_Combined_{ts}.xlsx"

    print(f"Combining {len(args.run_ids)} runs: {args.run_ids}")
    combine(args.run_ids, out_path)


if __name__ == "__main__":
    cli()
