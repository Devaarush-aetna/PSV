"""Reconstruct Standard output XLSX from on-disk Trace JSON files.

Use when a run was killed before emitter.flush() wrote the Standard file.
The trace JSONs are the source of truth for outcome/reason; the input Excel
supplies provider identity fields (name, license_id, etc.) not stored in traces.

Usage:
    python reconstruct_standard.py --run-id 20260727_1630_CT_001 \
        --input RunQueue/Input_CT_500.xlsx

Outputs:
    Output/YYYYMM/<run_id>/Standard/Standard_<run_id>_reconstructed.xlsx
    Output/YYYYMM/<run_id>/Standard/Standard_<run_id>_reconstructed.csv
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

PSV_DEV = Path(__file__).parents[3]

# PSV Tab column indices (must match psv_test.py)
C_FIRST_NAME   = 0
C_MIDDLE_NAME  = 1
C_LAST_NAME    = 2
C_EPDB_PIN     = 3
C_PROV_TYPE    = 4
C_MAINTAINED_BY = 7
C_LIC_STATE    = 9
C_LIC_TYPE     = 10
C_LIC_ID       = 11
C_LIC_EXPIRY   = 12

_NPI_HEADER_ALIASES = {"NPI_NO", "NPI", "NPI ID", "NPI_ID", "NPI NUMBER"}
_CAPTCHA_REASONS = frozenset({
    "state_captcha_blocked", "prov_type_captcha_blocked",
    "board_skip_captcha", "board_skipped",
})
_OUTCOME_RANK = {"Pass": 0, "Fail": 1, "Skip": 2}


def _load_input_rows(xlsx_path: Path, sheet_name: str = "PSV Tab") -> list[dict]:
    """Read PSV Tab into a list of row dicts (same mapping as psv_test.py)."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else (
         wb["PSV Tab"] if "PSV Tab" in wb.sheetnames else wb.active)
    raw = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not raw:
        return []

    # Detect header row
    first = raw[0]
    start = 0
    v = str(first[C_LIC_STATE] or "").strip().upper() if len(first) > C_LIC_STATE else ""
    if v in ("LIC_STATE", "C_LIC_STATE", "LICENSE STATE", "STATE", ""):
        start = 1

    # Discover NPI column index dynamically
    npi_col_idx = None
    if start == 1:
        for i, cell in enumerate(first):
            if cell is None:
                continue
            header = str(cell).strip().upper().replace("_", " ").replace(".", "")
            if header in _NPI_HEADER_ALIASES:
                npi_col_idx = i
                break

    rows = []
    for r in raw[start:]:
        def _cv(idx):
            if r is None or len(r) <= idx:
                return ""
            v = r[idx]
            return "" if v is None else str(v).strip()

        npi = ""
        if npi_col_idx is not None and len(r) > npi_col_idx:
            v = r[npi_col_idx]
            npi = "" if v is None else str(v).strip()

        rows.append({
            "first_name":    _cv(C_FIRST_NAME),
            "middle_name":   _cv(C_MIDDLE_NAME),
            "last_name":     _cv(C_LAST_NAME),
            "epdb_pin":      _cv(C_EPDB_PIN),
            "prov_type":     _cv(C_PROV_TYPE),
            "maintained_by": _cv(C_MAINTAINED_BY),
            "lic_state":     _cv(C_LIC_STATE).upper(),
            "lic_type":      _cv(C_LIC_TYPE),
            "license_id":    _cv(C_LIC_ID),
            "input_expiry":  _cv(C_LIC_EXPIRY),
            "npi_no":        npi,
        })
    return rows


def _infer_match_method(trace: dict, status: str) -> str:
    final_reason = (trace.get("final_reason") or "").strip()
    if status != "Pass" and final_reason in _CAPTCHA_REASONS:
        return "Captcha Based Board"
    if status != "Pass":
        return "none"
    if trace.get("escalate_to_ai_reason"):
        return "ai_fuzzy"
    if trace.get("nppes_used"):
        return "npi_substituted_exact"
    for att in reversed(trace.get("attempts", [])):
        wp = att.get("weight_profile_used") or ""
        if "license" in wp:
            return "exact_license"
        if wp:
            return "exact_name"
    return "exact_name"


def _board_source_id(trace: dict) -> str:
    for att in trace.get("attempts", []):
        if att.get("source_id"):
            sid = att["source_id"]
    return locals().get("sid", "")


def _build_standard_row(trace: dict, master: dict | None) -> dict:
    m = master or {}
    final_outcome = trace.get("final_outcome", "Fail")
    final_reason  = (trace.get("final_reason") or "").strip()
    attempts      = trace.get("attempts", [])

    if final_reason in _CAPTCHA_REASONS:
        status = "Skip"
    elif final_outcome in ("Pass", "Fail", "Skip"):
        status = final_outcome
    else:
        status = "Fail"

    match_method = _infer_match_method(trace, status)
    source_id    = _board_source_id(trace)

    return {
        "master_row_id":        trace.get("master_row_id", ""),
        "first_name":           m.get("first_name",    trace.get("first_name", "")),
        "middle_name":          m.get("middle_name",   ""),
        "last_name":            m.get("last_name",     trace.get("last_name", "")),
        "lic_state":            trace.get("state",     m.get("lic_state", "")),
        "prov_type":            trace.get("prov_type", m.get("prov_type", "")),
        "lic_type":             m.get("lic_type",      ""),
        "license_id":           m.get("license_id",   ""),
        "npi_no":               trace.get("npi_no",   m.get("npi_no", "")),
        "status":               status,
        "license_expiry":       "",   # not stored in trace
        "matched_license":      "",   # not stored in trace
        "matched_first":        "",   # not stored in trace
        "matched_last":         "",   # not stored in trace
        "board_name":           source_id,
        "match_method":         match_method,
        "fuzzy_score":          "",   # not stored in trace
        "weight_profile":       "",
        "ai_fallback_used":     bool(trace.get("escalate_to_ai_reason")),
        "npi_substituted":      bool(trace.get("nppes_used")),
        "nppes_used":           bool(trace.get("nppes_used")),
        "attempts_used":        len(attempts),
        "reason":               final_reason,
        "epdb_pin":             m.get("epdb_pin", ""),
        "epdb_name_score":      trace.get("epdb_name_score") or "",
        "nppes_name_score":     trace.get("nppes_name_score") or "",
        "trace_path":           "",   # filled after sorting
    }


def reconstruct(run_id: str, input_path: Path, sheet: str) -> Path:
    ym = run_id[:6]
    trace_dir = PSV_DEV / "Output" / ym / run_id / "Traces"
    std_dir   = PSV_DEV / "Output" / ym / run_id / "Standard"
    std_dir.mkdir(parents=True, exist_ok=True)

    if not trace_dir.is_dir():
        print(f"ERROR: Traces directory not found: {trace_dir}", file=sys.stderr)
        sys.exit(1)

    # Load input rows, build NPI → row and (first,last,state) → row lookups
    print(f"Loading input: {input_path}")
    input_rows = _load_input_rows(input_path, sheet)
    print(f"  {len(input_rows)} input rows loaded")

    npi_lookup: dict[str, dict] = {}
    name_lookup: dict[tuple, dict] = {}
    for r in input_rows:
        npi = r["npi_no"].strip()
        if npi:
            npi_lookup.setdefault(npi, r)
        key = (r["first_name"].upper(), r["last_name"].upper(), r["lic_state"])
        if key[0] or key[1]:
            name_lookup.setdefault(key, r)

    # Load all trace JSONs
    trace_files = sorted(trace_dir.glob("*.json"))
    print(f"  {len(trace_files)} trace files found")

    # Deduplicate: for the same NPI that appears in multiple traces
    # (from repeated resume runs), keep the one with the best outcome.
    best: dict[str, dict] = {}  # master_row_id -> trace dict
    dupe_npis: dict[str, list[dict]] = {}

    for tf in trace_files:
        try:
            t = json.loads(tf.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  WARN: could not read {tf.name}: {e}")
            continue

        npi = (t.get("npi_no") or "").strip()
        mrid = t.get("master_row_id", tf.stem)

        if npi and npi != "000":
            dupe_npis.setdefault(npi, []).append(t)
        else:
            best[mrid] = t

    for npi, traces in dupe_npis.items():
        if len(traces) == 1:
            t = traces[0]
        else:
            # Keep trace with best outcome; tie-break by most attempts (more complete run)
            traces.sort(key=lambda x: (
                _OUTCOME_RANK.get(x.get("final_outcome", "Fail"), 1),
                -len(x.get("attempts", [])),
            ))
            t = traces[0]
        mrid = t.get("master_row_id", npi)
        best[mrid] = t

    print(f"  {len(best)} unique rows after deduplication (removed {len(trace_files) - len(best)} duplicates)")

    # Build standard rows
    rows = []
    for mrid, trace in best.items():
        npi = (trace.get("npi_no") or "").strip()
        master = None
        if npi and npi != "000":
            master = npi_lookup.get(npi)
        if master is None:
            fn = trace.get("first_name", "")
            ln = trace.get("last_name", "")
            st = trace.get("state", "")
            master = name_lookup.get((fn.upper(), ln.upper(), st))

        row = _build_standard_row(trace, master)
        row["trace_path"] = str(trace_dir / f"{mrid}.json")
        rows.append(row)

    # Sort by row index embedded in master_row_id (row_NNNN_...)
    def _sort_key(r):
        m = re.search(r"row_(\d+)_", r.get("master_row_id", ""))
        return int(m.group(1)) if m else 9999999

    rows.sort(key=_sort_key)

    # Counters
    n_pass = sum(1 for r in rows if r["status"] == "Pass")
    n_fail = sum(1 for r in rows if r["status"] == "Fail")
    n_skip = sum(1 for r in rows if r["status"] == "Skip")
    rate   = 100 * n_pass / (n_pass + n_fail) if (n_pass + n_fail) > 0 else 0.0
    print(f"\nSummary: {len(rows)} rows  |  Pass={n_pass}  Fail={n_fail}  Skip={n_skip}  "
          f"PassRate={rate:.1f}%")

    # Write XLSX
    dt_tag = run_id  # use run_id as timestamp tag in filename
    xlsx_out = std_dir / f"Standard_{dt_tag}_reconstructed.xlsx"
    csv_out  = std_dir / f"Standard_{dt_tag}_reconstructed.csv"

    headers = list(rows[0].keys()) if rows else []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PSV Results"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    yel   = PatternFill("solid", fgColor="FFEB9C")
    for r in rows:
        ws.append([str(r.get(h, "") or "") for h in headers])
        fill = green if r["status"] == "Pass" else (yel if r["status"] == "Skip" else red)
        ri = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=ri, column=col).fill = fill
    wb.save(str(xlsx_out))
    print(f"\nXLSX -> {xlsx_out}")

    # Write CSV
    with open(csv_out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)
    print(f"CSV  -> {csv_out}")

    return xlsx_out


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--run-id",  required=True, help="Run ID, e.g. 20260727_1630_CT_001")
    p.add_argument("--input",   required=True, help="Path to input Excel file")
    p.add_argument("--sheet",   default="PSV Tab", help="Sheet name (default: PSV Tab)")
    args = p.parse_args()
    reconstruct(args.run_id, Path(args.input), args.sheet)


if __name__ == "__main__":
    main()
