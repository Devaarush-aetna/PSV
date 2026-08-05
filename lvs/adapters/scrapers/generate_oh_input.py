"""Generate clean input Excel for the 42 OH failing cases + 2 pass cases."""
import pandas as pd, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    ("Stawa",    "Rajab",     "Blanco",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0026867", "9999-12-31","2027-01-01","1073126314","OH",True),
    ("Nicole",   "Fritz",     "Fritz",        "OH","LPC", "A","2026-01-01","AHP","","OH","OPERATING",    "E.2404431",        "9999-12-31","2027-01-01","1326534975","OH",True),
    ("Riley",    "James",     "Miller",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0039506", "9999-12-31","2027-01-01","1508750746","OH",True),
    ("Erin",     "",          "Meckley",      "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0037578", "9999-12-31","2027-01-01","1104298249","OH",True),
    ("Nekia",    "",          "Jackson",      "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0033268", "9999-12-31","2027-01-01","1821606658","OH",True),
    ("Matthew",  "Edward",    "Raymond",      "OH","RNA", "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CRNA.0021407","9999-12-31","2027-01-01","1710561147","OH",True),
    ("Laura",    "",          "Babbitt",      "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.14484",   "9999-12-31","2027-01-01","1326428715","OH",True),
    ("Kenneth",  "L",         "Moore",        "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.14566",   "9999-12-31","2027-01-01","1013995323","OH",True),
    ("Ciara",    "",          "Haas",         "OH","PN",  "A","2026-01-01","AHP","","OH","OPERATING",    "0039670",          "9999-12-31","2027-01-01","1568059715","OH",True),
    ("Joanna",   "",          "Overholt",     "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0030764", "9999-12-31","2027-01-01","1194161687","OH",True),
    ("Sarah",    "",          "Jernigan",     "OH","DT",  "A","2026-01-01","AHP","","OH","OPERATING",    "86050463",         "9999-12-31","2027-01-01","1427714013","OH",True),
    ("Nadia",    "M.",        "Robinson",     "OH","MW",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0030218", "9999-12-31","2027-01-01","1407460488","OH",True),
    ("Tia",      "",          "King",         "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0036979", "9999-12-31","2027-01-01","1841740206","OH",True),
    ("Carol",    "",          "Green",        "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0028065", "9999-12-31","2027-01-01","1568026011","OH",True),
    ("Leanne",   "",          "Carman",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0039312", "9999-12-31","2027-01-01","1174315592","OH",True),
    ("Gregory",  "Michael",   "McDonnell",    "OH","DC",  "A","2026-01-01","AHP","","OH","OPERATING",    "ACUP-00177",       "9999-12-31","2027-01-01","1285174151","OH",True),
    ("Douglas",  "S",         "Larner",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0037279", "9999-12-31","2027-01-01","1467447896","OH",True),
    ("McKenna",  "",          "Waltenbaugh",  "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0036725", "9999-12-31","2027-01-01","1902650518","OH",True),
    ("Rachel",   "L.",        "Singer",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.13339",   "9999-12-31","2027-01-01","1093073546","OH",True),
    ("Maria",    "",          "Surmachevska", "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0039161", "9999-12-31","2027-01-01","1730897554","OH",True),
    ("Tiffany",  "",          "Colston",      "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0028286", "9999-12-31","2027-01-01","1083203665","OH",True),
    ("Cierra",   "Christine", "Ramey",        "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.026234",  "9999-12-31","2027-01-01","1265078307","OH",True),
    ("Lucille",  "Elizabeth", "Nathwani",     "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0032280", "9999-12-31","2027-01-01","1235740556","OH",True),
    ("April",    "",          "Kline",        "OH","LC",  "A","2026-01-01","AHP","","OH","OPERATING",    "10030039",         "9999-12-31","2027-01-01","1164752457","OH",True),
    ("Angela",   "",          "Lis",          "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0034610", "9999-12-31","2027-01-01","1245758978","OH",True),
    ("Carol",    "",          "Green",        "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0028065", "9999-12-31","2027-01-01","1023791449","OH",True),
    ("Stacey",   "L",         "Childress",    "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "COA.12146-NP",     "9999-12-31","2027-01-01","1083018030","OH",True),
    ("Christina","",          "Serger",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0038559", "9999-12-31","2027-01-01","1295466084","OH",True),
    ("Kayla",    "",          "McAdam",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0035678", "9999-12-31","2027-01-01","1457396368","OH",True),
    ("Kayla",    "",          "McAdam",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0035678", "9999-12-31","2027-01-01","1457113144","OH",True),
    ("Marco",    "",          "More",         "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0037287", "9999-12-31","2027-01-01","1659882470","OH",True),
    ("Veronica", "",          "Galaszewski",  "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0039555", "9999-12-31","2027-01-01","1194619700","OH",True),
    ("Maria",    "E",         "Carico",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.17686",   "9999-12-31","2027-01-01","1265813141","OH",True),
    ("Sylvie",   "",          "Riley",        "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0037507", "9999-12-31","2027-01-01","1285469361","OH",True),
    ("Jennifer", "L",         "Lakeberg",     "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.14477",   "9999-12-31","2027-01-01","1295168409","OH",True),
    ("Aissata",  "",          "Diallo",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0041396", "9999-12-31","2027-01-01","1336993351","OH",True),
    ("Sydney",   "",          "Waldon",       "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0036240", "9999-12-31","2027-01-01","1760952873","OH",True),
    ("Sarah",    "Michelle",  "Foltz",        "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.022942",  "9999-12-31","2027-01-01","1306315007","OH",True),
    ("Vickie",   "",          "Knueven",      "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0039311", "9999-12-31","2027-01-01","1083406516","OH",True),
    ("Umera",    "",          "Paracha",      "OH","PH",  "A","2026-01-01","AHP","","OH","STATE MEDICAL","25MA11601000",     "9999-12-31","2027-01-01","1558717850","OH",True),
    ("Michelle", "",          "Lepsesty",     "OH","NP",  "A","2026-01-01","AHP","","OH","OPERATING",    "APRN.CNP.0032579", "9999-12-31","2027-01-01","1811604556","OH",True),
    ("Katelyn",  "Emily",     "Dudenhoeffer", "OH","SH",  "A","2026-01-01","AHP","","OH","OPERATING",    "SP/11765",         "9999-12-31","2027-01-01","1689073421","OH",True),
]

COLS = [
    "First Name","Middle Name","Last Name","EPDB PIN","Provider Type","Source Code",
    "Status Eff Date","Maintained By","Netids","License State","LIC_TYPE_NM",
    "License ID","LIC_EXPRTN_DT","LIC_PRDEXPN_DT","NPI_NO",
    "Service Location State","Serv Loc Match Par State",
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Input"

hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ci, h in enumerate(COLS, 1):
    c = ws.cell(1, ci, h)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

for ri, row in enumerate(rows, 2):
    for ci, val in enumerate(row, 1):
        c = ws.cell(ri, ci, val)
        c.font = Font(size=9)
        c.alignment = Alignment(vertical="center")
        c.border = border

col_widths = [14,12,16,10,10,8,13,10,8,12,14,22,13,13,14,14,18]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

OUT = r"C:\Users\n676150\Downloads\OH_Input_42Cases.xlsx"
wb.save(OUT)
print("Saved:", OUT, "| rows:", len(rows))
