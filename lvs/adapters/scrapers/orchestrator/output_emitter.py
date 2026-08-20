"""Multi-channel output writer + per-row trace persister.

All files land under:
    PSV_DEV/Output/{YYYYMM}/{run_id}/{Channel}/{ChannelName}_{YYYYMMDD_HHMM}.ext

Channels:
  - Standard        Excel + CSV   every input row (includes routed_to column)
  - NPPES           CSV           every row, full NPPES record + diff vs master
  - AIFallback      CSV           every row where the AI agent ran
  - Manual          CSV           every unresolved row, with structured failure_reason
  - AddLicense      Excel         clean Pass rows ready for upload
  - AIAddLicense    Excel         "almost sure" rows — AI-resolved, partial/numeric
                                  matches, NPI-based, cross-row reconciliation;
                                  same AddLicense format + VerificationReason column.
                                  These rows are NOT duplicated in Manual.
  - RunSummary      CSV           per-state counters for the run
  - Drift           CSV           site-drift reports flagged by the AI agent
  - Traces/         JSON          per-row attempt log ({master_row_id}.json)
  - FallOut         Excel         client-facing error report — all Manual rows with
                                  provider identity, run metadata, and failure reason
"""
from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import date as _date
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill

from . import config as cfg
from . import disambiguator as disamb
from .disambiguator import (
    license_numerics_match as _lic_num_match,
    _NAME_PREFIXES_NORM,
    _NAME_SUFFIXES_NORM,
)
from .ai_agent import AiAgentResult
from .ladder import LadderResult
from .nppes_client import NpiDiscrepancy, NppesRecord
from .trace import RowTrace
from engine.models import LicenseStatus as _LicenseStatus

log = logging.getLogger(__name__)


def _clean_matched_name(raw: str) -> str:
    """Strip leading honorific prefixes (DR, MR, etc.) and trailing credential suffixes
    (MD, DM, RN, PHD, etc.) from a board-returned name part.

    Preserves the original casing and internal tokens; only whole leading/trailing
    tokens that exactly match the prefix/suffix sets are removed.
    """
    if not raw:
        return ""
    toks = str(raw).split()

    def _norm_tok(t: str) -> str:
        return re.sub(r"[.\-,]", "", t).upper()

    while toks and _norm_tok(toks[0]) in _NAME_PREFIXES_NORM:
        toks = toks[1:]
    while toks and _norm_tok(toks[-1]) in _NAME_SUFFIXES_NORM:
        toks = toks[:-1]
    return " ".join(toks)


# sites/ directory — used for lazy board_name lookups
_SITES_DIR = Path(__file__).resolve().parents[1] / "sites"

# Reason codes that produce status="Skip" in standard output.
# Superset — includes both captcha-blocked and transient-outage reasons.
_CAPTCHA_REASONS: frozenset[str] = frozenset({
    "state_captcha_blocked",
    "prov_type_captcha_blocked",
    "board_skip_captcha",
    "board_skipped",      # skip:true in board identity (e.g. BACB registry down)
    "board_unavailable",  # board site down/erroring at run time (timeout / HTTP 5xx)
})

# Subset of _CAPTCHA_REASONS whose match_method label is "Captcha Based Board".
# board_unavailable and board_skipped are transient outages, not captcha blocks —
# they get status=Skip but match_method="none".
_CAPTCHA_LABEL_REASONS: frozenset[str] = frozenset({
    "state_captcha_blocked",
    "prov_type_captcha_blocked",
    "board_skip_captcha",
})

# Manual-reason strings that route a row exclusively to AI_ADD_LICENSE (not Manual).
# These are "almost sure" matches a reviewer can approve with a quick look.
# Rows land in Manual only if _collect_ai_add_license returns False (no expiry date).
_REASONS_FOR_AI_ADD_LICENSE: frozenset[str] = frozenset({
    "AI fallback passed: manual review required to confirm verification result before use",
    "Numeric License ID matched",
    "License matched but Name mismatched",
    # "Name matched but License mismatched" intentionally excluded:
    # cross-name match without license confirmation must go to Manual, not AIAddLicense.
    "NPI used to fetch - manual review required",
    "Name mismatch after license match: EPDB and NPPES name scores both below 0.70 threshold",
    "Name verified: board record does not expose license number",
    "Name match accepted: board uses different license numbering",
    "Name verified: license number changed on renewal — review required",
})

# BACB boards are captcha-blocked — skip automated verification for any row
# routed exclusively to one of these source_ids.
_BACB_SOURCE_IDS: frozenset[str] = frozenset({"BACB"})
_BACB_CAPTCHA_REASON: str = (
    "Captcha Based Board: BACB (Behavior Analyst Certification Board) "
    "does not permit automated verification. Manual check required at bacb.com."
)

# National / multi-state registries that use their own credential numbering system
# (not state-issued license numbers). When the found record comes from one of these
# boards AND both name components match perfectly (≥ 1.0), skip the license-number
# comparison — the input may carry a state registration number (e.g. "NV20211995691")
# while the board stores its own certification ID (e.g. "L-163604").
_NATIONAL_REGISTRY_SOURCE_IDS: frozenset[str] = frozenset({
    "IBCLC_COMMISSION",  # IBCLCE certification; NV/other states may store state reg numbers
})

_board_name_cache: dict[str, str] = {}


def _get_board_name(source_id: str) -> str:
    """Return human-readable board_name for a source_id, reading config.yaml once."""
    if not source_id:
        return ""
    if source_id in _board_name_cache:
        return _board_name_cache[source_id]
    import yaml  # noqa: PLC0415
    cfg_path = _SITES_DIR / source_id / "config.yaml"
    try:
        data = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))
        name = data.get("identity", {}).get("board_name", source_id)
    except Exception:
        name = source_id
    _board_name_cache[source_id] = name
    return name


# Per-row outcome bundle that the runner builds and hands to emit_row().
@dataclass
class RowOutcome:
    master_row: dict
    master_row_id: str
    trace: RowTrace
    nppes: Optional[NppesRecord] = None
    discrepancy: Optional[NpiDiscrepancy] = None
    ladder_result: Optional[LadderResult] = None
    ai_result: Optional[AiAgentResult] = None

    @property
    def status(self) -> str:
        if self.ai_result and self.ai_result.outcome == "resolved":
            bd = self.ai_result.chosen_breakdown
            if bd is not None and not bd.gate_passed:
                return "Fail"
            return "Pass"
        if self.ladder_result and self.ladder_result.status == "Pass":
            return "Pass"
        return "Fail"

    @property
    def chosen_record(self) -> Optional[Any]:
        if self.ai_result and self.ai_result.outcome == "resolved":
            return self.ai_result.chosen_candidate
        # When a license-mismatch fallback found the right candidate but the AI gave
        # up (outcome reset to "gave_up"), use that candidate for expiry checks and
        # display so the unrelated expired ladder record doesn't surface as the match.
        if (self.ai_result
                and self.ai_result.chosen_candidate is not None
                and (self.trace.final_reason or "").strip() == "AI found License ID mismatched"):
            return self.ai_result.chosen_candidate
        if self.ladder_result:
            return self.ladder_result.best_record
        return None

    @property
    def chosen_breakdown(self) -> Optional[disamb.ScoreBreakdown]:
        if self.ai_result and self.ai_result.outcome == "resolved":
            return self.ai_result.chosen_breakdown
        if self.ladder_result:
            return self.ladder_result.best_breakdown
        return None

    @property
    def reason(self) -> str:
        # No-licensure-required always wins — overrides ai_result/ladder reason codes.
        if getattr(self.trace, "no_licensure_required", False) and self.trace.final_reason:
            return self.trace.final_reason
        if self.status == "Pass":
            # Surface the out-of-state verification state when applicable (FL T-licenses).
            if self.ladder_result and self.ladder_result.reason and self.ladder_result.reason.startswith("out_of_state:"):
                return self.ladder_result.reason
            return ""
        if self.ai_result and self.ai_result.reason:
            return self.ai_result.reason
        if self.ladder_result and self.ladder_result.reason:
            return self.ladder_result.reason
        if self.trace.final_reason:
            if self.trace.final_reason == "board_skipped" and self.trace.skip_reason_text:
                return self.trace.skip_reason_text
            return self.trace.final_reason
        return "no_records"


def _blank_state_stats() -> dict:
    return {
        "total":          0,
        "pass_rule":      0,  # Pass via rule-based (no AI, no NPI)
        "pass_ai":        0,  # AI resolved Pass (goes to manual not add_license)
        "pass_npi":       0,  # NPI substituted Pass
        "fail_rule":      0,  # Rule-based Fail
        "fail_ai":        0,  # AI attempted but failed
        "captcha":        0,  # Captcha / WAF blocked
        "mismatch":       0,  # Name/license cross-validation override to Fail
        "same_expiry":    0,  # Expiry same as input (no update needed)
        "expired_after_fetch": 0,  # Board expiry is in the past — not added to add_license
        "no_expiry":      0,  # Pass but board returned no expiry date
        "manual":           0,  # Total rows in manual channel
        "add_license":      0,  # Total rows in add_license channel
        "ai_add_license":   0,  # Total rows in ai_add_license channel
        "ai_used":        0,  # Any row where AI agent ran
        "ai_resolved":    0,  # AI resolved (outcome == "resolved")
        "ai_failed":      0,  # AI ran but did not resolve
        "npi_substituted": 0, # NPI was used to find the board record
        # AI token / cost aggregates (summed across all rows in this state)
        "ai_input_tokens":  0,
        "ai_output_tokens": 0,
        "ai_usd_cost":      0.0,
    }


@dataclass
class OutputEmitter:
    run_id: str
    dirs: dict[str, Path] = field(default_factory=dict)
    _standard_rows: list[dict] = field(default_factory=list)
    _nppes_rows: list[dict] = field(default_factory=list)
    _ai_rows: list[dict] = field(default_factory=list)
    _manual_rows: list[dict] = field(default_factory=list)
    _add_license_rows: list[dict] = field(default_factory=list)
    _ai_add_license_rows: list[dict] = field(default_factory=list)
    _fallout_rows: list[dict] = field(default_factory=list)
    _state_stats: dict[str, dict] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if not self.dirs:
            self.dirs = cfg.ensure_channel_dirs(self.run_id)

    # ----- Per-row entry point -----

    @staticmethod
    def _resolve_board_name_parts(rec: Optional[Any], master_last: str = "") -> tuple[str, str]:
        """Return (first, last) from a board record, falling back to splitting
        licensee_full_name when separate first/last fields are absent (e.g. IN_PLA).
        Returns ("", "") when rec is None.
        """
        if rec is None:
            return "", ""
        first = (getattr(rec, "licensee_first_name", "") or "").strip()
        last  = (getattr(rec, "licensee_last_name",  "") or "").strip()
        if not first and not last:
            full = (getattr(rec, "licensee_full_name", "") or "").strip()
            if full:
                first, last = disamb._split_full_name(full, master_last)
        elif not first and last and " " in last:
            # Board stored full name in licensee_last_name (e.g. "George Joseph Vesper").
            # Split it so first/last columns are populated correctly in output.
            first, last = disamb._split_full_name(last, master_last)
        elif first and last and " " in last and last.split()[0].strip().upper() == first.strip().upper():
            # Board populated first AND dumped the full name into licensee_last_name
            # (e.g. NC_SLP_AUD detail "Name" field → first="John", last="John
            # Rutherfoord Smith"). The last field leads with the first name, so it is
            # really the full name — re-split so the last column holds only the surname.
            _f, _l = disamb._split_full_name(last, master_last)
            if _l:
                last = _l
                first = first or _f
        return first, last

    @staticmethod
    def _name_license_mismatch_reason(outcome: "RowOutcome") -> str | None:
        """Name matched but license numbers are present on both sides and don't align."""
        bd = outcome.chosen_breakdown
        rec = outcome.chosen_record
        if bd is None or rec is None:
            return None
        if bd.license_numerics >= 1.0:
            return None  # numeric match — no issue
        board_lic = (getattr(rec, "license_number", "") or "").strip()
        input_lic = (outcome.master_row.get("license_id", "") or "").strip()
        if not board_lic or not input_lic:
            return None  # one side has no license to compare
        # Re-check numeric equivalence directly on the current record values.
        # The ScoreBreakdown may have been computed before the detail page populated
        # license_number (e.g. IN_PLA results table has no license in the summary row;
        # detail page sets it later). In that case license_numerics is stale (0.0) even
        # though board and input licenses are actually the same — avoid a false mismatch.
        if disamb.license_numerics_match(input_lic, board_lic):
            return None
        # National registries (IBCLCE etc.) use their own credential numbering that differs
        # from state-issued registration numbers. When both name components match perfectly
        # and the record comes from a known national registry, the license-number difference
        # is an expected cross-system artifact — not a wrong-person mismatch.
        src = getattr(rec, "source_id", "") or ""
        if src in _NATIONAL_REGISTRY_SOURCE_IDS and bd.first_name >= 1.0 and bd.last_name >= 1.0:
            return None
        return "Name matched but License mismatched"

    @staticmethod
    def _same_expiry_check(outcome: "RowOutcome") -> str | None:
        """If the board expiry matches the input expiry, return a manual reason.
        Future match -> 'Provider has the same Expiry as input, still in 90 days'
        Past match   -> 'Expired and same date in the State board'
        Returns None when either date is absent or they differ.
        """
        rec = outcome.chosen_record
        if rec is None:
            return None
        input_expiry_str = (outcome.master_row.get("input_expiry", "") or "").strip()
        if not input_expiry_str:
            return None
        board_expiry_str = _expiry_str(rec)
        if not board_expiry_str:
            return None
        try:
            input_date = _date.fromisoformat(input_expiry_str[:10])
            board_date = _date.fromisoformat(board_expiry_str[:10])
        except Exception:
            return None
        if input_date != board_date:
            return None
        today = _date.today()
        if board_date >= today:
            return "Check again later for updates, same as input"
        return "Expired and same as input"

    @staticmethod
    def _license_name_mismatch_reason(outcome: "RowOutcome") -> str | None:
        """License matched numerically but first or last name doesn't align.
        Only fires when both input and board have the relevant name fields.
        """
        bd = outcome.chosen_breakdown
        rec = outcome.chosen_record
        if bd is None or rec is None:
            return None
        if bd.license_numerics < 1.0:
            return None  # license didn't match — different check handles this
        # NAME_FUZZ_MIN default is 0.7 — same threshold the disambiguator gate uses
        _THRESHOLD = 0.7
        input_first = (outcome.master_row.get("first_name", "") or "").strip()
        input_last = (outcome.master_row.get("last_name", "") or "").strip()
        first_fail = input_first and bd.first_name < _THRESHOLD
        last_fail = input_last and bd.last_name < _THRESHOLD
        if first_fail or last_fail:
            # Low last name (<0.40) but strong first name (≥0.85): name-change case
            # (marriage/divorce) — same person, different last name.  Route to
            # AIAddLicense for human review, not Manual.
            # Low last name AND low/absent first name: truly different person with
            # coincidentally matching license digits — route to Manual.
            if input_last and bd.last_name < 0.40:
                if bd.first_name >= 0.85:
                    return "License matched but Name mismatched"
                return "Wrong provider matched: license matched but last name is completely different"
            return "License matched but Name mismatched"
        return None

    @staticmethod
    def _numeric_not_exact_license_reason(outcome: "RowOutcome") -> str | None:
        """Name matched AND license numerics aligned, but the license strings are not
        identical (e.g. input '12345' vs board 'LC-12345', or '017371' vs '17371').
        The digits agree but the full value differs — flag for human confirmation.
        Returns None when the license is an exact alphanumeric match or when
        license numerics did not align at all.
        """
        import re as _re
        bd = outcome.chosen_breakdown
        rec = outcome.chosen_record
        if bd is None or rec is None:
            return None
        if bd.license_numerics < 1.0:
            return None  # no numeric alignment — handled elsewhere
        board_lic = (getattr(rec, "license_number", "") or "").strip()
        input_lic = (outcome.master_row.get("license_id", "") or "").strip()
        if not board_lic or not input_lic:
            return None
        # Strip all non-alphanumeric and compare case-insensitively
        _strip = lambda s: _re.sub(r"[^a-z0-9]", "", s.lower())
        _si = _strip(input_lic)
        _sb = _strip(board_lic)
        if _si == _sb:
            return None  # effectively identical — no review needed
        # Pure-numeric leading-zero mismatch (e.g. EPDB "7278" vs board "007278")
        if _si.isdigit() and _sb.isdigit() and _si.lstrip("0") == _sb.lstrip("0"):
            return None
        # Dot-prefix artifact in input (e.g. EPDB "14.015029" vs board "015029")
        if "." in input_lic and _sb.isdigit():
            for _seg in input_lic.split("."):
                _ss = _strip(_seg)
                if _ss.isdigit() and _ss.lstrip("0") == _sb.lstrip("0"):
                    return None
        # Letter-prefix + digit core vs double-prefix.digit.suffix
        # (e.g. WA: "RN61176701" vs "RN.RN.61176701.MSL")
        # Do NOT suppress when input is pure digits but board has a letter prefix
        # (e.g. input "2391" vs board "LCPC 2391") — the prefix is significant.
        _digit_run = lambda s: max((_re.findall(r"\d+", s) or [""]), key=len)
        _di = _digit_run(_si)
        _db = _digit_run(_sb)
        _input_pure_digit = _si.isdigit()
        _board_has_letters = not _sb.isdigit()
        if (len(_di) >= 4 and _di.lstrip("0") == _db.lstrip("0")
                and not (_input_pure_digit and _board_has_letters)):
            return None
        return "Numeric License ID matched"

    @staticmethod
    def _expired_after_fetch_reason(outcome: "RowOutcome") -> str | None:
        """Return 'Provider fetch after Expiry' when the board record is expired.

        Two triggers (either is sufficient):
          1. Board record's status field is LicenseStatus.EXPIRED — catches boards that
             mark a license "Expired" without populating an expiry date.
          2. Board-returned expiry date is in the past — catches boards that return a
             date but no explicit status (or an ambiguous status like "Inactive").
        Only fires on Pass rows where a matched board record exists.
        """
        rec = outcome.chosen_record
        if rec is None:
            return None
        # Check 1: board explicitly returned status=Expired
        rec_status = getattr(rec, "status", None)
        if rec_status == _LicenseStatus.EXPIRED:
            return "Provider fetch after Expiry"
        # Check 2: board-returned expiry date is in the past
        board_expiry_str = _expiry_str(rec)
        if not board_expiry_str:
            return None
        try:
            board_date = _date.fromisoformat(board_expiry_str[:10])
        except Exception:
            return None
        if board_date < _date.today():
            return "Provider fetch after Expiry"
        return None

    # Human-readable manual-reason messages, keyed by final_reason code.
    _CAPTCHA_MANUAL_REASONS: dict[str, str] = field(default_factory=lambda: {
        "state_captcha_blocked": (
            "Captcha Based Board: Entire state is blocked from automated verification "
            "(CAPTCHA / IP restriction). Manual verification required on the state board website."
        ),
        "prov_type_captcha_blocked": (
            "Captcha Based Board: This provider type is blocked from automated verification "
            "(CAPTCHA / IP restriction on state board). Manual verification required."
        ),
        "board_skip_captcha": (
            "Captcha Based Board: State board blocks automated access via CAPTCHA, McAfee, "
            "DataDome, or WAF. Manual verification required on the state board website."
        ),
        "board_skipped": (
            "Board Skipped: Registry is currently unavailable (under maintenance or access "
            "restriction). Automated verification was not attempted. "
            "Manual verification required on the board's website."
        ),
        "board_unavailable": (
            "Board Unavailable: The board website was unreachable at verification time "
            "(connection timeout or HTTP server error). This is a temporary board-side "
            "outage — re-run this row once the site is back online."
        ),
    })

    def collect(self, outcome: RowOutcome) -> None:
        """Capture one row's outputs into the in-memory channel buffers.

        Architecture:
          1. Standard channel  — every row, always.
          2. NPPES / AI channels — when applicable.
          3. Compute a single manual_reason (first match wins).
          4. If manual_reason -> manual channel only, never add_license.
             If no manual_reason AND Pass with expiry -> add_license only.
          Standard and add_license are populated independently;
          manual and add_license are always mutually exclusive.
        """
        # Override trace outcome to Fail before persisting the JSON so that the
        # trace file is consistent with the Standard CSV status column.
        # The ladder sets final_outcome="Pass" whenever a board record is found,
        # but an expired record must surface as Fail everywhere.
        if self._expired_after_fetch_reason(outcome):
            outcome.trace.final_outcome = "Fail"
            outcome.trace.final_reason = "Provider fetch after Expiry"

        outcome.trace.write_json(self.dirs["trace"])
        self._collect_standard(outcome)
        self._collect_nppes(outcome)
        if outcome.ai_result is not None:
            self._collect_ai(outcome)

        manual_reason = self._compute_manual_reason(outcome)

        # Apply any retroactive standard-row overrides driven by the manual reason
        if manual_reason in (
            "Name matched but License mismatched",
            "License matched but Name mismatched",
        ):
            self._standard_rows[-1]["status"] = "Fail"
            self._standard_rows[-1]["match_method"] = "name_license_mismatch"
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == "Temporary License ID":
            # Keep status as Pass — temp cert holders are verified by name match.
            # Route to Manual so reviewers can record the board's permanent license.
            self._standard_rows[-1]["match_method"] = "temporary_license_id"
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == "Expired and same as input":
            self._standard_rows[-1]["status"] = "Fail"
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == "Check again later for updates, same as input":
            self._standard_rows[-1]["status"] = "Fail"
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == "Provider fetch after Expiry":
            self._standard_rows[-1]["status"] = "Fail"
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == "Numeric License ID matched":
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == (
            "Name mismatch after license match: "
            "EPDB and NPPES name scores both below 0.70 threshold"
        ):
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason and manual_reason.startswith("Low match score"):
            self._standard_rows[-1]["reason"] = manual_reason
        elif manual_reason == (
            "AI fallback passed: manual review required to confirm "
            "verification result before use"
        ):
            self._standard_rows[-1]["reason"] = manual_reason

        # Route: manual XOR add_license; AI_ADD_LICENSE rows are excluded from manual
        # (AIAddLicense IS the manual file for these — reviewers only need one place to look).
        # If an AI_ADD_LICENSE-qualifying row has no expiry, it falls back to manual only.
        went_manual = False
        went_add_license = False
        went_ai_add_license = False
        _epdb = str(outcome.master_row.get("epdb_pin", "") or "").strip()
        if manual_reason:
            if "not in Service Location State" in manual_reason:
                # Row belongs to RemoveLicense channel — do not route to Manual.
                self._standard_rows[-1]["routed_to"] = "RemoveLicense"
                self._accumulate_state_stats(outcome, manual_reason, False, False, False)
                return
            elif manual_reason in _REASONS_FOR_AI_ADD_LICENSE:
                if not _epdb:
                    self._collect_manual(outcome, failure_reason="EPDB is blanks")
                    went_manual = True
                elif self._collect_ai_add_license(outcome, manual_reason):
                    went_ai_add_license = True
                else:
                    self._collect_manual(outcome, failure_reason=manual_reason)
                    went_manual = True
            else:
                self._collect_manual(outcome, failure_reason=manual_reason)
                went_manual = True
        elif outcome.status == "Pass":
            if not _epdb:
                self._collect_manual(outcome, failure_reason="EPDB is blanks")
                went_manual = True
            else:
                added = self._collect_add_license(outcome)
                if added:
                    went_add_license = True
                else:
                    # No expiry returned — board verified the person but gave no expiry date.
                    # Treat as Fail: we cannot confirm the license is current without a term date.
                    _no_expiry_reason = (
                        "no_expiry_date: license verified on state board but expiration "
                        "date not returned - manual review required to confirm LicenseTermDate"
                    )
                    self._standard_rows[-1]["status"] = "Fail"
                    self._standard_rows[-1]["reason"] = _no_expiry_reason
                    self._collect_manual(outcome, failure_reason=_no_expiry_reason)
                    went_manual = True

        # Catch-all: if any manual_reason was not matched by the explicit override block
        # above, backfill it now so no future reason string silently slips through.
        if went_manual and manual_reason and not self._standard_rows[-1].get("reason"):
            self._standard_rows[-1]["reason"] = manual_reason

        # Tag the standard row with its routing destination
        if went_ai_add_license:
            self._standard_rows[-1]["routed_to"] = "AIAddLicense"
            # Rows routed to AIAddLicense represent "almost sure" matches approved for
            # upload — treat as Pass regardless of any intermediate Fail flag (e.g. name
            # mismatch that was resolved by fuzzy score above threshold).
            self._standard_rows[-1]["status"] = "Pass"
        elif went_add_license:
            self._standard_rows[-1]["routed_to"] = "AddLicense"
        elif went_manual:
            self._standard_rows[-1]["routed_to"] = "Manual"

        # Per-state stats accumulation
        self._accumulate_state_stats(outcome, manual_reason, went_manual, went_add_license, went_ai_add_license)

    def _compute_manual_reason(self, outcome: RowOutcome) -> str | None:
        """Single point that decides whether a row goes to manual and why.
        Returns a human-readable reason string, or None (-> eligible for add_license).

        Priority order (first match wins):
          1. Captcha / board-skip blocked
          1.5. BACB captcha-blocked registry
          1.7. Low fuzzy/AI score (< 0.70) AND license ID numerics don't match
          2. AI fallback used — any layer (search or disambiguator) — Pass or Fail
          3. NPI substituted
          4. Rule-based Fail (no match found)
          5. Name ↔ license cross-validation mismatch (Pass overridden to Fail)
          6. Same expiry as input (no update needed)
        """
        _final_reason = (outcome.trace.final_reason or "").strip()

        # 1. Captcha / WAF block
        if _final_reason in _CAPTCHA_REASONS:
            if _final_reason == "board_skipped" and outcome.trace.skip_reason_text:
                return outcome.trace.skip_reason_text
            return self._CAPTCHA_MANUAL_REASONS.get(_final_reason, _final_reason)

        # 1.5. BACB board — captcha-blocked registry; skip automated verification
        if outcome.status != "Pass":
            _bacb_sources = {a.source_id for a in outcome.trace.attempts}
            if _bacb_sources and _bacb_sources.issubset(_BACB_SOURCE_IDS):
                return _BACB_CAPTCHA_REASON

        # 1.7. Low match score + no license match → Manual regardless of method.
        # Fires when: the match score (rule-based fuzzy OR AI confidence) is below 0.70
        # AND the license ID numerics didn't align. Both conditions together indicate the
        # match is too uncertain to be upload-ready — route to manual regardless of
        # whether the ladder or AI agent accepted it.
        # Skipped when the board record is expired (that surfaces as the primary reason).
        #
        # Exception — renewal pattern: perfect first AND last name (both == 1.0) with
        # no license match usually means the provider received a new license number on
        # renewal (EPDB still carries the old one).  Route to AIAddLicense so a reviewer
        # can confirm quickly rather than sending the whole row to Manual.
        _bd_check = outcome.chosen_breakdown
        _input_lic_check = (outcome.master_row.get("license_id", "") or "").strip()
        if (outcome.status == "Pass"
                and _bd_check is not None
                and _bd_check.total < 0.70
                and _bd_check.license_numerics < 1.0
                and _input_lic_check
                and not self._expired_after_fetch_reason(outcome)):
            if (_bd_check.first_name >= 1.0 and _bd_check.last_name >= 1.0
                    and not (outcome.ladder_result
                             and outcome.ladder_result.npi_substituted)):
                return "Name verified: license number changed on renewal — review required"
            return (
                f"Low match score ({round(_bd_check.total, 3)}) with no license ID match "
                f"— manual review required"
            )

        # 2. AI fallback used — layer 1 (search) or layer 2 (disambiguator)
        #    Applies regardless of Pass/Fail outcome; never goes to add_license.
        #    Expiry checks run first for Pass rows: an expired license must show
        #    'Provider fetch after Expiry' / same-expiry reason, not 'AI fallback passed'.
        if outcome.ai_result is not None:
            if outcome.status == "Pass":
                expired = self._expired_after_fetch_reason(outcome)
                if expired:
                    return expired
                same = self._same_expiry_check(outcome)
                if same:
                    return same
                return (
                    "AI fallback passed: manual review required to confirm "
                    "verification result before use"
                )
            else:
                # License mismatch detected by AI: route to Manual with board data preserved.
                # Must be checked BEFORE gate_passed — prov_type=0 can make gate_passed=False
                # even when the name matches well, which must not suppress the mismatch reason.
                if _final_reason == "AI found License ID mismatched":
                    return "AI found License ID mismatched"
                # If the AI picked a candidate whose name doesn't match the input
                # (gate_passed=False), don't route to AIAddLicense — send to Manual.
                _ai_bd_check = outcome.ai_result.chosen_breakdown
                if _ai_bd_check is not None and not _ai_bd_check.gate_passed:
                    _ai_fail_reason = (outcome.ai_result.reason or "no_candidates")
                    _ai_fail_reason = _ai_fail_reason.replace("—", "-").replace(";", ",")
                    return f"AI fallback failed - manual review required ({_ai_fail_reason})"
                # When the root cause was a license-found-but-name-mismatch, report
                # that directly so analysts see WHY rather than a generic AI fail.
                if (_final_reason == "name_mismatch"
                        and outcome.trace.license_attempts_returned_records()):
                    return "License matched but Name mismatched"
                _ai_fail_reason = (outcome.ai_result.reason or "no_candidates")
                _ai_fail_reason = _ai_fail_reason.replace("—", "-").replace(";", ",")
                return f"AI fallback failed - manual review required ({_ai_fail_reason})"

        # 3. Board record expired — checked before NPI/name-gate routing so that an
        # expired license always surfaces as "Provider fetch after Expiry" regardless
        # of name score or NPI substitution. (AI fallback already checks this at step 2;
        # this catches the rule-based, NPI-substituted, and name-gate pass paths.)
        if outcome.status == "Pass":
            expired_after_fetch = self._expired_after_fetch_reason(outcome)
            if expired_after_fetch:
                return expired_after_fetch

        # 4. NPI substituted (standard stays Pass; human confirms the NPI-derived match)
        if outcome.status == "Pass" and outcome.ladder_result and outcome.ladder_result.npi_substituted:
            return "NPI used to fetch - manual review required"

        # 4.5. Post-license name gate: max(nppes_score, epdb_score) < 0.70 — manual only
        _gate_reason = getattr(outcome.trace, "name_gate_reason", None)
        if outcome.status == "Pass" and _gate_reason == "name_gate_manual":
            # When the last name score is very low (< 0.40) the matched record is almost
            # certainly a different person, not a name-variant of the same person.
            # Return a reason NOT in _REASONS_FOR_AI_ADD_LICENSE so it routes to Manual.
            _gate_bd = outcome.chosen_breakdown
            _gate_input_last = (outcome.master_row.get("last_name", "") or "").strip()
            if _gate_bd is not None and _gate_input_last and _gate_bd.last_name < 0.40:
                return (
                    "Wrong provider matched: license matched but last name is completely different"
                )
            return (
                "Name mismatch after license match: "
                "EPDB and NPPES name scores both below 0.70 threshold"
            )

        # 5. Rule-based Fail
        if outcome.status == "Fail":
            _fail_reason_code = outcome.reason or "no_records"
            # When a license-mode rung found a record but the name doesn't match,
            # report the mismatch clearly instead of the raw reason code.
            if (_fail_reason_code == "name_mismatch"
                    and outcome.trace.license_attempts_returned_records()):
                return "License matched but Name mismatched"
            return _fail_reason_code

        # 5.5. KY temporary/internal license prefix — TC###, TP###, TSA### are client
        # tracking codes that never appear on the public board.  The board stores a
        # different permanent number (e.g. PA3822, 06248, SA464); name-only match is
        # the best we can do.  Surface as "Temporary License ID" so reviewers have the
        # board's actual license and expiry available.
        _input_lic = (outcome.master_row.get("license_id", "") or "").strip().upper()
        if _input_lic.startswith("TC") or (
            _input_lic.startswith("TP") and _input_lic[2:].isdigit()
        ) or (
            _input_lic.startswith("TSA") and _input_lic[3:].isdigit()
        ):
            return "Temporary License ID"

        # 5a. Name-only match: input has a license_id but the board record carries no
        #     license number at all (some boards don't surface it on the results table).
        #     Name + provider type were confirmed by the disambiguator; route to
        #     AIAddLicense so a reviewer can confirm identity before upload.
        if outcome.status == "Pass":
            _no_lic_bd = outcome.chosen_breakdown
            _no_lic_rec = outcome.chosen_record
            _no_lic_input = (outcome.master_row.get("license_id", "") or "").strip()
            _no_lic_board = (getattr(_no_lic_rec, "license_number", "") or "").strip() if _no_lic_rec else ""
            if (_no_lic_bd is not None
                    and _no_lic_bd.license_numerics < 1.0
                    and _no_lic_input
                    and not _no_lic_board):
                return "Name verified: board record does not expose license number"

        # 5b. Name-only high-confidence match: board returned a license number that
        #     doesn't match the input (different board numbering system). The
        #     disambiguator still accepted the name (gate_passed=True, score >=
        #     threshold). Route to AIAddLicense so a reviewer can confirm identity.
        if outcome.status == "Pass":
            _diff_bd = outcome.chosen_breakdown
            _diff_rec = outcome.chosen_record
            _diff_input = (outcome.master_row.get("license_id", "") or "").strip()
            _diff_board = (getattr(_diff_rec, "license_number", "") or "").strip() if _diff_rec else ""
            if (_diff_bd is not None
                    and _diff_bd.license_numerics < 1.0
                    and _diff_input
                    and _diff_board):
                return "Name match accepted: board uses different license numbering"

        # 5. Name ↔ license cross-validation (Pass rows only beyond this point)
        mismatch = (
            self._name_license_mismatch_reason(outcome)
            or self._license_name_mismatch_reason(outcome)
        )
        if mismatch:
            return mismatch

        # 5b. Name matched + numeric license aligned but strings not identical
        numeric_not_exact = self._numeric_not_exact_license_reason(outcome)
        if numeric_not_exact:
            return numeric_not_exact

        # 6. Expiry unchanged vs input
        same_expiry = self._same_expiry_check(outcome)
        if same_expiry:
            return same_expiry

        return None  # clean Pass — eligible for add_license

    # ----- Per-state stats -----

    def _accumulate_state_stats(
        self,
        outcome: RowOutcome,
        manual_reason: str | None,
        went_manual: bool,
        went_add_license: bool,
        went_ai_add_license: bool = False,
    ) -> None:
        state = (outcome.master_row.get("lic_state") or "UNKNOWN").upper()
        if state not in self._state_stats:
            self._state_stats[state] = _blank_state_stats()
        s = self._state_stats[state]
        s["total"] += 1

        _final_reason = (outcome.trace.final_reason or "").strip()
        _ai_used = outcome.ai_result is not None
        _npi_used = bool(outcome.ladder_result and outcome.ladder_result.npi_substituted)
        _is_bacb = (
            outcome.status != "Pass"
            and bool(outcome.trace.attempts)
            and all(a.source_id in _BACB_SOURCE_IDS for a in outcome.trace.attempts)
        )

        # AI counters
        if _ai_used:
            s["ai_used"] += 1
            if outcome.ai_result.outcome == "resolved":
                s["ai_resolved"] += 1
                s["pass_ai"] += 1
            else:
                s["ai_failed"] += 1
            # Accumulate token / cost for every row where AI ran
            s["ai_input_tokens"]  += getattr(outcome.ai_result, "input_tokens", 0) or 0
            s["ai_output_tokens"] += getattr(outcome.ai_result, "output_tokens", 0) or 0
            s["ai_usd_cost"]      += getattr(outcome.ai_result, "usd_cost", 0.0) or 0.0

        # NPI
        if _npi_used:
            s["npi_substituted"] += 1
            s["pass_npi"] += 1

        # Captcha (includes BACB-blocked rows)
        if _final_reason in _CAPTCHA_REASONS or _is_bacb:
            s["captcha"] += 1

        # Mismatch override
        if manual_reason in (
            "Name matched but License mismatched",
            "License matched but Name mismatched",
            "Numeric License ID matched",
        ):
            s["mismatch"] += 1

        # Same expiry
        if manual_reason in (
            "Check again later for updates, same as input",
            "Expired and same as input",
        ):
            s["same_expiry"] += 1

        # Expired after fetch
        if manual_reason == "Provider fetch after Expiry":
            s["expired_after_fetch"] += 1

        # No expiry
        if went_manual and manual_reason and "no_expiry_date" in manual_reason:
            s["no_expiry"] += 1

        # Pass / Fail in standard (after any retroactive overrides)
        std = self._standard_rows[-1] if self._standard_rows else {}
        if std.get("status") == "Pass":
            if not _ai_used and not _npi_used:
                s["pass_rule"] += 1
        else:
            if not _ai_used and _final_reason not in _CAPTCHA_REASONS and not _is_bacb:
                s["fail_rule"] += 1

        # Channel counters
        if went_manual:
            s["manual"] += 1
        if went_add_license:
            s["add_license"] += 1
        if went_ai_add_license:
            s["ai_add_license"] += 1

    # ----- Channel collectors -----

    def _collect_standard(self, o: RowOutcome) -> None:
        m = o.master_row
        rec = o.chosen_record
        bd = o.chosen_breakdown
        ev = ""
        for a in reversed(o.trace.attempts):
            if a.evidence_dir:
                ev = a.evidence_dir
                break

        # Match method
        _final_reason = (o.trace.final_reason or "").strip()
        _is_bacb = (
            o.status != "Pass"
            and bool(o.trace.attempts)
            and all(a.source_id in _BACB_SOURCE_IDS for a in o.trace.attempts)
        )
        if o.status != "Pass" and (_final_reason in _CAPTCHA_LABEL_REASONS or _is_bacb):
            match_method = "Captcha Based Board"
        elif o.status != "Pass":
            match_method = "none"
        elif o.ai_result and o.ai_result.outcome == "resolved":
            match_method = "ai_fuzzy"
        elif o.ladder_result and o.ladder_result.npi_substituted:
            match_method = "npi_substituted_exact"
        elif o.ladder_result and o.ladder_result.tiebreaker_used:
            match_method = "tiebreak_provider_type"
        else:
            match_method = "exact_license" if (
                bd and bd.license_numerics >= 1.0
            ) else "exact_name"

        _ml = _matched_license(rec, m, o.status)
        # For Fail rows with no chosen record, surface the best board candidate so
        # reviewers can see what name the board actually had (e.g. license matched
        # but name failed the gate). Pass rows always use the chosen record directly.
        _name_rec = rec if rec is not None else (
            _best_fail_candidate(o.trace) if o.status != "Pass" else None
        )
        # When AI picked a candidate but the license-mismatch override forced Fail,
        # surface the AI's candidate so the reviewer sees the board data (name +
        # board license number). _best_fail_candidate only sees ladder attempts and
        # would otherwise show an unrelated record from the license search.
        if (rec is None
                and _final_reason == "AI found License ID mismatched"
                and o.ai_result is not None
                and o.ai_result.chosen_candidate is not None):
            _name_rec = o.ai_result.chosen_candidate
            _ml = (getattr(o.ai_result.chosen_candidate, "license_number", "") or "").strip() or _ml
        row = {
            "master_row_id": o.master_row_id,
            "first_name": m.get("first_name", ""),
            "middle_name": m.get("middle_name", ""),
            "last_name": m.get("last_name", ""),
            "lic_state": m.get("lic_state", ""),
            "prov_type": m.get("prov_type", ""),
            "lic_type": m.get("lic_type", ""),
            "license_id": m.get("license_id", ""),
            "npi_no": m.get("npi_no", ""),
            "status": (
                "Skip" if (_final_reason in _CAPTCHA_REASONS or _is_bacb)
                else "Fail" if self._expired_after_fetch_reason(o)
                else o.trace.final_outcome if o.trace.final_outcome in ("N/A", "Skip")
                else o.status
            ),
            "license_expiry": _expiry_str(rec),
            "matched_license": _ml,
            "matched_first": _matched_name_part(_name_rec, m, o.status, 0),
            "matched_last":  _matched_name_part(_name_rec, m, o.status, 1),
            "board_name": _get_board_name(getattr(rec, "source_id", "") or "") if rec else "",
            "match_method": match_method,
            "fuzzy_score": (round(bd.total, 3) if bd else ""),
            "weight_profile": (bd.weight_profile if bd else ""),
            "tiebreaker_used": bool(o.ladder_result and o.ladder_result.tiebreaker_used)
                                or bool(o.ai_result and o.ai_result.outcome == "resolved"),
            "ai_fallback_used": o.ai_result is not None,
            "ai_outcome": (o.ai_result.outcome if o.ai_result else ""),
            "npi_substituted": bool(o.ladder_result and o.ladder_result.npi_substituted),
            "nppes_used": o.trace.nppes_used,
            "secondary_check_passed": bool(bd and bd.gate_passed),
            "provider_type_matched": bool(bd and bd.provider_type >= 1.0),
            "board_provider_type": (
                getattr(rec, "license_type", "") or getattr(rec, "profession_code", "") or ""
            ) if rec else "",
            "attempts_used": len(o.trace.attempts),
            "evidence_dir": ev,
            "trace_path": str(self.dirs["trace"] / f"{o.master_row_id}.json"),
            "routed_to": "",  # filled after routing decision in collect()
            "reason": o.reason,
            "fuzzy_breakdown": json.dumps(bd.to_dict()) if bd else "",
            "epdb_pin": str(m.get("epdb_pin", "") or ""),
            "epdb_name_score": (
                round(o.trace.epdb_name_score, 3)
                if o.trace.epdb_name_score is not None else ""
            ),
            "nppes_name_score": (
                round(o.trace.nppes_name_score, 3)
                if o.trace.nppes_name_score is not None else ""
            ),
            # AI telemetry — populated for every row; blank when AI did not run
            "ai_model": getattr(o.ai_result, "model", "") if o.ai_result else "",
            "ai_input_tokens": getattr(o.ai_result, "input_tokens", "") if o.ai_result else "",
            "ai_output_tokens": getattr(o.ai_result, "output_tokens", "") if o.ai_result else "",
            "ai_usd_cost": (
                round(o.ai_result.usd_cost, 6)
                if o.ai_result and getattr(o.ai_result, "usd_cost", None) is not None else ""
            ),
            "ai_confidence_score": (
                getattr(o.ai_result, "confidence_score", "")
                if o.ai_result and getattr(o.ai_result, "confidence_score", None) is not None else ""
            ),
        }
        self._standard_rows.append(row)

    def _collect_nppes(self, o: RowOutcome) -> None:
        n = o.nppes
        d = o.discrepancy
        if n is None:
            row = {
                "master_row_id": o.master_row_id,
                "npi_no": o.master_row.get("npi_no", ""),
                "fetch_status": "empty_input" if not o.master_row.get("npi_no") else "not_found",
            }
        else:
            primary_lic = n.primary_license
            row = {
                "master_row_id": o.master_row_id,
                "npi_no": n.npi,
                "nppes_first": n.first_name,
                "nppes_last": n.last_name,
                "nppes_middle": n.middle_name or "",
                "nppes_credential": n.credential or "",
                "nppes_primary_taxonomy": n.primary_taxonomy_desc,
                "nppes_primary_taxonomy_code": n.primary_taxonomy_code,
                "nppes_primary_license_no": primary_lic.get("number", "") if primary_lic else "",
                "nppes_primary_license_state": primary_lic.get("state", "") if primary_lic else "",
                "extra_license_count": max(0, len(n.license_numbers) - 1),
                "has_other_names": len(n.other_names) > 0,
                "fetch_status": n.fetch_status,
            }
            if d:
                df = d.differing_fields
                row["diff_first_name"] = _diff_cell(df.get("first_name"))
                row["diff_last_name"] = _diff_cell(df.get("last_name"))
                row["diff_license_number"] = _diff_cell(df.get("license_number"))
                row["other_name_used"] = d.other_name_used
            else:
                row.update({"diff_first_name": "", "diff_last_name": "",
                            "diff_license_number": "", "other_name_used": False})
        self._nppes_rows.append(row)

    def _collect_ai(self, o: RowOutcome) -> None:
        ai = o.ai_result
        assert ai is not None
        row = {
            "master_row_id": o.master_row_id,
            "lic_state": o.master_row.get("lic_state", ""),
            "prov_type": o.master_row.get("prov_type", ""),
            "first_name": o.master_row.get("first_name", ""),
            "last_name": o.master_row.get("last_name", ""),
            "license_id": o.master_row.get("license_id", ""),
            "outcome": ai.outcome,
            "reason": ai.reason,
            "turns_used": ai.turns_used,
            "tools_used": ",".join(ai.tools_used),
            "chosen_source_id": ai.chosen_source_id or "",
            "chosen_license": (getattr(ai.chosen_candidate, "license_number", "") or ""
                                if ai.chosen_candidate else ""),
            "drift_count": len(ai.drift_reports),
            # Token / cost telemetry
            "model": getattr(ai, "model", ""),
            "input_tokens": getattr(ai, "input_tokens", 0),
            "output_tokens": getattr(ai, "output_tokens", 0),
            "usd_cost": (
                round(getattr(ai, "usd_cost", 0.0), 6)
                if getattr(ai, "usd_cost", None) is not None else ""
            ),
            # Confidence: disambiguation ScoreBreakdown.total (0..1)
            "confidence_score": (
                getattr(ai, "confidence_score", None)
                if getattr(ai, "confidence_score", None) is not None else ""
            ),
            # Groundedness
            "groundedness_score": getattr(ai, "groundedness_score", 0),
            "hallucination_risk": getattr(ai, "hallucination_risk", "high"),
        }
        self._ai_rows.append(row)

    def _collect_manual(self, o: RowOutcome, failure_reason: str | None = None) -> None:
        reason = failure_reason if failure_reason is not None else o.reason
        row = {
            "master_row_id": o.master_row_id,
            "first_name": o.master_row.get("first_name", ""),
            "middle_name": o.master_row.get("middle_name", ""),
            "last_name": o.master_row.get("last_name", ""),
            "lic_state": o.master_row.get("lic_state", ""),
            "prov_type": o.master_row.get("prov_type", ""),
            "lic_type": o.master_row.get("lic_type", ""),
            "license_id": o.master_row.get("license_id", ""),
            "npi_no": o.master_row.get("npi_no", ""),
            "failure_reason": reason,
            "LicenseTermDate": _expiry_str(o.chosen_record),
            "attempts_used": len(o.trace.attempts),
            "trace_path": str(self.dirs["trace"] / f"{o.master_row_id}.json"),
            "nppes_used": o.trace.nppes_used,
        }
        self._manual_rows.append(row)
        self._collect_fallout(o, reason)

    def _collect_fallout(self, o: RowOutcome, failure_reason: str) -> None:
        """Client FallOut Report — one row per Manual-routed record."""
        m = o.master_row
        # Always use the input license_id — never the board-fetched license_number.
        lic_num = str(m.get("license_id", "") or "")
        self._fallout_rows.append({
            "FirstName":        str(m.get("first_name", "") or ""),
            "LastName":         str(m.get("last_name", "") or ""),
            "EPDB_PIN":         str(m.get("epdb_pin", "") or ""),
            "State":            str(m.get("lic_state", "") or ""),
            "LicenseNumber":    str(lic_num),
            "LicenseTermDate":  _expiry_str(o.chosen_record),
            "VerificationDate": _run_date_text(self.run_id),
            "CheckedBy":        f"Automation ({self.run_id})",
            "Outcome":          failure_reason,
        })

    def _collect_add_license(self, o: RowOutcome) -> bool:
        """AddLicense channel — one row per Pass result with a confirmed expiry date.

        Column rules (per AddLicense.xlsx template):
          EPDB                    -> Input  (EPDB PIN from master row)
          State                   -> Input  (License State from master row)
          MaintBy                 -> Input  (Maintained By from master row)
          LicenseNumber           -> Input  (verified license number from board)
          LicenseEffDate          -> Blanks (intentionally empty)
          LicenseTermDate         -> Updated Exp Date (expiry from board record)
          LicenseType             -> Operating (LIC_TYPE_NM from master row)
          OriginalLicenseDate     -> Blanks (intentionally empty)
          OverrideExistingLicense -> Yes
          EPDBDone                -> Blanks (filled manually post-upload)

        Returns True if the row was added, False if skipped (no expiry found -> manual).
        All values written as text strings; dates as MM/DD/YYYY.
        """
        m = o.master_row
        rec = o.chosen_record

        expiry = _expiry_text(rec)
        if not expiry:
            log.info(
                "[add_license] Skipping %s %s — no expiry date returned by board (-> manual review)",
                m.get("first_name", ""), m.get("last_name", ""),
            )
            return False

        # Always use the input license_id (matches what is already in EPDB).
        # The board may format its license differently; the input value is authoritative for upload.
        verified_license = (m.get("license_id", "") or "").strip()

        row = {
            "EPDB": str(m.get("epdb_pin", "") or ""),
            "State": str(m.get("lic_state", "") or ""),
            "MaintBy": str(m.get("maintained_by", "") or ""),
            "LicenseNumber": str(verified_license),
            "LicenseEffDate": "",
            "LicenseTermDate": expiry,          # MM/DD/YYYY text
            "LicenseType": str(m.get("lic_type", "") or ""),
            "OriginalLicenseDate": "",
            "OverrideExistingLicense": "Yes",
            "EPDBDone": "",
        }
        self._add_license_rows.append(row)
        return True

    def _collect_ai_add_license(self, o: RowOutcome, manual_reason: str) -> bool:
        """AI AddLicense channel — same base columns as AddLicense plus match context.

        Gathers rows that are "almost sure" matches: AI-resolved, partial name/license
        matches, NPI-based lookups.  These rows are NOT duplicated in Manual —
        AIAddLicense IS the manual-review file for these cases.

        Returns True if the row was added (expiry found), False if skipped.
        """
        m = o.master_row
        rec = o.chosen_record
        bd = o.chosen_breakdown

        expiry = _expiry_text(rec)
        if not expiry:
            return False

        # LicenseNumber uses the input license_id (matches what is already in EPDB).
        # board_lic_raw is the raw value from the board record — may differ in format
        # (e.g. input '029' vs board '29') and is used for the diff columns/reason label.
        verified_license = (m.get("license_id", "") or "").strip()
        board_lic_raw    = (getattr(rec, "license_number", "") or "").strip() if rec else ""

        input_first = (m.get("first_name", "") or "").strip()
        input_last  = (m.get("last_name",  "") or "").strip()
        input_name  = f"{input_first} {input_last}".strip()
        input_lic   = (m.get("license_id", "") or "").strip()

        board_first = (getattr(rec, "licensee_first_name", "") or "") if rec else ""
        board_last  = (getattr(rec, "licensee_last_name",  "") or "") if rec else ""
        board_name  = f"{board_first} {board_last}".strip() if (board_first or board_last) else (
            (getattr(rec, "full_name", "") or "") if rec else ""
        )

        row = {
            "EPDB":                   str(m.get("epdb_pin",      "") or ""),
            "State":                  str(m.get("lic_state",     "") or ""),
            "MaintBy":                str(m.get("maintained_by", "") or ""),
            "LicenseNumber":          str(verified_license),
            "LicenseEffDate":         "",
            "LicenseTermDate":        expiry,
            "LicenseType":            str(m.get("lic_type", "") or ""),
            "OriginalLicenseDate":    "",
            "OverrideExistingLicense": "Yes",
            "EPDBDone":               "",
            "VerificationReason":     _ai_add_license_reason_label(
                                          manual_reason, input_lic, board_lic_raw,
                                          input_name, board_name),
            "InputName":              input_name,
            "BoardMatchedName":       board_name,
            "InputLicense":           input_lic,
            "BoardMatchedLicense":    board_lic_raw or str(verified_license),
            "MatchScore":             (round(bd.total, 3) if bd else ""),
            "master_row_id":          o.master_row_id,
            "nppes_used":             o.trace.nppes_used,
        }
        self._ai_add_license_rows.append(row)
        return True

    # ----- Flush all channels to disk -----

    def flush(self) -> dict[str, Path]:
        self._reconcile_by_name()
        self._rebuild_add_license_after_reconciliation()
        paths: dict[str, Path] = {}
        if self._standard_rows:
            paths["standard_xlsx"] = self._write_standard_xlsx()
            paths["standard_csv"] = self._write_csv("standard", self._standard_rows)
        if self._nppes_rows:
            paths["nppes"] = self._write_csv("nppes", self._nppes_rows)
        if self._ai_rows:
            paths["ai_fallback"] = self._write_csv("ai_fallback", self._ai_rows)
        if self._manual_rows:
            paths["manual"] = self._write_csv("manual", self._manual_rows)
        if self._add_license_rows:
            paths["add_license"] = self._write_add_license_xlsx()
        if self._ai_add_license_rows:
            paths["ai_add_license"] = self._write_ai_add_license_xlsx()
        if self._fallout_rows:
            paths["fall_out"] = self._write_fallout_xlsx()
        if self._state_stats:
            paths["run_summary"] = self._write_run_summary()
        log.info("Output flushed: %s", {k: str(v) for k, v in paths.items()})
        return paths

    def _rebuild_add_license_after_reconciliation(self) -> None:
        """Route cross_row_name_match rows to AIAddLicense (never AddLicense).
        Cross-row reconciled rows are never "clean" — the input license ID did not
        match the board directly, so a human must confirm before upload.
        Mutual-exclusivity: any license already in manual or ai_add_license is skipped.
        """
        existing_licenses = {r["LicenseNumber"] for r in self._add_license_rows}
        existing_licenses |= {r.get("LicenseNumber", "") for r in self._ai_add_license_rows}
        # Licenses that are already going to manual must not also go to ai_add_license
        manual_licenses = {r.get("license_id", "") for r in self._manual_rows if r.get("license_id")}
        for row in self._standard_rows:
            if row.get("match_method") != "cross_row_name_match":
                continue
            if row.get("status") != "Pass":
                continue
            # Prefer input license_id for the upload-ready LicenseNumber column.
            lic = row.get("license_id") or row.get("matched_license", "")
            if lic in existing_licenses or lic in manual_licenses:
                continue

            expiry_iso = row.get("license_expiry", "")
            expiry = _iso_to_text(expiry_iso)

            if not expiry:
                # No expiry — cannot confirm the license is current; treat as Fail.
                _no_expiry_reason = (
                    "no_expiry_date: license verified via cross_row_name_match but "
                    "expiration date not available - manual review required to confirm LicenseTermDate"
                )
                row["status"] = "Fail"
                row["reason"] = _no_expiry_reason
                self._manual_rows.append({
                    "master_row_id": row.get("trace_path", ""),
                    "first_name": row.get("first_name", ""),
                    "middle_name": row.get("middle_name", ""),
                    "last_name": row.get("last_name", ""),
                    "lic_state": row.get("lic_state", ""),
                    "prov_type": row.get("prov_type", ""),
                    "lic_type": row.get("lic_type", ""),
                    "license_id": lic,
                    "npi_no": row.get("npi_no", ""),
                    "failure_reason": _no_expiry_reason,
                    "LicenseTermDate": "",
                    "attempts_used": row.get("attempts_used", ""),
                    "trace_path": row.get("trace_path", ""),
                })
                continue

            # Past-expiry check — do not add expired licenses to add_license
            _manual_stub = {
                "master_row_id": row.get("master_row_id", "") or row.get("trace_path", ""),
                "first_name": row.get("first_name", ""),
                "middle_name": row.get("middle_name", ""),
                "last_name": row.get("last_name", ""),
                "lic_state": row.get("lic_state", ""),
                "prov_type": row.get("prov_type", ""),
                "lic_type": row.get("lic_type", ""),
                "license_id": lic,
                "npi_no": row.get("npi_no", ""),
                "LicenseTermDate": expiry_iso,
                "attempts_used": row.get("attempts_used", ""),
                "trace_path": row.get("trace_path", ""),
            }
            try:
                if _date.fromisoformat(expiry_iso[:10]) < _date.today():
                    self._manual_rows.append({
                        **_manual_stub,
                        "failure_reason": "Provider fetch after Expiry",
                    })
                    continue
            except Exception:
                pass

            # EPDB is required — use the value propagated from master_row via standard row
            _recon_epdb = str(row.get("epdb_pin", "") or "").strip()
            if not _recon_epdb:
                self._manual_rows.append({
                    **_manual_stub,
                    "failure_reason": "EPDB is blanks",
                })
                continue

            input_name = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
            self._ai_add_license_rows.append({
                "EPDB":                    _recon_epdb,
                "State":                   str(row.get("lic_state", "") or ""),
                "MaintBy":                 "",
                "LicenseNumber":           str(lic),
                "LicenseEffDate":          "",
                "LicenseTermDate":         expiry,
                "LicenseType":             str(row.get("lic_type", "") or ""),
                "OriginalLicenseDate":     "",
                "OverrideExistingLicense": "Yes",
                "EPDBDone":                "",
                "VerificationReason": (
                    f"Cross-row name match: same provider verified via another row "
                    f"in this batch (matched license: {row.get('matched_license', lic)!r})"
                ),
                "InputName":           input_name,
                "BoardMatchedName": (
                    f"{row.get('matched_first', '')} {row.get('matched_last', '')}".strip()
                ),
                "InputLicense":        str(row.get("license_id", "") or ""),
                "BoardMatchedLicense": str(row.get("matched_license", "") or lic),
                "MatchScore":          str(row.get("fuzzy_score", "") or ""),
                "master_row_id":       str(row.get("master_row_id", "") or ""),
                "nppes_used":          row.get("nppes_used", False),
            })
            row["routed_to"] = "AIAddLicense"
            existing_licenses.add(lic)

    def _reconcile_by_name(self) -> None:
        """Post-run pass: for every non-Pass row, check if another row in this
        run already passed for the same (first, last, state, prov_type). If so,
        copy the matched record fields from the passing row and mark as Pass with
        match_method='cross_row_name_match'. Handles the common case where the
        same provider appears twice with different license-ID formats (e.g. '46101'
        and '5346101') — the license fetch for one format succeeds; the other
        resolves here via name match rather than hitting AI/manual.
        """
        def _norm(s: str) -> str:
            return (s or "").strip().upper()

        # Build lookup of passing rows by (first, last, state, prov_type).
        # When multiple passing rows share the same name key (genuinely different
        # providers with the same name), skip reconciliation for that key to avoid
        # cross-contamination.
        pass_lookup: dict[tuple, dict] = {}
        ambiguous_keys: set[tuple] = set()
        for row in self._standard_rows:
            if row.get("status") != "Pass":
                continue
            key = (
                _norm(row.get("first_name", "")),
                _norm(row.get("last_name", "")),
                _norm(row.get("lic_state", "")),
                _norm(row.get("prov_type", "")),
            )
            if not key[0] or not key[1]:
                continue  # no name -> cannot reconcile
            if key in pass_lookup:
                ambiguous_keys.add(key)
            else:
                pass_lookup[key] = row

        reconciled = 0
        for row in self._standard_rows:
            if row.get("status") == "Pass":
                continue
            key = (
                _norm(row.get("first_name", "")),
                _norm(row.get("last_name", "")),
                _norm(row.get("lic_state", "")),
                _norm(row.get("prov_type", "")),
            )
            if not key[0] or not key[1]:
                continue
            if key in ambiguous_keys or key not in pass_lookup:
                continue
            # Do not override rows that failed for a definitive business reason.
            # These rows were reached on the board and rejected deliberately.
            _BLOCK_RECONCILE = {
                "Provider fetch after Expiry",
            }
            if row.get("reason") in _BLOCK_RECONCILE:
                continue
            src = pass_lookup[key]
            # Guard: license IDs must be compatible before promoting via name-only.
            # Two-stage comparison:
            #   1. Alphanumeric-normalised (strip hyphens/spaces, lowercase) — handles
            #      "CH-1445" == "CH1445" and keeps meaningful letter suffixes like "-C".
            #   2. Digit-only fallback — handles prefix differences ("CH1445" vs "1445")
            #      and leading-zero variants; blocks "CH1445" vs "CH10445" (1445≠10445).
            # If both sides have a license and neither stage passes, the licenses are
            # unrelated — leave the row as Fail/Manual.
            _input_lic     = (row.get("license_id", "") or "").strip()
            _src_board_lic = (src.get("matched_license", "") or src.get("license_id", "")).strip()
            if _input_lic and _src_board_lic:
                import re as _re
                _alnum = lambda s: _re.sub(r"[-\s]", "", s.lower())
                if _alnum(_input_lic) != _alnum(_src_board_lic):
                    # Block digit-only fallback when both licenses end with a
                    # different letter — the suffix is meaningful, not a format variant.
                    _trail = lambda s: _re.search(r"[A-Za-z]$", s.strip())
                    _tA = _trail(_input_lic)
                    _tB = _trail(_src_board_lic)
                    if _tA and _tB and _input_lic.strip()[-1].upper() != _src_board_lic.strip()[-1].upper():
                        continue
                    if not _lic_num_match(_input_lic, _src_board_lic):
                        continue
            row["status"] = "Pass"
            row["match_method"] = "cross_row_name_match"
            row["matched_license"] = src.get("matched_license", "") or src.get("license_id", "")
            row["matched_first"] = src.get("matched_first", "")
            row["matched_last"] = src.get("matched_last", "")
            row["license_expiry"] = src.get("license_expiry", "")
            row["fuzzy_score"] = src.get("fuzzy_score", "")
            row["weight_profile"] = src.get("weight_profile", "")
            row["source_run_row"] = src.get("license_id", "")  # for audit
            reconciled += 1

        if reconciled:
            log.info("Post-run name reconciliation: resolved %d row(s) via cross_row_name_match",
                     reconciled)
            # Remove reconciled rows from manual channel (they are now Pass).
            # Also patch per-state stats so RunSummary reflects final channel routing
            # (manual→add_license) rather than the pre-reconciliation state.
            reconciled_ids = {
                r["license_id"]
                for r in self._standard_rows
                if r.get("match_method") == "cross_row_name_match"
            }
            for mr in self._manual_rows:
                if mr.get("license_id") in reconciled_ids:
                    state = (mr.get("lic_state") or "UNKNOWN").upper()
                    if state in self._state_stats:
                        self._state_stats[state]["manual"] = max(
                            0, self._state_stats[state]["manual"] - 1
                        )
                        self._state_stats[state]["ai_add_license"] += 1
            self._manual_rows = [
                r for r in self._manual_rows
                if r.get("license_id") not in reconciled_ids
            ]

    def _write_fallout_xlsx(self) -> Path:
        """Write Client FallOut Report — all Manual-routed rows in a client-readable Excel."""
        dt = cfg.date_time_from_run_id(self.run_id)
        out = self.dirs["fall_out"] / f"ClientFallOut_{dt}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ClientFallOut"
        _FO_COLS = [
            "FirstName", "LastName", "EPDB_PIN", "State", "LicenseNumber",
            "VerificationDate", "CheckedBy", "Outcome",
        ]
        ws.append(_FO_COLS)
        hdr_fill = PatternFill("solid", fgColor="F4CCCC")  # light red
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = hdr_fill
        for r in self._fallout_rows:
            row_num = ws.max_row + 1
            for col_idx, h in enumerate(_FO_COLS, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=str(r.get(h, "") or ""))
                cell.number_format = "@"
        wb.save(str(out))
        log.info("Client FallOut report written: %s (%d row(s))", out, len(self._fallout_rows))
        return out

    def _write_run_summary(self) -> Path:
        """Write one row per state to Output/{YYYYMM}/{run_id}/RunSummary/RunSummary_{dt}.csv."""
        dt = cfg.date_time_from_run_id(self.run_id)
        out = self.dirs["run_summary"] / f"RunSummary_{dt}.csv"
        _COLS = [
            "run_id", "state",
            "total", "pass_rule", "pass_ai", "pass_npi",
            "fail_rule", "fail_ai", "captcha",
            "mismatch", "same_expiry", "expired_after_fetch", "no_expiry",
            "manual", "add_license", "ai_add_license",
            "ai_used", "ai_resolved", "ai_failed", "npi_substituted",
            "ai_input_tokens", "ai_output_tokens", "ai_usd_cost",
        ]
        rows = []
        for state in sorted(self._state_stats):
            s = self._state_stats[state]
            row = {"run_id": self.run_id, "state": state}
            row.update({k: s.get(k, 0) for k in _COLS if k not in ("run_id", "state")})
            # Round USD cost to 6 decimal places for readability
            row["ai_usd_cost"] = round(row["ai_usd_cost"], 6)
            rows.append(row)

        # Append a TOTAL row across all states
        if len(rows) > 1:
            total_row = {"run_id": self.run_id, "state": "TOTAL"}
            for k in _COLS:
                if k in ("run_id", "state"):
                    continue
                total_row[k] = sum(r.get(k, 0) for r in rows)
            total_row["ai_usd_cost"] = round(total_row["ai_usd_cost"], 6)
            rows.append(total_row)

        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=_COLS)
            w.writeheader()
            w.writerows(rows)

        log.info("Run summary written: %s (%d state(s))", out, len(self._state_stats))
        return out

    def _write_csv(self, channel: str, rows: list[dict]) -> Path:
        if not rows:
            return Path()
        dt = cfg.date_time_from_run_id(self.run_id)
        folder_name = self.dirs[channel].name  # e.g. "NPPES", "AIFallback"
        out = self.dirs[channel] / f"{folder_name}_{dt}.csv"
        # Union of all keys across rows
        keys: list[str] = []
        seen: set[str] = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            for r in rows:
                w.writerow(r)
        return out

    def _write_add_license_xlsx(self) -> Path:
        dt = cfg.date_time_from_run_id(self.run_id)
        out = self.dirs["add_license"] / f"AddLicense_{dt}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AddLicense"
        _AL_COLS = [
            "EPDB", "State", "MaintBy", "LicenseNumber",
            "LicenseEffDate", "LicenseTermDate", "LicenseType",
            "OriginalLicenseDate", "OverrideExistingLicense", "EPDBDone",
        ]
        # Header row (bold, no text format needed for headers)
        ws.append(_AL_COLS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Data rows — every cell forced to TEXT format so Excel never
        # auto-converts dates, license numbers, or leading zeros.
        for r in self._add_license_rows:
            row_num = ws.max_row + 1
            for col_idx, h in enumerate(_AL_COLS, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=str(r.get(h, "") or ""))
                cell.number_format = "@"   # Excel TEXT format
        wb.save(str(out))
        return out

    def _write_ai_add_license_xlsx(self) -> Path:
        """Write AI_ADD_LICENSE Excel — same base columns as AddLicense plus match context."""
        dt = cfg.date_time_from_run_id(self.run_id)
        out = self.dirs["ai_add_license"] / f"AIAddLicense_{dt}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AIAddLicense"
        _AAL_COLS = [
            "EPDB", "State", "MaintBy", "LicenseNumber",
            "LicenseEffDate", "LicenseTermDate", "LicenseType",
            "OriginalLicenseDate", "OverrideExistingLicense", "EPDBDone",
            # Match context columns (informational — do not upload these)
            "VerificationReason", "InputName", "BoardMatchedName",
            "InputLicense", "BoardMatchedLicense", "MatchScore", "master_row_id",
            "nppes_used",
        ]
        # Header row
        ws.append(_AAL_COLS)
        hdr_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = hdr_fill
        # Mark the context-only columns with a lighter fill so reviewers know to stop at EPDBDone
        context_start = _AAL_COLS.index("VerificationReason") + 1  # 1-based col index
        ctx_fill = PatternFill("solid", fgColor="FFF2CC")  # pale yellow
        for col_idx in range(context_start, len(_AAL_COLS) + 1):
            ws.cell(row=1, column=col_idx).fill = ctx_fill
        # Data rows — every cell TEXT format
        for r in self._ai_add_license_rows:
            row_num = ws.max_row + 1
            for col_idx, h in enumerate(_AAL_COLS, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=str(r.get(h, "") or ""))
                cell.number_format = "@"
        wb.save(str(out))
        return out

    def _write_standard_xlsx(self) -> Path:
        dt = cfg.date_time_from_run_id(self.run_id)
        out = self.dirs["standard"] / f"Standard_{dt}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PSV Results"
        if not self._standard_rows:
            wb.save(str(out))
            return out
        headers = list(self._standard_rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        for r in self._standard_rows:
            ws.append([r.get(h, "") for h in headers])
            fill = green if r.get("status") == "Pass" else red
            row_idx = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill
        wb.save(str(out))
        return out


# ----- Helpers ------------------------------------------------------------

def _matched_license(rec: Optional[Any], master_row: dict, status: str) -> str:
    """Board license number; falls back to master-row license_id on Pass when blank."""
    val = (getattr(rec, "license_number", "") or "").strip() if rec else ""
    if not val and status == "Pass":
        val = (master_row.get("license_id", "") or "").strip()
    return val



_LIC_MODES_FOR_FAIL_DISPLAY: frozenset = frozenset({
    "license_number", "license_number_exact", "license_numeric_only",
    "license_formatted", "license_first_last", "license_and_last",
    "license_and_first",
})


def _best_fail_candidate(trace: Any) -> Optional[Any]:
    """Best board record from a failed attempt, for display in Fail rows.

    When the ladder fails but attempts returned candidates that didn't pass the
    gate (name/license mismatch), we surface what the board actually had so
    reviewers see the discrepancy in matched_first/matched_last.
    Prefers the most recent license-mode attempt (more specific) over name-only.
    """
    for attempt in reversed(trace.attempts):
        if attempt.candidates and attempt.mode in _LIC_MODES_FOR_FAIL_DISPLAY:
            return attempt.candidates[0]
    for attempt in reversed(trace.attempts):
        if attempt.candidates:
            return attempt.candidates[0]
    return None


def _matched_name_part(rec: Optional[Any], master_row: dict, status: str, idx: int) -> str:
    """Board first (idx=0) or last (idx=1) name; validates and falls back on Pass when blank/garbage."""
    raw_val = OutputEmitter._resolve_board_name_parts(rec, master_row.get("last_name", ""))[idx]
    cleaned = _clean_matched_name(raw_val)
    # Reject pure-numeric tokens (e.g. '1' from garbled detail-page extraction).
    if cleaned and cleaned.strip().isdigit():
        cleaned = ""
    if not cleaned and status == "Pass":
        fallback_key = "first_name" if idx == 0 else "last_name"
        cleaned = (master_row.get(fallback_key, "") or "").strip()
    return cleaned


def _expiry_str(rec: Optional[Any]) -> str:
    """ISO date string — used for standard output channel."""
    if rec is None:
        return ""
    d = getattr(rec, "expiration_date", None)
    if d is None:
        return ""
    try:
        return d.isoformat()
    except Exception:
        return str(d)


def _expiry_text(rec: Optional[Any]) -> str:
    """MM/DD/YYYY text string for AddLicense output. Returns '' if no expiry."""
    if rec is None:
        return ""
    d = getattr(rec, "expiration_date", None)
    if d is None:
        return ""
    try:
        if hasattr(d, "strftime"):
            return d.strftime("%m/%d/%Y")
        if isinstance(d, str) and d:
            parsed = _date.fromisoformat(d[:10])
            return parsed.strftime("%m/%d/%Y")
        return str(d)
    except Exception:
        return str(d) if d else ""


def _run_date_text(run_id: str) -> str:
    """Extract run date from run_id (YYYYMMDD...) → MM/DD/YYYY text."""
    try:
        if len(run_id) >= 8 and run_id[:8].isdigit():
            d = _date(int(run_id[:4]), int(run_id[4:6]), int(run_id[6:8]))
            return d.strftime("%m/%d/%Y")
    except Exception:
        pass
    return ""


def _iso_to_text(iso: str) -> str:
    """Convert an ISO date string (YYYY-MM-DD) to MM/DD/YYYY text. Returns '' if unparseable."""
    if not iso:
        return ""
    try:
        parsed = _date.fromisoformat(iso[:10])
        return parsed.strftime("%m/%d/%Y")
    except Exception:
        return iso  # pass through as-is if already formatted or unparseable


def _ai_add_license_reason_label(
    manual_reason: str,
    input_lic: str,
    board_lic: str,
    input_name: str,
    board_name: str,
) -> str:
    """Return a short, human-readable VerificationReason string for AI_ADD_LICENSE rows."""
    if manual_reason.startswith("AI fallback passed"):
        return "AI agent resolved match - verify before upload"
    if manual_reason == "Numeric License ID matched":
        return (
            f"Numeric license match: format differs "
            f"(input: {input_lic!r}: board: {board_lic!r})"
        )
    if manual_reason == "License matched but Name mismatched":
        return (
            f"License exact match: name on board differs "
            f"(input: {input_name!r}: board: {board_name!r})"
        )
    if manual_reason == "Name matched but License mismatched":
        return (
            f"Name exact match: license on board differs "
            f"(input: {input_lic!r}: board: {board_lic!r})"
        )
    if manual_reason == "Name verified: board record does not expose license number":
        return (
            f"Name-only verified (board has no license field): "
            f"confirm {input_name!r} is the correct provider before upload"
        )
    if manual_reason == "Name match accepted: board uses different license numbering":
        return (
            f"Name match: board license {board_lic!r} differs from input {input_lic!r}. "
            f"Confirm {input_name!r} is the correct provider before upload"
        )
    if manual_reason.startswith("NPI used to fetch"):
        return "NPI-verified match - confirm license before upload"
    return manual_reason


def _diff_cell(pair: Optional[tuple[str, str]]) -> str:
    if not pair:
        return ""
    return f"{pair[0]} | {pair[1]}"
