"""PSV batch tester — reads input Excel, routes each row to matching board
configs, calls verify_license(), writes Pass/Fail + reason incrementally.
Processes rows in batches (default 10) so results are saved after each batch.

Search strategy per row:
  1. Try license_number mode first (fast for active/known licenses).
  2. If no records returned, fall back to last_name mode (catches expired licenses
     still in board DB, and handles format mismatches in license IDs).

Browser reuse:
  For browser-based boards (classic_html_form), one Playwright browser is launched
  per board config at startup and reused across all searches. This avoids the 30-60s
  per-call Chromium launch overhead on Windows. Context+page are recreated per search
  (~0.5s) instead of restarting the whole browser (~35s).

  For non-browser boards (csv_bulk, socrata_api, etc.), verify_license() is used as
  normal — these open no browser.
"""
from __future__ import annotations

import argparse
import asyncio
import copy
import csv
import logging
import re
import sys
import time
import types
from datetime import datetime as _dt
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from playwright.async_api import async_playwright, Browser

sys.path.insert(0, str(Path(__file__).parent))

from engine.browser import _REAL_UA, _STEALTH, _STEALTH_ARGS
from engine.evidence import capture_evidence
from engine.extractor import extract_results_table, extract_detail, extract_th_td_multi
from engine.post_processors import apply_field_map
from engine.models import BoardUnavailableError, LicenseStatus, SearchQuery
from engine.output import map_to_license_record
from engine.proxy import get_proxy_config
from engine.validate import load_config
from engine.navigator import navigate_to_search, fill_search_form
from archetypes._shared import _wait_for_detail_content, _navigate_back, _try_out_of_state_tab, _scrape_pdf_detail
from archetypes.browser_form import _fetch_detail_via_api
from run import verify_license

log = logging.getLogger("psv_test")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
)

# PSV Tab column indices (0-based) — fixed layout
C_FIRST_NAME = 0
C_MIDDLE_NAME = 1
C_LAST_NAME = 2
C_EPDB_PIN = 3      # EPDB PIN — used by AddLicense output channel
C_PROV_TYPE = 4
C_MAINTAINED_BY = 7 # Maintained By — used by AddLicense output channel
C_LIC_STATE = 9
C_LIC_TYPE = 10
C_LIC_ID = 11
C_LIC_EXPIRY = 12  # LIC_EXPRTN_DT — input expiry date used for same-expiry comparison
C_SVC_LOC_STATE = 15  # Service Location State — license state must appear here

# NPI_NO is looked up DYNAMICALLY by header name — Input.xlsx may not always
# include it. If a header cell matches one of these names, that column index
# is used; otherwise npi_no stays empty and NPPES enrichment is skipped per row.
_NPI_HEADER_ALIASES = ("NPI_NO", "NPI", "NPI ID", "NPI_ID", "NPI Number")

# Browser-based archetypes (share one Playwright browser per board)
_BROWSER_ARCHETYPES = {"classic_html_form", "aspnet_webforms", "angular_spa", "react_spa", "ag_grid_spa"}

# Per-(state, prov_type) combos where the state does not license that provider type.
# The board is still routed and visited, but the outcome is forced to N/A after
# the scrape completes — the message below is written to the output regardless of
# what the board returns.
NA_PROV_TYPES: dict[tuple[str, str], str] = {
    ("NJ", "DT"):  "NJ does not license Dietitians (DT) — Professional License step N/A",
    ("NJ", "NUT"): "NJ does not license Nutritionists (NUT) — Professional License step N/A",
    ("CT", "DT"):  "The state of Connecticut does not require DT/NUT to hold state licensure",
    ("CT", "NUT"): "The state of Connecticut does not require DT/NUT to hold state licensure",
    ("CO", "DT"):  "The state of Colorado does not license Dietitians — Professional License step N/A",
    ("CO", "NUT"): "The state of Colorado does not license Nutritionists — Professional License step N/A",
    ("CO", "ABA"): "The state of Colorado does not issue state ABA licenses — Professional License step N/A",
    ("MI", "DT"):  "Michigan does not license Dietitians (DT) at this time — no state primary source (per LARA, Mary Hess 517-335-4084). Professional License step N/A",
    ("MI", "NUT"): "Michigan does not license Nutritionists (NUT) at this time — no state primary source (per LARA, Mary Hess 517-335-4084). Professional License step N/A",
    ("NY", "DT"):  "The state of New York does not currently license DT/NUT, but they do issue certification — Professional License step N/A",
    ("NY", "NUT"): "The state of New York does not currently license DT/NUT, but they do issue certification — Professional License step N/A",
}

# (state, prov_type) combos to Skip immediately — the state does not license the
# profession, so no primary source exists and visiting a board is pointless. The
# row is written as Skip with the NA_PROV_TYPES reason above (no browser launched).
SKIP_UNLICENSED_PROV_TYPES: set[tuple[str, str]] = {
    ("CO", "DT"), ("CO", "NUT"), ("CO", "ABA"),
    ("MI", "DT"), ("MI", "NUT"),
    ("NY", "DT"), ("NY", "NUT"),
}

# Per-(state, prov_type) combos where the board site blocks automated access.
# Rows matching these are written as Fail immediately without launching a browser.
CAPTCHA_PROV_TYPES: dict[tuple[str, str], str] = {
    ("KY", "NP"):  "KY Board of Nursing requires CAPTCHA — automated access blocked",
    ("KY", "PN"):  "KY Board of Nursing requires CAPTCHA — automated access blocked",
    ("KY", "RNA"): "KY Board of Nursing requires CAPTCHA — automated access blocked",
    # McAfee Web Gateway blocks both proxy and direct access (502 both routes 2026-06-29)
    ("NV", "OT"):  "NV OT board (occupationaltherapy.nv.gov) blocked by McAfee Web Gateway",
    # Nevada State Board of Nursing (nevadanursingboard.org) — connection forcibly closed
    # by McAfee Web Gateway from corporate network (WinError 10054, 2026-06-29).
    ("NV", "NP"):  "NV Board of Nursing (nevadanursingboard.org) blocked by McAfee Web Gateway",
    ("NV", "PN"):  "NV Board of Nursing (nevadanursingboard.org) blocked by McAfee Web Gateway",
    ("NV", "RNA"): "NV Board of Nursing (nevadanursingboard.org) blocked by McAfee Web Gateway",
    # Nevada Board of Examiners for Social Workers — DNS resolution fails from corporate
    # network (getaddrinfo failed on all known nv.gov/swbn variants, 2026-06-29).
    ("NV", "SW"):  "NV Social Work board (swbn.nv.gov) — DNS blocked at corporate network layer",
    # KY Board of Pharmacy (kybopp.aithent.com) — McAfee Web Gateway returns 502 from
    # corporate network; pharmacy.ky.gov also returns 403 (confirmed 2026-06-30).
    # KY PH removed 2026-07-01: PH now routes to KY_MEDBOARD first (no CAPTCHA).
    ("KY", "PM"):  "KY Board of Pharmacy (kybopp.aithent.com) blocked by McAfee Web Gateway",
    # KY Board of Dentistry: kybde.ky.gov was McAfee-blocked; replaced 2026-07-04 by
    # kbd.portalus.thentiacloud.net (KY_DENTAL) which IS accessible — DN entry removed.
    # KY Board of Social Work (kscsw.org) — CAPTCHA-protected, automated access blocked.
    ("KY", "SW"):  "KY Board of Social Work (kscsw.org) — CAPTCHA-protected, automated access blocked",
    # KY Board of Physical Therapy (secure.kentucky.gov/formservices/PT) — CAPTCHA-protected;
    # PT licenses are NOT on KY_MULTIBOARD despite partial name hits there.
    ("KY", "PT"):  "KY Board of Physical Therapy (secure.kentucky.gov/formservices/PT) — CAPTCHA-protected, automated access blocked",
    # NV Board of Nursing (nvbn.boardsofnursing.org) — same block as NV/PN/RNA/NP.
    # MW (midwife/CNM) licenses carry RN-prefix numbers issued by the nursing board.
    ("NV", "MW"):  "NV Board of Nursing (nvbn.boardsofnursing.org) blocked — connection times out from corporate network",
    # MD Board of Occupational Therapy Practice (mdbnc.health.maryland.gov/OTVerification) —
    # CAPTCHA-protected, automated access blocked from corporate network.
    ("MD", "OT"):  "MD Board of Occupational Therapy Practice (mdbnc.health.maryland.gov) — CAPTCHA-protected, automated access blocked",
    # MD ABA (Behavior Analysts) — no automated-access board in inventory; site requires CAPTCHA.
    ("MD", "ABA"): "MD Behavior Analyst board — CAPTCHA-protected, automated access blocked",
    # WY Board of Nursing (nursing.state.wy.us) — CAPTCHA-protected, no automated access.
    # NP/PN/RNA must not fall through to WY_PHYSICIAN (physician board).
    ("WY", "NP"):  "WY Board of Nursing (nursing.state.wy.us) — CAPTCHA-protected, automated access blocked",
    ("WY", "PN"):  "WY Board of Nursing (nursing.state.wy.us) — CAPTCHA-protected, automated access blocked",
    ("WY", "RNA"): "WY Board of Nursing (nursing.state.wy.us) — CAPTCHA-protected, automated access blocked",
    # NC Marriage and Family Therapy Board (ncbmft.org) — CAPTCHA-protected.
    # Previously mis-routed to NC_MENTAL_HEALTH (ncblcmhc.org) which is the wrong board.
    # Correct URL: https://www.ncbmft.org/licensure/verify-a-licensee — CAPTCHA-blocked.
    ("NC", "MT"):   "NC Board of Marriage and Family Therapy (ncbmft.org) — CAPTCHA-protected, automated access blocked",
    # NC Board of Nursing (ncbon.com) — CAPTCHA-protected. Covers RN, LPN, CRNA, NP, NPB, MW.
    ("NC", "RN"):   "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "LPN"):  "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "PN"):   "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "CRNA"): "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "RNA"):  "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "NP"):   "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "NPB"):  "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    ("NC", "MW"):   "NC Board of Nursing (ncbon.com) — CAPTCHA-protected, automated access blocked",
    # NC ABA board — CAPTCHA-protected, no automated access.
    ("NC", "ABA"):  "NC Applied Behavior Analyst board — CAPTCHA-protected, automated access blocked",
    # NC Board of Pharmacy (ncbop.org) — CAPTCHA-protected. Covers PM.
    ("NC", "PM"):   "NC Board of Pharmacy (ncbop.org) — CAPTCHA-protected, automated access blocked",
    # NC Medical Board (ncmedboard.org) — CAPTCHA-protected. Covers MD, DO, PA, PAS, PAH, PAB, PH.
    ("NC", "MD"):   "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    ("NC", "DO"):   "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    ("NC", "PA"):   "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    ("NC", "PAS"):  "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    ("NC", "PAH"):  "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    ("NC", "PAB"):  "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    ("NC", "PH"):   "NC Medical Board (ncmedboard.org) — CAPTCHA-protected, automated access blocked",
    # NC Social Work Certification and Licensure Board (ncswboard.org) — CAPTCHA-protected, no routing configured.
    ("NC", "SW"):   "NC Social Work Certification and Licensure Board (ncswboard.org) — CAPTCHA-protected, automated access blocked",
    # NC Art Therapy — verified manually by emailing the board contact.
    ("NC", "AP"):   "License will be verified by emailing to pat@smvt.com",
    # AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 explicit on
    # every search; blocks all headless/automated access (site key 6LdG0VIUA...).
    ("AR", "RN"):   "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "LPN"):  "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "APRN"): "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "CRNA"): "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "NP"):   "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "NPB"):  "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "PN"):   "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
    ("AR", "RNA"):  "AR State Board of Nursing (arsbn.boardsofnursing.org) — reCAPTCHA v2 blocks automated access",
}

# Maps (board_source_id, license_prefix_uppercase) → skip_reason.
# When a license starts with the given prefix, that board is bypassed entirely.
BOARD_LICENSE_PREFIX_SKIP: dict[tuple[str, str], str] = {
    # TSA = Temporary Surgical Assistant — different credential class from permanent KCSA.
    ("KY_SA", "TSA"): (
        "TSA (Temporary Surgical Assistant) is a different credential class — "
        "KY_SA tracks permanent KCSA credentials only"
    ),
    # PSYPACT is a multistate compact credential, not a state-issued license.
    ("NV_BOP", "PSYPACT"): (
        "PSYPACT is a multistate compact license — not issued by NV state board (NV_BOP not applicable)"
    ),
    ("KS_BSRB", "PSYPACT"): (
        "PSYPACT is a multistate compact license — not issued by KS state board (KS_BSRB not applicable)"
    ),
}

# Project root — used for PSYPACT evidence paths and psypact_scraper import
_PSV_DEV = Path(__file__).parents[3]

# Routing table: (state_abbr, psv_prov_type) -> [source_id, ...]
_ROUTING: dict[tuple[str, str], list[str]] = {}
_ROUTING_CSV = Path(__file__).parent / "board_routing_master.csv"


def _parse_psypact_expiry(date_str: str):
    """Parse PSYPACT expiry string ('April 16, 2027') to datetime.date."""
    if not date_str:
        return None
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%b. %d, %Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return _dt.strptime(date_str.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return date_str  # fall back to raw string


async def _verify_psypact_row(row: dict, trace, run_id: str = "") -> "LadderResult":
    """Verify a PSYPACT E.Passport license via directory.psypact.gov.

    Handles both common license_id formats:
      - "PSYPACT18696"  →  mobility number "18696"
      - "APIT-18240"    →  mobility number "18240"  (Authority to Practice
                            Interjurisdictional Telepsychology)
    """
    from orchestrator.ladder import LadderResult  # noqa: PLC0415
    from orchestrator.disambiguator import ScoreBreakdown  # noqa: PLC0415
    from orchestrator.config import yyyy_mm_from_run_id  # noqa: PLC0415

    first = (row.get("first_name") or "").strip()
    middle = (row.get("middle_name") or "").strip()
    last = (row.get("last_name") or "").strip()
    license_id = (row.get("license_id") or "").strip()

    # Strip known PSYPACT prefixes before extracting the mobility number.
    # "PSYPACT18696" → "18696", "APIT-18240" → "18240"
    _stripped = re.sub(r"^(?:PSYPACT|APIT[-\s]*)", "", license_id.upper())
    expected_num = re.sub(r"[^\d]", "", _stripped)
    ym = yyyy_mm_from_run_id(run_id) if run_id else ""
    evidence_dir = str(_PSV_DEV / "Evidence" / ym / run_id)

    if str(_PSV_DEV) not in sys.path:
        sys.path.insert(0, str(_PSV_DEV))

    try:
        from psypact_scraper import run_scraper as _run_psypact  # noqa: PLC0415

        results = await _run_psypact(
            first, middle, last, output_dir=evidence_dir, headless=True
        )

        for result in results:
            ld = result.get("license_data")
            if not ld:
                continue
            mobility = (ld.get("mobility_number") or "").strip()
            if not mobility or mobility != expected_num:
                continue

            expiry = _parse_psypact_expiry(ld.get("expiration_date", ""))
            rec = types.SimpleNamespace(
                license_number=license_id,
                expiration_date=expiry,
                licensee_first_name=first,
                licensee_last_name=last,
                source_id="PSYPACT_DIRECTORY",
            )
            bd = ScoreBreakdown(
                license_numerics=1.0, first_name=1.0, last_name=1.0,
                gate_passed=True, total=1.0,
            )
            trace.final_outcome = "Pass"
            return LadderResult(status="Pass", best_record=rec, best_breakdown=bd)

        trace.final_outcome = "Fail"
        trace.final_reason = "no_records"
        return LadderResult(status="Fail", reason="no_records")

    except Exception as exc:
        log.warning("[PSYPACT] Error for %s %s (%s): %s", first, last, license_id, exc,
                    exc_info=True)
        trace.final_outcome = "Fail"
        trace.final_reason = "no_records"
        return LadderResult(status="Fail", reason="no_records")


def _load_routing() -> None:
    if _ROUTING_CSV.exists():
        with open(_ROUTING_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                key = (row["state"].strip().upper(), row["psv_prov_type"].strip().upper())
                _ROUTING.setdefault(key, []).append(row["source_id"].strip())
        log.info("Loaded routing for %d (state,prov_type) pairs from %s", len(_ROUTING), _ROUTING_CSV.name)
    else:
        # CSV not present — fall back to the hardcoded dict in board_routing.py
        from board_routing import ROUTING as _HARDCODED  # noqa: PLC0415
        _ROUTING.update(_HARDCODED)
        log.info("CSV not found — loaded %d routing entries from board_routing.py", len(_ROUTING))


STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming",
}


_SUFFIX_RE = re.compile(r"\s+(?:Jr\.?|Sr\.?|II|III|IV|2nd|3rd|4th|Esq\.?)$", re.IGNORECASE)

# Socrata board type filters: (source_id, psv_prov_type) -> exact license_type/provider_type string.
# Populated from live API distinct-value queries (2026-06-22).
# IL_LICENSING uses license_type_selector="license_type" (board-level categories).
# WA_HEALTH uses provider_type_selector="credentialtype" (credential type strings).
# CO_DORA uses opaque short codes — no safe mapping; omitted (unfiltered search still works).
# DE_LICENSING has 297 granular types — omitted to avoid false-negative risk.
_SOCRATA_TYPE_MAP: dict[tuple[str, str], str] = {
    # IL_LICENSING — board-level category (safe 1:1 mapping)
    ("IL_LICENSING", "ABA"): "ADVISORY BD OF BEHAV",
    ("IL_LICENSING", "AP"):  "ACUPUNCTURE",
    ("IL_LICENSING", "AU"):  "SPEECH-LANGUAGE PATH",
    ("IL_LICENSING", "CP"):  "CLIN PSYCHOLOGIST",
    ("IL_LICENSING", "DDS"): "DENTAL",
    ("IL_LICENSING", "DMD"): "DENTAL",
    ("IL_LICENSING", "DO"):  "MEDICAL BOARD",
    ("IL_LICENSING", "DP"):  "PODIATRY",
    ("IL_LICENSING", "DT"):  "DIETETIC AND NUTRITION",
    ("IL_LICENSING", "LC"):  "PROF. COUNSELOR",
    ("IL_LICENSING", "LPC"): "PROF. COUNSELOR",
    ("IL_LICENSING", "MD"):  "MEDICAL BOARD",
    ("IL_LICENSING", "MT"):  "MASSAGE LICENSING BD",
    ("IL_LICENSING", "NP"):  "ADV PRACTICE NURSE",
    ("IL_LICENSING", "NPB"): "ADV PRACTICE NURSE",
    ("IL_LICENSING", "NPS"): "ADV PRACTICE NURSE",
    ("IL_LICENSING", "NSA"): "NURSING BOARD",
    ("IL_LICENSING", "OD"):  "OPTOMETRY",
    ("IL_LICENSING", "OT"):  "OCCUPATIONAL THERAPY",
    ("IL_LICENSING", "PA"):  "PHYSICIAN ASSISTANT",
    ("IL_LICENSING", "PAB"): "PHYSICIAN ASSISTANT",
    ("IL_LICENSING", "PAS"): "PHYSICIAN ASSISTANT",
    ("IL_LICENSING", "PH"):  "MEDICAL BOARD",
    ("IL_LICENSING", "MW"):  "NURSING BOARD",
    ("IL_LICENSING", "PN"):  "NURSING BOARD",
    ("IL_LICENSING", "PT"):  "PHYSICAL THERAPY",
    ("IL_LICENSING", "RN"):  "NURSING BOARD",
    ("IL_LICENSING", "RNA"): "NURSING BOARD",
    ("IL_LICENSING", "SH"):  "SPEECH-LANGUAGE PATH",
    ("IL_LICENSING", "SW"):  "SOCIAL WORKER",
    # WA_HEALTH — credential type string (exact match)
    ("WA_HEALTH", "DDS"): "Dentist License",
    ("WA_HEALTH", "DMD"): "Dentist License",
    ("WA_HEALTH", "DO"):  "Osteopathic Physician & Surgeon License",
    ("WA_HEALTH", "DP"):  "Podiatric Physician And Surgeon License",
    ("WA_HEALTH", "MD"):  "Physician And Surgeon License",
    ("WA_HEALTH", "NP"):  "Advanced Registered Nurse Practitioner License",
    ("WA_HEALTH", "NPB"): "Advanced Registered Nurse Practitioner License",
    ("WA_HEALTH", "OT"):  "Occupational Therapist License",
    ("WA_HEALTH", "PA"):  "Physician Assistant License",
    ("WA_HEALTH", "PAS"): "Physician Assistant License",
    ("WA_HEALTH", "PH"):  "Physician And Surgeon License",
    ("WA_HEALTH", "PM"):  "Pharmacist License",
    ("WA_HEALTH", "PN"):  "Licensed Practical Nurse",
    ("WA_HEALTH", "PT"):  "Physical Therapist License",
    ("WA_HEALTH", "RNA"): "Registered Nurse License",
    ("CO_DORA", "AP"):  "ACU",
    ("CO_DORA", "CP"):  "PSY",
    ("CO_DORA", "ND"):  "ND",
    ("CO_DORA", "RFA"): "SA",
    # WA SW omitted: too many LICSW subtypes; exact match would risk false negatives
    # KS_NURSING_KSBN — KSBN dropdown values (not prov_type codes)
    # Dropdown options: RN, LPN, LMHT, RNA, NP, NMW, CNS
    ("KS_NURSING_KSBN", "NP"):  "NP",
    ("KS_NURSING_KSBN", "NPB"): "NP",
    ("KS_NURSING_KSBN", "NPS"): "NP",
    ("KS_NURSING_KSBN", "NSA"): "NP",
    ("KS_NURSING_KSBN", "PN"):  "LPN",
    ("KS_NURSING_KSBN", "RN"):  "RN",
    ("KS_NURSING_KSBN", "RNA"): "RNA",
    ("KS_NURSING_KSBN", "MW"):  "NMW",
}

# Fallback license types to try when the primary type-filtered name search returns 0.
# Handles career-progression scenarios (e.g. PN→RN): the board may have moved the
# person to a different license class while the input still shows the original prov_type.
# Only fires for browser boards after ALL standard name searches return 0.
_BOARD_TYPE_FALLBACKS: dict[tuple[str, str], list[str]] = {
    # KSBN: PN/LPN nurses who upgraded to RN or APRN keep only the new license class.
    ("KS_NURSING_KSBN", "PN"):  ["RN", "NP"],
    ("KS_NURSING_KSBN", "NP"):  ["RN"],
    ("KS_NURSING_KSBN", "RN"):  ["LPN", "NP"],
}


def _normalize(s: str) -> str:
    # Replace commas, hyphens, apostrophes, periods with space; collapse whitespace
    return re.sub(r"\s+", " ", re.sub(r"[,\-.']+", " ", s.upper())).strip()


# Credential suffixes and honorific prefixes stripped from board-returned names
# before comparison to avoid false name_mismatch when the board appends "MD", "RN",
# etc. or prepends "Dr." to the licensee name.
_CRED_SUFFIXES: frozenset[str] = frozenset({
    "II", "III", "IV", "V", "JR", "SR", "ESQ",
    "MD", "DO", "DPM", "DDS", "DMD", "OD", "PHD", "PSYD",
    "DPT", "DC", "ND",
    "RN", "LPN", "LVN", "APRN", "DNP", "CNM", "NP", "PA", "CRNA",
    "LCSW", "LMFT", "LPC", "LCPC", "LMHC", "BCBA", "BCABA", "RBT",
    "PT", "OT", "SLP", "AUD",
    "PHARMD", "RPH",
    "FACP", "FACS", "FACOG", "FAAP",
})
_CRED_PREFIXES: frozenset[str] = frozenset({
    "DR", "MR", "MRS", "MS", "MISS", "PROF", "REV",
    "PASTOR", "RABBI", "SISTER", "BROTHER",
})


def _strip_name_credentials(s: str) -> str:
    """Remove leading honorific prefixes and trailing credential suffixes from
    an already-normalized (uppercase, comma-free) name string."""
    tokens = s.split()
    while tokens and tokens[0] in _CRED_PREFIXES:
        tokens = tokens[1:]
    while tokens and tokens[-1] in _CRED_SUFFIXES:
        tokens = tokens[:-1]
    return " ".join(tokens)


def _full_name(rec) -> str:
    fn = getattr(rec, "licensee_full_name", None)
    if fn:
        return fn
    parts = []
    f = getattr(rec, "licensee_first_name", None)
    l = getattr(rec, "licensee_last_name", None)
    if f:
        parts.append(f)
    if l:
        parts.append(l)
    return " ".join(parts)


def _name_matches(rec, last: str, first: str) -> bool:
    full = _strip_name_credentials(_normalize(_full_name(rec)))
    if not full:
        return False
    # Space-collapsed form so multi-token surnames whose spacing differs between the
    # board and the input still match (board "DI NITTO"/"DE LEON"/"MONTES DE OCA" vs
    # input "Dinitto"/"Deleon"/"Montesdeoca", and vice-versa).
    full_ns = full.replace(" ", "")
    last_norm = _strip_name_credentials(_normalize(last)) if last else ""
    if last_norm and last_norm not in full and last_norm.replace(" ", "") not in full_ns:
        # Hyphenated surname fallback: board may store only one component of the name
        # (e.g. board shows "BATES, AMY J" while PSV has "Bates-Daly" → try "BATES" or "DALY").
        if "-" in last:
            parts = [_strip_name_credentials(_normalize(p)) for p in last.split("-") if p.strip()]
            if not any(p and (p in full or p.replace(" ", "") in full_ns) for p in parts):
                return False
        else:
            return False
    if first:
        first_norm = _strip_name_credentials(_normalize(first))
        if first_norm and first_norm not in full and first_norm.replace(" ", "") not in full_ns:
            return False
    return True


def _license_matches(rec, lid: str) -> bool:
    if not lid:
        return True
    lic = getattr(rec, "license_number", None) or ""
    if not lic:
        # Board doesn't expose license numbers in results — accept name-only match
        return True
    lid_u = lid.upper().strip()
    rec_u = lic.upper().strip()
    if lid_u == rec_u:
        return True
    # Legacy license number fallback: some boards keep an old numeric ID alongside the
    # current license number (e.g. KY_MULTIBOARD col3 "Legacy Number" vs col4 "License Number").
    # When the board-config maps that column to legacy_license_number, check it here.
    _raw = getattr(rec, "raw_fields", None) or {}
    _legacy = (_raw.get("legacy_license_number") or "").upper().strip()
    if _legacy:
        if lid_u == _legacy:
            return True
        if lid_u.isdigit() and _legacy.isdigit() and lid_u.lstrip("0") == _legacy.lstrip("0"):
            return True
    # Substring match only when at least one side is alphanumeric (prefix/suffix case
    # like "8901" ⊂ "LC8901").  Two all-digit strings must be an exact match — otherwise
    # "3940" spuriously matches inside "13940", "23940", etc.
    if lid_u.isdigit() and rec_u.isdigit():
        # Middle-group exception: board returns just the center digits of a longer license
        # (e.g. KSBN returns "84236" for full input "5384236101").
        # Require returned value ≥ 4 shorter than input to prevent "3940" ⊂ "13940" (diff=1).
        if len(rec_u) >= 4 and len(lid_u) - len(rec_u) >= 4 and rec_u in lid_u:
            return True
        # Leading-zero tolerance for pure-digit pairs: "4102" == "04102".
        # The lstrip check below is unreachable in this branch without this explicit guard.
        if lid_u.lstrip("0") == rec_u.lstrip("0"):
            return True
        return False
    if lid_u in rec_u:
        if lid_u.isdigit():
            _dig_rec = re.sub(r"\D", "", rec_u)
            if len(_dig_rec) > len(lid_u) + 2:
                pass  # skip — year-prefix or totally different number (e.g. "2561" in "17-02561")
            elif not _dig_rec.endswith(lid_u):
                pass  # skip — digits appear at front of board value, not trailing ("1495" in "14959")
            else:
                return True
        else:
            return True
    # Middle-digits-only boards (e.g. KSBN): board returns the center group of a dashed
    # license — "81920" ⊂ "53-81920-022". Only reached when input is NOT all-digit (dashes
    # skip the all-digit guard above), so min-length of 4 prevents accidental short matches.
    if rec_u and len(rec_u) >= 4 and rec_u in lid_u:
        return True
    # Numeric-only fallback: "LPC 04643" should match "LCPC 04643" since the digit
    # portion is identical (different type-prefix conventions across PSV vs board).
    lid_num = re.sub(r"\D", "", lid_u)
    rec_num = re.sub(r"\D", "", rec_u)
    if lid_num and rec_num and len(lid_num) >= 4 and lid_num == rec_num:
        return True
    # Leading-zero tolerance: "01041" == "1041" after stripping leading zeros
    if lid_num and rec_num and len(lid_num) >= 3 and lid_num.lstrip("0") == rec_num.lstrip("0"):
        return True
    # Leading numeric group match: handles NNNNN-XX-V versioned credential formats such as
    # NV_DIETITIAN ("39673-DI-3" vs input "39673-D1-2" or "40902-DI-1" vs "40902-DI-0").
    # Board renumbers credentials on each renewal cycle; the first digit block is the stable key.
    # Only fires when BOTH sides have a leading digit group of ≥ 4 followed by a non-digit,
    # so pure-digit comparisons (already handled above) and short IDs are unaffected.
    m_in = re.match(r"^(\d{4,})\D", lid_u)
    m_rec = re.match(r"^(\d{4,})\D", rec_u)
    if m_in and m_rec and m_in.group(1) == m_rec.group(1):
        return True
    # Cross-format match: handles TYPE-prefixed board IDs where the board stores
    # "DI-40215" (prefix-number) but input carries "40215-DI-1" (number-prefix-version).
    # Finds the first ≥ 5-digit group in each side; short IDs (≤ 4 digits) are excluded to
    # avoid false matches on common short sequences (e.g. "0076" in "A-0076").
    m_in_any = re.search(r"(\d{5,})", lid_u)
    m_rec_any = re.search(r"(\d{5,})", rec_u)
    if m_in_any and m_rec_any and m_in_any.group(1) == m_rec_any.group(1):
        return True
    return False


class PsvBrowser:
    """Shared Playwright browser for one board config. Avoids per-call browser launch overhead."""

    def __init__(self, config, browser: Browser, proxy_cfg):
        self.config = config
        self._browser = browser
        self._proxy = proxy_cfg

    async def search(self, query: SearchQuery, timeout_ms: int = 45000,
                     run_id: str = "") -> list:
        """Open a new context+page, search, extract grid rows, close context. ~15s per call.
        Captures per-rung evidence (HTML + screenshot) for browser archetypes —
        otherwise the orchestrator's per-attempt evidence column points to a
        path that was never written.
        """
        src = self.config.identity.source_id
        state = self.config.identity.state
        ua = self.config.transport.user_agent
        if ua == "LVS-LicenseVerifier/1.0":
            ua = _REAL_UA
        # Empty ua means "don't override" — lets Chromium send its native UA so that
        # User-Agent and Sec-CH-UA headers stay consistent (mismatch triggers WAF 403).
        ctx_kwargs: dict = dict(
            viewport={"width": 1280, "height": 900},
            proxy=self._proxy,
            locale="en-US",
            timezone_id="America/New_York",
            ignore_https_errors=self.config.transport.ignore_https_errors,
        )
        if ua:
            ctx_kwargs["user_agent"] = ua
        ctx = await self._browser.new_context(**ctx_kwargs)
        ctx.set_default_timeout(timeout_ms)
        ctx.set_default_navigation_timeout(min(timeout_ms, 30000))
        if _STEALTH is not None:
            await _STEALTH.apply_stealth_async(ctx)
        page = await ctx.new_page()
        _search_start = time.monotonic()
        _timeout_s = timeout_ms / 1000
        try:
            log.info("[%s] search mode=%s query=%s", src, query.mode, query.query)
            await navigate_to_search(page, self.config)
            has_results = await fill_search_form(page, self.config, query)
            if not has_results:
                log.info("[%s] No results for query=%s", src, query.query)
                if run_id:
                    try:
                        await capture_evidence(page, self.config.evidence,
                                                stage="search_results",
                                                run_id=run_id, source_id=src,
                                                state=state, query=query)
                    except Exception:
                        pass
                return []
            # Single-result redirect: portal goes directly to detail page instead of results list.
            single_pat = getattr(self.config.results, "single_result_url_pattern", None)
            if single_pat and single_pat in page.url:
                raw = await extract_detail(page, self.config.detail)
                await _try_out_of_state_tab(page, self.config, raw)
                if run_id:
                    try:
                        await capture_evidence(page, self.config.evidence,
                                                stage="detail_page",
                                                run_id=run_id, source_id=src,
                                                state=state, query=query)
                    except Exception:
                        pass
                return [map_to_license_record(raw, self.config, {})]
            if self.config.results.type == "th_td_multi":
                raw_rows_multi = await extract_th_td_multi(page, self.config.results)
                if run_id:
                    try:
                        await capture_evidence(page, self.config.evidence,
                                                stage="search_results",
                                                run_id=run_id, source_id=src,
                                                state=state, query=query)
                    except Exception:
                        pass
                mapped = [
                    map_to_license_record(
                        apply_field_map(r, self.config.detail.field_map),
                        self.config, {}
                    )
                    for r in raw_rows_multi
                ]
                return mapped
            raw_rows, _warn = await extract_results_table(page, self.config.results)
            if _warn:
                log.warning("[%s] extract_results_table partial: %s", src, _warn)

            # Opt-in name-mode pagination (results.paginate_summary_rows): boards like
            # AR_MEDBOARD order their name-search GridView alphabetically and show ~10
            # rows/page, so the target person is frequently on a later page. Detail-
            # clicking each row (below) breaks the ASP.NET pager, so we'd only ever see
            # page 1. Instead collect summary rows across ALL pages here and match on the
            # summary table (which already carries full_name + license_number); expiry for
            # the winning row is fetched on demand by the caller (_fetch_detail_expiry).
            _NAME_QUERY_MODES_PG = {
                "first_name", "last_name", "first_and_last", "first_and_last_typed",
            }
            _pg = self.config.results.pagination
            _paginated_summary = False
            if (getattr(self.config.results, "paginate_summary_rows", False)
                    and query.mode in _NAME_QUERY_MODES_PG
                    and raw_rows and _pg and _pg.enabled
                    and _pg.strategy == "next_button" and _pg.next_selector):
                _max_pages = 60
                _page_n = 1
                _row_sel = self.config.results.table.row_selector if self.config.results.table else None

                async def _first_row_sig() -> str:
                    # Content sentinel for detecting an ASP.NET postback page advance —
                    # networkidle is unreliable for GridView __doPostBack (the connection
                    # can stay open). Watch the first result row's text instead.
                    if not _row_sel:
                        return ""
                    try:
                        loc = page.locator(_row_sel).first
                        if await loc.count() > 0:
                            return (await loc.inner_text())[:120]
                    except Exception:
                        pass
                    return ""

                while _page_n < _max_pages:
                    # Leave headroom in the row's time budget for matching + expiry fetch.
                    if time.monotonic() - _search_start >= _timeout_s * 0.6:
                        log.warning("[%s] summary pagination time budget reached at page %d",
                                    src, _page_n)
                        break
                    try:
                        _nxt = page.locator(_pg.next_selector).first
                        if await _nxt.count() == 0:
                            break
                        _cls = (await _nxt.get_attribute("class") or "").lower()
                        if _pg.disabled_class and _pg.disabled_class.lower() in _cls:
                            break
                        if (await _nxt.get_attribute("aria-disabled")) == "true":
                            break
                        _sig_before = await _first_row_sig()
                        await _nxt.click()
                        # Wait for the results table to actually change (postback complete),
                        # not just for the network to idle.
                        _advanced = False
                        for _ in range(30):  # up to ~15s
                            await asyncio.sleep(0.5)
                            if await _first_row_sig() != _sig_before:
                                _advanced = True
                                break
                        if not _advanced:
                            log.info("[%s] pagination: page did not change after next-click at page %d — stopping",
                                     src, _page_n)
                            break
                        _more, _ = await extract_results_table(page, self.config.results)
                        if not _more:
                            break
                        raw_rows.extend(_more)
                        _page_n += 1
                    except Exception as _pg_err:
                        log.warning("[%s] summary pagination stopped at page %d: %s",
                                    src, _page_n, _pg_err)
                        break
                _paginated_summary = True
                log.info("[%s] name-mode summary pagination: %d row(s) across %d page(s)",
                         src, len(raw_rows), _page_n)
            # AG-Grid returns rows keyed by column-header text ("License Number", "First Name", …).
            # apply_field_map normalises these to snake_case so targeting and map_to_license_record
            # can find them via the expected keys ("license_number", "first_name", …).
            if self.config.results.type == "ag_grid" and self.config.detail.field_map:
                raw_rows = [apply_field_map(r, self.config.detail.field_map) for r in raw_rows]
            if run_id:
                try:
                    await capture_evidence(page, self.config.evidence,
                                            stage="search_results",
                                            run_id=run_id, source_id=src,
                                            state=state, query=query)
                except Exception:
                    pass
            # Follow detail pages when the board stores expiry/full data only on the detail page.
            # For license_number searches: visit up to 10 detail pages.
            # For name-mode searches (first_name, last_name, first_and_last, etc.): visit ALL
            # returned rows — a first_name search may return several records and we must evaluate
            # every one (checking both name AND license) before declaring no match and moving to
            # the next board. Capping at 1 was causing false negatives when the correct record
            # happened not to be the first row returned.
            _has_detail = (
                self.config.results.has_detail_page
                and self.config.results.detail_trigger
            )
            _NAME_QUERY_MODES = {
                "first_name", "last_name", "first_and_last", "first_and_last_typed",
            }
            _detail_limit = (
                len(raw_rows) if query.mode in _NAME_QUERY_MODES else 10
            )
            # Boards like MD_PHYSICIANS show a single view-button (#btnLICNO2) instead
            # of a results table. extract_results_table returns 0 rows, but the trigger
            # button IS present. Synthesise one placeholder row so the detail loop runs.
            if _has_detail and not raw_rows:
                trigger_sel = self.config.results.detail_trigger.selector
                try:
                    if await page.locator(trigger_sel).count() > 0:
                        raw_rows = [{}]
                except Exception:
                    pass
            # Name-hint narrowing: when the query carries first/last name alongside a
            # license_number search (detail-expiry re-fetch), find the single matching
            # row and visit only that detail page — avoids O(N) visits on multi-name boards.
            _detail_hint_fn = (getattr(query, "first_name", None) or "").upper().strip()
            _detail_hint_ln = (getattr(query, "last_name", None) or "").upper().strip()
            _detail_targeted_idx: int | None = None
            if (query.mode == "license_number" and (_detail_hint_fn or _detail_hint_ln)
                    and _has_detail and raw_rows and len(raw_rows) > 1):
                for _ri, _rw in enumerate(raw_rows):
                    _rw_ln = (_rw.get("last_name", "") or "").upper().strip()
                    _rw_fn = (_rw.get("first_name", "") or "").upper().strip()
                    # Boards that return full_name instead of split fields (e.g. MD_PT
                    # returns "MCDERMOTT, KYLE M.") — split on the first comma.
                    if not (_rw_ln or _rw_fn):
                        _rw_full = (_rw.get("full_name", "") or "").upper().strip()
                        if "," in _rw_full:
                            _rw_ln = _rw_full.split(",", 1)[0].strip()
                            _rw_fn = _rw_full.split(",", 1)[1].strip()
                        elif _rw_full:
                            _rw_fn = _rw_full
                    _ln_ok = not _detail_hint_ln or _rw_ln == _detail_hint_ln
                    _fn_ok = not _detail_hint_fn or _rw_fn == _detail_hint_fn
                    if _ln_ok and _fn_ok:
                        _detail_targeted_idx = _ri
                        log.info("[%s] Name-hint: targeting detail idx=%d (%s %s)",
                                 src, _ri, _rw_fn, _rw_ln)
                        break
            # License-number-based targeting: fires when 2+ rows are returned from a
            # license_number search and name-hint didn't resolve (e.g. boards that only
            # expose full_name without names in the query, or license-only queries).
            # Exact string match is tried first; numeric-only match (leading-zero tolerance)
            # is the fallback, but ONLY when both sides share the same alpha/numeric type —
            # this prevents "21524" from falsely matching "CP021524T" via stripped digits.
            if (_detail_targeted_idx is None and query.mode == "license_number"
                    and _has_detail and raw_rows and len(raw_rows) > 1):
                _lic_hint = (getattr(query, "license_number", None) or "").strip()
                if _lic_hint:
                    _lic_hint_u = _lic_hint.upper()
                    _lic_hint_num = re.sub(r'\D', '', _lic_hint_u).lstrip('0') or '0'
                    _lic_hint_has_alpha = bool(re.search(r'[A-Za-z]', _lic_hint_u))
                    for _ri, _rw in enumerate(raw_rows):
                        _rw_lic = (_rw.get("license_number", "") or "").strip()
                        if not _rw_lic:
                            continue
                        _rw_lic_u = _rw_lic.upper()
                        _rw_lic_num = re.sub(r'\D', '', _rw_lic_u).lstrip('0') or '0'
                        _rw_lic_has_alpha = bool(re.search(r'[A-Za-z]', _rw_lic_u))
                        _exact_match = _rw_lic_u == _lic_hint_u
                        _numeric_match = (
                            _rw_lic_num != '0' and _lic_hint_num != '0'
                            and len(_rw_lic_num) >= 3
                            and _rw_lic_num == _lic_hint_num
                            and _rw_lic_has_alpha == _lic_hint_has_alpha
                        )
                        if _exact_match or _numeric_match:
                            _detail_targeted_idx = _ri
                            log.info("[%s] License-hint: targeting detail idx=%d (lic=%s)",
                                     src, _ri, _rw_lic)
                            break
            # Name-mode license-hint narrowing: when searching by name but the query
            # still carries the original license number (e.g. license_number search
            # returned no results so the ladder fell back to first_name), scan the
            # summary table for rows whose license_number matches. Visiting only those
            # prevents an O(N) detail sweep when a common first name returns hundreds
            # of rows.
            _name_lic_indices: list | None = None
            if (query.mode in _NAME_QUERY_MODES and _detail_targeted_idx is None
                    and query.license_number and _has_detail and raw_rows):
                _nlh = query.license_number.strip().upper()
                _nlh_num = re.sub(r'\D', '', _nlh).lstrip('0') or '0'
                _nlh_alpha = bool(re.search(r'[A-Za-z]', _nlh))
                _nlh_hits = []
                for _ri, _rw in enumerate(raw_rows):
                    _rl = (_rw.get("license_number", "") or "").strip().upper()
                    if not _rl:
                        continue
                    _rn = re.sub(r'\D', '', _rl).lstrip('0') or '0'
                    if _rl == _nlh or (
                            _rn != '0' and _nlh_num != '0' and _rn == _nlh_num
                            and bool(re.search(r'[A-Za-z]', _rl)) == _nlh_alpha):
                        _nlh_hits.append(_ri)
                if _nlh_hits:
                    log.info("[%s] Name-mode lic-hint: %d/%d rows match lic=%s",
                             src, len(_nlh_hits), len(raw_rows), query.license_number)
                    _name_lic_indices = _nlh_hits
                else:
                    log.info("[%s] Name-mode lic-hint: no summary rows match lic=%s — visiting all %d rows",
                             src, query.license_number, len(raw_rows))

            _visit_indices = (
                [_detail_targeted_idx] if _detail_targeted_idx is not None
                else list(_name_lic_indices) if _name_lic_indices is not None
                else range(len(raw_rows))
            )
            if _has_detail and raw_rows and not _paginated_summary and (
                    _detail_targeted_idx is not None
                    or _name_lic_indices is not None
                    or len(raw_rows) <= _detail_limit):
                trigger_sel = self.config.results.detail_trigger.selector
                detailed = []
                for _idx in _visit_indices:
                    if time.monotonic() - _search_start >= _timeout_s * 0.85:
                        log.warning("[%s] Detail loop time budget (%.0fs) nearly exhausted at idx=%d — stopping early",
                                    src, _timeout_s, _idx)
                        break
                    try:
                        btn = page.locator(trigger_sel).nth(_idx)
                        if not await btn.is_visible(timeout=3000):
                            break
                        # PDF detail: the trigger links to a PDF (e.g. NC_DAC's
                        # /PractitionerLookup/Detail/{id} serves a "Credential Status"
                        # letter as application/pdf even though the href lacks a .pdf
                        # suffix). Download and parse it in-place instead of navigating —
                        # navigating would let extract_detail capture the page's
                        # "Credential Status" heading as the licensee name. No back-
                        # navigation is needed since the browser never leaves the results.
                        _dt = self.config.results.detail_trigger
                        _href = (await btn.get_attribute("href") or "").strip()
                        _is_pdf = (
                            getattr(_dt, "force_pdf", False)
                            or _href.lower().endswith(".pdf")
                            or "pdf" in _href.lower().split("?")[0]
                        )
                        # BRANCH A: Direct JSON API detail (e.g. PA_PALS)
                        # When config.detail.api is set, the board's "detail link"
                        # opens a new _blank tab — Playwright's URL-change wait
                        # never fires. Skip the click entirely; call the backing
                        # JSON API directly (in the current page context so session
                        # cookies are sent automatically) and merge the response
                        # into the summary row. No back-navigation needed.
                        # See _fetch_detail_via_api in browser_form.py for docs.
                        if self.config.detail.api:
                            _api_raw = await _fetch_detail_via_api(page, self.config, _idx)
                            # Backfill fields missing from the API response with the
                            # corresponding summary-row value (e.g. full_name and
                            # license_type come from the results table, not the API).
                            if _idx < len(raw_rows):
                                _sr = raw_rows[_idx]
                                for _k in ("full_name", "first_name", "last_name",
                                           "license_number", "license_type",
                                           "status", "board", "address"):
                                    if not _api_raw.get(_k) and _sr.get(_k):
                                        _api_raw[_k] = _sr[_k]
                            detailed.append(map_to_license_record(_api_raw, self.config, {}))
                            continue

                        if _is_pdf:
                            if not _href:
                                log.warning("[%s] force_pdf but empty href at idx=%d — using summary row",
                                            src, _idx)
                                if _idx < len(raw_rows):
                                    detailed.append(map_to_license_record(raw_rows[_idx], self.config, {}))
                                continue
                            _pdf_raw = await _scrape_pdf_detail(page, _href, self.config)
                            _pdf_mapped = apply_field_map(_pdf_raw, self.config.detail.field_map)
                            # Backfill from the summary row for anything the letter omitted
                            # (e.g. the inactive-credential letter carries no license number).
                            if _idx < len(raw_rows):
                                _sr = raw_rows[_idx]
                                for _k in ("full_name", "first_name", "last_name",
                                           "license_number", "license_type",
                                           "city", "state", "status",
                                           "issue_date", "expiration_date"):
                                    if not _pdf_mapped.get(_k) and _sr.get(_k):
                                        _pdf_mapped[_k] = _sr[_k]
                            detailed.append(map_to_license_record(_pdf_mapped, self.config, {}))
                            continue
                        if getattr(self.config.results.detail_trigger, "opens_modal", False):
                            # Modal detail (no navigation): fire the row's own click handler
                            # via JS so a cookie/consent overlay can't intercept the pointer,
                            # then wait directly for the modal body to fill. Skips the
                            # URL-change wait, which never fires for a modal and would burn
                            # the full detail timeout on every row.
                            try:
                                await btn.evaluate("el => el.click()")
                            except Exception:
                                await btn.click()
                        else:
                            await btn.evaluate("el => el.removeAttribute('target')")
                            url_before = page.url
                            await btn.click()
                            try:
                                await page.wait_for_function(
                                    "url => window.location.href !== url",
                                    url_before,
                                    timeout=self.config.detail.wait.timeout_ms,
                                )
                            except Exception:
                                pass
                        await _wait_for_detail_content(page, self.config)
                        raw = await extract_detail(page, self.config.detail)
                        await _try_out_of_state_tab(page, self.config, raw)
                        # Supplement with summary-row fields not captured on the detail
                        # page (e.g. license_type on NC_SLP_AUD lives only in the table).
                        if _idx < len(raw_rows):
                            _sr = raw_rows[_idx]
                            for _k in ("license_type", "city", "state_code"):
                                if not raw.get(_k) and _sr.get(_k):
                                    raw[_k] = _sr[_k]
                        if run_id:
                            try:
                                await capture_evidence(page, self.config.evidence,
                                                        stage="detail_page",
                                                        run_id=run_id, source_id=src,
                                                        state=state, query=query)
                            except Exception:
                                pass
                        detailed.append(map_to_license_record(raw, self.config, {}))
                    except Exception as det_err:
                        log.warning("[%s] detail_click idx=%d failed: %s — using summary row",
                                    src, _idx, det_err)
                        if _idx < len(raw_rows):
                            detailed.append(map_to_license_record(raw_rows[_idx], self.config, {}))
                    # Navigate back to results page before clicking the next row's detail link.
                    if _idx < len(raw_rows) - 1:
                        try:
                            await _navigate_back(page, self.config)
                            try:
                                await page.wait_for_load_state("networkidle", timeout=10000)
                            except Exception:
                                await asyncio.sleep(2)
                        except Exception as back_err:
                            log.warning("[%s] navigate_back failed at detail idx=%d: %s",
                                        src, _idx, back_err)
                            break
                if detailed:
                    return detailed
            return [map_to_license_record(r, self.config, {}) for r in raw_rows]
        except BoardUnavailableError:
            # Board site is down/erroring — propagate so the ladder classifies the
            # row as Skip (board_unavailable), rather than swallowing to [] which
            # looks identical to a genuine no-records result.
            log.warning("[%s] Board unavailable mode=%s query=%s", src, query.mode, query.query)
            if run_id:
                try:
                    await capture_evidence(page, self.config.evidence,
                                            stage="error",
                                            run_id=run_id, source_id=src,
                                            state=state, query=query)
                except Exception:
                    pass
            raise
        except Exception as exc:
            log.warning("[%s] Error mode=%s query=%s: %s", src, query.mode, query.query, exc)
            if run_id:
                try:
                    await capture_evidence(page, self.config.evidence,
                                            stage="error",
                                            run_id=run_id, source_id=src,
                                            state=state, query=query)
                except Exception:
                    pass
            return []
        finally:
            try:
                await page.close()
                await ctx.close()
            except Exception:
                pass


async def _try_search_psv(psv_b: PsvBrowser, query: SearchQuery, timeout: int) -> list:
    """Wrap PsvBrowser.search with asyncio timeout. Never raises."""
    src = psv_b.config.identity.source_id
    try:
        return await asyncio.wait_for(
            psv_b.search(query, timeout_ms=timeout * 1000),
            timeout=float(timeout),
        )
    except asyncio.TimeoutError:
        log.warning("[%s] Timeout for mode=%s query=%s", src, query.mode, query.query)
        return []
    except BoardUnavailableError:
        raise  # let the caller classify the board as down (Skip)
    except Exception as exc:
        log.warning("[%s] Error for mode=%s query=%s: %s", src, query.mode, query.query, exc)
        return []


async def _try_search_api(config, query: SearchQuery, timeout: int) -> list:
    """Wrap verify_license for non-browser boards (csv_bulk, socrata, etc.). Never raises."""
    src = config.identity.source_id
    try:
        records = await asyncio.wait_for(
            verify_license(config, query, db=None),
            timeout=float(timeout),
        )
        return records or []
    except asyncio.TimeoutError:
        log.warning("[%s] Timeout for mode=%s query=%s", src, query.mode, query.query)
        return []
    except Exception as exc:
        log.warning("[%s] Error for mode=%s query=%s: %s", src, query.mode, query.query, exc)
        return []


def _epdb_name_gate_note(rec, last_name: str, first_name: str) -> str:
    """Return a bracketed gate note for standalone psv_test.py reason strings.

    Empty when the gate approves or skips (names already match).
    Called only inside run_row() — EPDB-only (no NPPES in standalone mode).
    """
    try:
        from orchestrator.name_gate import evaluate_name_gate  # noqa: PLC0415
        ng = evaluate_name_gate(
            master_row={"first_name": first_name, "last_name": last_name},
            board_record=rec,
            nppes=None,
        )
        if ng.skipped or ng.verdict == "approve":
            return ""
        score_str = f"{ng.epdb_score:.2f}" if ng.epdb_score is not None else "N/A"
        return f" [gate:{ng.verdict} epdb={score_str}]"
    except Exception:
        return ""


# Status preference for choosing among multiple matching records (lower = better).
# Boards commonly return several rows for the same person — e.g. NC counseling boards
# list a superseded "A#####" associate credential (status "Transitioned" → INACTIVE)
# *before* the current full-licence row (ACTIVE). Selecting by table order would pick
# the stale/expired associate record; ranking by status surfaces the ACTIVE one instead.
_STATUS_PREFERENCE = {
    LicenseStatus.ACTIVE: 0,
    LicenseStatus.PROBATION: 1,
    LicenseStatus.UNKNOWN: 2,   # status not parsed — don't penalise below a known-bad row
    LicenseStatus.INACTIVE: 3,
    LicenseStatus.EXPIRED: 3,
    LicenseStatus.SUSPENDED: 3,
    LicenseStatus.REVOKED: 3,
}


def _status_rank(rec) -> int:
    """Preference rank for a record's status (lower = prefer). Unknown statuses rank
    with UNKNOWN so an unparsed row is never chosen over an ACTIVE one."""
    return _STATUS_PREFERENCE.get(getattr(rec, "status", LicenseStatus.UNKNOWN), 2)


def _prefer_active(recs: list) -> list:
    """Stable-sort matching records so ACTIVE rows come first, preserving the board's
    original order for records of equal status (Python's sort is stable)."""
    return sorted(recs, key=_status_rank)


def _match_analysis(records, last_name, first_name, license_id):
    """Return (both, name_hits, lic_hits) lists from records.

    Each list is ordered ACTIVE-first (stable), so callers that take element [0]
    verify against the current/active record rather than a superseded one that the
    board happened to list first (e.g. NC "Transitioned" A-prefixed associate rows)."""
    both = [r for r in records if _name_matches(r, last_name, first_name) and _license_matches(r, license_id)]
    name_hits = [r for r in records if _name_matches(r, last_name, first_name)]
    lic_hits = [r for r in records if _license_matches(r, license_id)]
    return _prefer_active(both), _prefer_active(name_hits), _prefer_active(lic_hits)


async def _try_first_and_last_psv(
    psv_b, first_name: str, last_name: str, middle_name: str,
    license_id: str, src_id: str, timeout: int,
    type_kwargs: dict | None = None,
    try_swap: bool = False,
) -> tuple | None:
    """Try first_and_last search if the board supports it. Returns (status, reason, expiry) or None.

    When try_swap=True (board has identity.try_name_swap=True), also retries with first/last
    names reversed if the standard order produces no match. Handles inputs where names are stored
    in "LastName FirstName" order without a comma delimiter.
    """
    cfg_modes = [m.mode for m in (psv_b.config.search.modes or [])]
    if "first_and_last" not in cfg_modes or not first_name or not last_name:
        return None

    async def _run_fal(f: str, l: str, label: str) -> tuple | None:
        q = SearchQuery(
            mode="first_and_last",
            query=f"{f} {l}",
            license_number=license_id,
            first_name=f,
            middle_name=middle_name or None,
            last_name=l,
            **(type_kwargs or {}),
        )
        recs = await _try_search_psv(psv_b, q, timeout)
        if not recs:
            return None
        both, name_hits, lic_hits = _match_analysis(recs, l, f, license_id)
        if both:
            _gn = _epdb_name_gate_note(both[0], l, f)
            return "Pass", f"Verified via {src_id} ({label}){_gn}", _get_expiry(both[0])
        if name_hits and not lic_hits:
            if license_id.upper().startswith("TC"):
                return "Pass", f"Verified via {src_id} (TC — {label} name match)", _get_expiry(name_hits[0])
        if lic_hits:
            last_only = [r for r in lic_hits if _name_matches(r, l, "")]
            if last_only:
                return "Pass", f"Verified via {src_id} ({label}, license+last name match)", _get_expiry(last_only[0])
            # License numerics matched but name didn't — route to AIAddLicense via
            # the "license match — name on board differs" reason so output_emitter
            # catches it as "License matched but Name mismatched".
            return "Pass", f"Verified via {src_id} ({label}, license match — name on board differs)", _get_expiry(lic_hits[0])
        return None

    result = await _run_fal(first_name, last_name, "first+last search")
    if result:
        return result

    # Name-swap fallback: input data sometimes stores names as "LastName FirstName" without
    # a comma separator, making it ambiguous which token is first vs last.
    if try_swap:
        log.info("[%s] Trying swapped name: first=%s last=%s", src_id, last_name, first_name)
        result = await _run_fal(last_name, first_name, "first+last swapped search")
        if result:
            return result

    return None


def _get_expiry(rec) -> str:
    """Return expiration_date as ISO string, or '' if not available."""
    d = getattr(rec, "expiration_date", None)
    if d is None:
        return ""
    try:
        return d.isoformat()
    except Exception:
        return str(d)


async def _fetch_detail_expiry(psv_b: "PsvBrowser", rec, timeout: int) -> str:
    """Secondary targeted lookup: re-search by the board's own license number to trigger
    detail_click_single and capture expiry.  Used when a multi-row search result matched
    but the detail page was never visited (so expiry is absent on the summary record).
    Only fires when the board has a detail page and the record carries a license number."""
    cfg = psv_b.config
    if not (cfg.results.has_detail_page and cfg.results.detail_trigger):
        return ""
    board_lic = getattr(rec, "license_number", None) or ""
    if not board_lic:
        return ""
    q = SearchQuery(
        mode="license_number", query=board_lic, license_number=board_lic,
        first_name=(getattr(rec, "licensee_first_name", None) or "").strip() or None,
        last_name=(getattr(rec, "licensee_last_name", None) or "").strip() or None,
    )
    detail_recs = await _try_search_psv(psv_b, q, timeout)
    # The re-search may again return several rows for the same person (e.g. NC boards
    # returning a superseded "A#####" Transitioned row alongside the ACTIVE one).
    # `board_lic` is the chosen record's own number, so a row matching it exactly is the
    # one we want — this excludes the A-prefixed superseded credential. Take that row's
    # expiry first; only if no exact-match row carries one do we fall back to any other
    # ACTIVE row (never an inactive/Transitioned one, whose stale date would false-Fail).
    board_lic_u = board_lic.upper().strip()
    exact = [dr for dr in detail_recs
             if (getattr(dr, "license_number", None) or "").upper().strip() == board_lic_u]
    for dr in _prefer_active(exact):
        exp = _get_expiry(dr)
        if exp:
            return exp
    for dr in detail_recs:
        if getattr(dr, "status", None) == LicenseStatus.ACTIVE and dr not in exact:
            exp = _get_expiry(dr)
            if exp:
                return exp
    return ""


async def run_row(
    row_data: dict,
    psv_browsers: dict,   # source_id -> PsvBrowser (browser boards)
    api_configs: dict,    # source_id -> SiteConfig (non-browser boards)
    all_configs: dict,    # source_id -> SiteConfig (for routing lookup)
    timeout: int,
) -> tuple[str, str, str]:
    """Returns (status, reason, expiry_date). expiry_date is ISO string or '' for Fail rows."""
    first_name = row_data["first_name"]
    last_name = _SUFFIX_RE.sub("", row_data["last_name"]).strip()
    middle_name = row_data["middle_name"]
    license_id = row_data["license_id"]
    prov_type = row_data.get("prov_type", "").upper().strip()
    lic_state = row_data.get("lic_state", "").upper().strip()

    if not license_id:
        return "Fail", "No license ID", ""

    # --- PSYPACT / APIT E.Passport: bypass state board, verify via directory.psypact.gov ---
    _lic_upper = license_id.upper()
    if _lic_upper.startswith("PSYPACT") or _lic_upper.startswith("APIT"):
        _trace_stub = types.SimpleNamespace(final_outcome=None, final_reason=None)
        _lr = await _verify_psypact_row(row_data, _trace_stub)
        if _lr.status == "Pass" and getattr(_lr, "best_record", None):
            _exp = getattr(_lr.best_record, "expiration_date", None)
            _exp_str = _exp.isoformat() if hasattr(_exp, "isoformat") else (str(_exp) if _exp else "")
            return "Pass", "Verified via PSYPACT_DIRECTORY (psypact.gov)", _exp_str
        return "Fail", "no_records", ""

    # --- ABA + BACB certification number: skip immediately (registry is CAPTCHA/maintenance-blocked) ---
    if prov_type == "ABA" and _is_bacb_license(license_id):
        return "Skip", "BACB Certificant Registry — CAPTCHA-based board (registry unavailable)", ""

    # --- NC LPC + LCAS prefix: wrong board (NCASPPB, not NC_MENTAL_HEALTH) ---
    # LCAS-##### is an NC Licensed Clinical Addiction Specialist credential issued by
    # the NC Addictions Specialist Professional Practice Board (NCASPPB), not by the
    # NC Board of Licensed Clinical Mental Health Counselors routed via NC_MENTAL_HEALTH.
    # No NCASPPB board config is available yet; return Fail with an actionable message.
    if lic_state == "NC" and prov_type == "LPC" and license_id.upper().startswith("LCAS"):
        return "Fail", (
            "License prefix LCAS indicates NC Addictions Specialist Professional Practice Board "
            "(NCASPPB) — no board config available for this credential type"
        ), ""

    # --- Cap: License State must appear in Service Location State ---
    _svc_raw = row_data.get("svc_loc_state", "")
    _svc_states = [s.strip().upper() for s in _svc_raw.split(",") if s.strip()]
    if _svc_states and lic_state not in _svc_states:
        return "N/A", (
            f"License State ({lic_state}) not in Service Location State "
            f"({_svc_raw}) — PSV step N/A"
        ), ""

    preferred_sids = _ROUTING.get((lic_state, prov_type), [])
    # NC LPC conditional routing: LCAS-prefix licenses belong to the NC Substance
    # Abuse Professional Practice Board (NC_DAC); all other LPC licenses belong to
    # the NC Mental Health Counselors board (NC_MENTAL_HEALTH). No secondary routing.
    if lic_state == "NC" and prov_type == "LPC":
        if re.sub(r"[^A-Z]", "", (license_id or "").upper()).startswith("LCAS"):
            preferred_sids = ["NC_DAC"]
        else:
            preferred_sids = ["NC_MENTAL_HEALTH"]
    if not preferred_sids:
        return "Fail", f"No board configured for prov_type '{prov_type}'", ""

    configs_to_try_ids = [s for s in preferred_sids if s in psv_browsers or s in api_configs]
    if not configs_to_try_ids:
        return "Fail", f"Config not loaded for boards: {preferred_sids}", ""

    last_fail_reason = f"No match found in {configs_to_try_ids}"
    # Fallback: exactly-1 first+last name match with no license match.
    # Only used after all boards fail — prevents short-circuiting a better match on a later board.
    _single_name_fallback: tuple | None = None

    for src_id in configs_to_try_ids:
        is_browser = src_id in psv_browsers
        psv_b = psv_browsers.get(src_id)
        api_cfg = api_configs.get(src_id)

        # Socrata type filter: narrow SoQL query to the correct profession/board category.
        # Only set when a verified mapping exists; omitted otherwise (unfiltered search works).
        _type_val = _SOCRATA_TYPE_MAP.get((src_id, prov_type))
        if _type_val:
            _cfg = all_configs.get(src_id)
            if _cfg and getattr(_cfg.identity, "provider_type_selector", None):
                _type_kwargs: dict = {"provider_type": _type_val}
            else:
                _type_kwargs = {"license_type": _type_val}
        else:
            _type_kwargs = {}

        # --- License normalization: prov_type-based prefix injection ---
        # Boards like NC_PT require a single-letter designation prefix before the number
        # (P=PT, A=PTA). Two cases handled:
        #   1. Bare numeric "12345" → "P12345"  (license_id[0].isdigit())
        #   2. Prov-type-prefixed "PT12345" → "P12345"  (input has full prov_type as prefix
        #      instead of the board's shorter letter code)
        _cfg_for_norm = all_configs.get(src_id)
        _pfx_map = (
            getattr(getattr(_cfg_for_norm, "search", None), "license_prov_type_prefix_map", None) or {}
        ) if _cfg_for_norm else {}
        if _pfx_map and license_id:
            _target_pfx = _pfx_map.get(prov_type.upper()) or _pfx_map.get(prov_type)
            if _target_pfx:
                _old_lic = license_id
                if license_id[0].isdigit():
                    # Case 1: bare numeric
                    license_id = f"{_target_pfx}{license_id}"
                elif license_id.upper().startswith(prov_type.upper()):
                    # Case 2: prov_type-prefixed (e.g. "PT12345" when board wants "P12345")
                    _rest = license_id[len(prov_type):]
                    if _rest and (_rest[0].isdigit() or _rest[0] in ("-", " ")):
                        license_id = f"{_target_pfx}{_rest.lstrip('- ')}"
                if license_id != _old_lic:
                    log.info("[%s] License prefix normalized: '%s' → '%s' (prov_type=%s)",
                             src_id, _old_lic, license_id, prov_type)

        # --- Pass 1: search by license number ---
        # Do NOT pass _type_kwargs here: the license number uniquely identifies
        # the record, and a board-level type filter causes misses when the input
        # prov_type differs from the actual license_type on the board (e.g. PN
        # input but APRN on board after a credential upgrade).
        q_lic = SearchQuery(
            mode="license_number",
            query=license_id,
            license_number=license_id,
            first_name=first_name or None,
            middle_name=middle_name or None,
            last_name=last_name or None,
        )
        if is_browser:
            records = await _try_search_psv(psv_b, q_lic, timeout)
        else:
            records = await _try_search_api(api_cfg, q_lic, timeout)

        lic_search_garbage = False  # set when license search returns unrelated records
        if records:
            both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, license_id)
            if both:
                expiry = _get_expiry(both[0])
                if not expiry and is_browser and psv_b:
                    expiry = await _fetch_detail_expiry(psv_b, both[0], timeout)
                _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                return "Pass", f"Verified via {src_id} (license search){_gn}", expiry
            if name_hits and not lic_hits:
                if license_id.upper().startswith("TC"):
                    return "Pass", f"Verified via {src_id} (TC temp cert — name match only)", _get_expiry(name_hits[0])
                # Exact name match in a license-number search: credential close but format differs.
                # Store as last-resort fallback (same logic as Pass 2 single-name fallback).
                if len(name_hits) == 1 and first_name and _single_name_fallback is None:
                    _single_name_fallback = (src_id, name_hits[0])
                last_fail_reason = f"name found but license not matched ({len(name_hits)} records) — {src_id}"
                continue
            if lic_hits and not name_hits:
                # Last-name-only check: first name may be a nickname/variant (e.g. Kathryn→Kathy)
                last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                if last_only:
                    return "Pass", f"Verified via {src_id} (license + last name match)", _get_expiry(last_only[0])
                # License found but name on board differs (e.g. married/maiden name change).
                # Trust the license number as the primary key when searched directly.
                return "Pass", f"Verified via {src_id} (license match — name on board differs)", _get_expiry(lic_hits[0])
            # Neither name nor license matched — board likely returned unrelated results
            # for an unrecognised license format. Fall through to last_name search.
            log.info("[%s] %d record(s) with no match at all — falling through to name search",
                     src_id, len(records))
            lic_search_garbage = True

        # --- Pass 1.5: strip non-digits and retry if license has a prefix/hyphens ---
        numeric_id = re.sub(r"\D", "", license_id)
        # Skip numeric-only retry when the license starts with alphabetic characters
        # fused directly to digits (no separator), e.g. "A01225" or "PA1728".
        # In those formats the letters are a meaningful credential-type prefix that
        # routes to a different board (e.g. MD_PSYCH_ASSOC for "A"-prefix licenses).
        # Stripping the prefix would match a different person on the wrong board.
        # Hyphenated formats like "PPC-1359" still get the retry (separator present).
        _alpha_prefix_fused = bool(re.match(r'^[A-Za-z]+\d', license_id)) and "-" not in license_id and "/" not in license_id
        if not lic_search_garbage and numeric_id and numeric_id != license_id and not _alpha_prefix_fused:
            log.info("[%s] No results for '%s', retrying with numeric-only: '%s'",
                     src_id, license_id, numeric_id)
            q_num = SearchQuery(
                mode="license_number",
                query=numeric_id,
                license_number=numeric_id,
                first_name=first_name or None,
                middle_name=middle_name or None,
                last_name=last_name or None,
                **_type_kwargs,
            )
            if is_browser:
                records = await _try_search_psv(psv_b, q_num, timeout)
            else:
                records = await _try_search_api(api_cfg, q_num, timeout)

            if records:
                both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, numeric_id)
                if both:
                    _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                    return "Pass", f"Verified via {src_id} (numeric license search){_gn}", _get_expiry(both[0])
                if name_hits and not lic_hits:
                    last_fail_reason = f"name found but numeric license not matched ({len(name_hits)} records) — {src_id}"
                    continue
                if lic_hits and not name_hits:
                    last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                    if last_only:
                        return "Pass", f"Verified via {src_id} (numeric license + last name match)", _get_expiry(last_only[0])
                    return "Pass", f"Verified via {src_id} (numeric license match — name on board differs)", _get_expiry(lic_hits[0])
                log.info("[%s] Numeric search also returned unrelated records — falling through to name search",
                         src_id)
                lic_search_garbage = True

        # --- Pass 1.51: strip trailing sequence/renewal suffix (e.g. "1380-3" → "1380") ---
        # Some boards store only the base license number without a trailing renewal suffix.
        # Pattern: all-digit prefix + hyphen + 1-2 trailing digits (e.g. "1380-3", "1373-3").
        # Confirmed for KS_OPTOMETRY where "-N" suffix is a renewal sequence, not part of the number.
        _seq_m = re.match(r'^(\d+)-(\d{1,2})$', license_id)
        if not lic_search_garbage and _seq_m:
            _base_id = _seq_m.group(1)
            if _base_id and _base_id != license_id and _base_id != numeric_id:
                log.info("[%s] Stripping sequence suffix: '%s' → '%s'", src_id, license_id, _base_id)
                q_base = SearchQuery(
                    mode="license_number",
                    query=_base_id,
                    license_number=_base_id,
                    first_name=first_name or None,
                    middle_name=middle_name or None,
                    last_name=last_name or None,
                    **_type_kwargs,
                )
                if is_browser:
                    records = await _try_search_psv(psv_b, q_base, timeout)
                else:
                    records = await _try_search_api(api_cfg, q_base, timeout)
                if records:
                    both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, _base_id)
                    if both:
                        _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                        return "Pass", f"Verified via {src_id} (base license — sequence suffix stripped){_gn}", _get_expiry(both[0])
                    if name_hits and not lic_hits:
                        last_fail_reason = (
                            f"name found but base license not matched ({len(name_hits)} records) — {src_id}")
                        continue
                    if lic_hits and not name_hits:
                        last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                        if last_only:
                            return "Pass", f"Verified via {src_id} (base license + last name)", _get_expiry(last_only[0])
                        return "Pass", f"Verified via {src_id} (base license — name on board differs)", _get_expiry(lic_hits[0])
                    log.info("[%s] Base-license search also returned unrelated records — falling through to name search",
                             src_id)
                    lic_search_garbage = True

        _board_cfg = (psv_b.config if psv_b else api_cfg)
        _dash_fmt = getattr(getattr(_board_cfg, "search", None), "license_dash_format", None)

        # --- Pass 1.53: extract middle group from already-dashed KSBN-style licenses ---
        # KSBN board note: "Enter the middle digits only (example: 00-000000-000)".
        # When input is "53-81920-022" or "5381920022", derive the center group ("81920")
        # and search with that alone. Only fires for 3-part dash formats (e.g. "2-5-3").
        if not lic_search_garbage and _dash_fmt:
            _fp = _dash_fmt.split("-")
            if len(_fp) == 3:
                try:
                    _fp0, _fp1 = int(_fp[0]), int(_fp[1])
                except ValueError:
                    _fp0 = _fp1 = None
                if _fp0 is not None and _fp1 is not None:
                    _raw153 = re.sub(r"\D", "", license_id)
                    if len(_raw153) >= _fp0 + _fp1:
                        _mid153 = _raw153[_fp0:_fp0 + _fp1]
                        # Skip if already just the middle group (e.g. input was "81920")
                        if _mid153 and _mid153 not in (license_id, _raw153):
                            log.info("[%s] Trying middle-group license: '%s'", src_id, _mid153)
                            q_mid153 = SearchQuery(
                                mode="license_number",
                                query=_mid153,
                                license_number=_mid153,
                                first_name=first_name or None,
                                middle_name=middle_name or None,
                                last_name=last_name or None,
                                **_type_kwargs,
                            )
                            records = await (_try_search_psv(psv_b, q_mid153, timeout) if is_browser
                                             else _try_search_api(api_cfg, q_mid153, timeout))
                            if records:
                                both, name_hits, lic_hits = _match_analysis(
                                    records, last_name, first_name, _mid153)
                                if both:
                                    expiry = _get_expiry(both[0])
                                    if not expiry and is_browser and psv_b:
                                        expiry = await _fetch_detail_expiry(psv_b, both[0], timeout)
                                    _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                                    return "Pass", f"Verified via {src_id} (middle-group license){_gn}", expiry
                                if name_hits and not lic_hits:
                                    last_fail_reason = (
                                        f"name found but middle-group license not matched — {src_id}")
                                if lic_hits:
                                    last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                                    if last_only:
                                        return "Pass", f"Verified via {src_id} (middle-group license + last name)", _get_expiry(last_only[0])
                                    return "Pass", f"Verified via {src_id} (middle-group license — name differs)", _get_expiry(lic_hits[0])

        # --- Pass 1.55: try dash-formatted license when board specifies license_dash_format ---
        # Handles pure-digit licenses that need grouping (e.g. "5383371052" → "53-83371-052" on KSBN).
        if not lic_search_garbage and _dash_fmt:
            _raw_digits = re.sub(r"\D", "", license_id)
            if _raw_digits and _raw_digits == license_id:  # pure-digit only (same guard as ladder)
                from orchestrator.ladder import _apply_dash_format
                _dash_id = _apply_dash_format(_raw_digits, _dash_fmt)
                if _dash_id and _dash_id != license_id:
                    log.info("[%s] Trying dash-formatted license: '%s'", src_id, _dash_id)
                    q_dash = SearchQuery(
                        mode="license_number",
                        query=_dash_id,
                        license_number=_dash_id,
                        first_name=first_name or None,
                        middle_name=middle_name or None,
                        last_name=last_name or None,
                        **_type_kwargs,
                    )
                    records = await (_try_search_psv(psv_b, q_dash, timeout) if is_browser
                                     else _try_search_api(api_cfg, q_dash, timeout))
                    if records:
                        both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, _dash_id)
                        if both:
                            _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                            return "Pass", f"Verified via {src_id} (dash-formatted license){_gn}", _get_expiry(both[0])
                        if name_hits and not lic_hits:
                            last_fail_reason = f"name found but dash-formatted license not matched — {src_id}"
                            continue
                        if lic_hits and not name_hits:
                            last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                            if last_only:
                                return "Pass", f"Verified via {src_id} (dash-formatted license + last name)", _get_expiry(last_only[0])
                            return "Pass", f"Verified via {src_id} (dash-formatted license match — name differs)", _get_expiry(lic_hits[0])
                        lic_search_garbage = True

        # --- Pass 1.56: derive KSBN license from IBCLC credential format ---
        # IBCLC credentials follow XX-XXXXXX-XXX (11 stripped digits, e.g. "14-126425-012").
        # For boards with license_dash_format (e.g. KSBN 2-5-3), the underlying nursing license
        # is sometimes the last 10 stripped digits formatted per the dash spec ("41-26425-012").
        # The KSBN site also accepts "middle digits only" — the 5-digit center group ("26425").
        # Tester note: "middle 5 digit from license#" = last 5 of the 6-digit IBCLC middle group.
        if not lic_search_garbage and _dash_fmt:
            _ibclc_m = re.match(r'^(\d{1,2})-(\d{6})-(\d{3})$', license_id)
            if _ibclc_m:
                _raw_11 = re.sub(r"\D", "", license_id)
                _derived_10 = _raw_11[-10:]
                from orchestrator.ladder import _apply_dash_format
                _derived_id = _apply_dash_format(_derived_10, _dash_fmt)
                # Also try the middle 5 digits alone (KSBN "middle digits only" search).
                _parts = _dash_fmt.split("-")
                _mid_len = int(_parts[1]) if len(_parts) == 3 else None
                _mid_start = int(_parts[0]) if len(_parts) == 3 else None
                _middle_only = (_derived_10[_mid_start:_mid_start + _mid_len]
                                if _mid_start is not None and _mid_len is not None else None)
                _ibclc_candidates = [
                    (_derived_id, f"last-10 → {_dash_fmt}"),
                    (_middle_only, "middle digits only"),
                ]
                _ibclc_any_records = False
                for _attempt_id, _attempt_label in _ibclc_candidates:
                    if not _attempt_id or _attempt_id == license_id:
                        continue
                    log.info("[%s] Trying IBCLC-derived license (%s): '%s'",
                             src_id, _attempt_label, _attempt_id)
                    q_deriv = SearchQuery(
                        mode="license_number",
                        query=_attempt_id,
                        license_number=_attempt_id,
                        first_name=first_name or None,
                        middle_name=middle_name or None,
                        last_name=last_name or None,
                        **_type_kwargs,
                    )
                    records = await (_try_search_psv(psv_b, q_deriv, timeout) if is_browser
                                     else _try_search_api(api_cfg, q_deriv, timeout))
                    if records:
                        _ibclc_any_records = True
                        both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, _attempt_id)
                        if both:
                            _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                            return "Pass", f"Verified via {src_id} (IBCLC-derived license){_gn}", _get_expiry(both[0])
                        if name_hits and not lic_hits:
                            last_fail_reason = f"name found but IBCLC-derived license not matched — {src_id}"
                        if lic_hits and not name_hits:
                            last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                            if last_only:
                                return "Pass", f"Verified via {src_id} (IBCLC-derived license + last name)", _get_expiry(last_only[0])
                            return "Pass", f"Verified via {src_id} (IBCLC-derived license — name differs)", _get_expiry(lic_hits[0])
                if not _ibclc_any_records:
                    lic_search_garbage = True

        # --- Pass 1.57: alpha-space-insert, digit-prefix, and digit-pad normalization ---
        # license_alpha_space_insert: "LCPC03720" → "LCPC 03720" (insert space after alpha prefix).
        # license_digit_prefixes: ["LCPC", "LPC"] for pure-digit inputs like "03192" → "LCPC 03192".
        # license_digit_pad: N — zero-pads digit portion to N digits:
        #   "D63352" → "D0063352" (pad=7); pure-digit with prefixes uses no space + padded digits.
        _alpha_space = getattr(getattr(_board_cfg, "search", None), "license_alpha_space_insert", False)
        _digit_pfxs = getattr(getattr(_board_cfg, "search", None), "license_digit_prefixes", None) or []
        _digit_pad = getattr(getattr(_board_cfg, "search", None), "license_digit_pad", None)
        if not lic_search_garbage and (_alpha_space or _digit_pfxs or _digit_pad):
            _p157_cands: list[tuple[str, str]] = []
            _m157_alpha = re.match(r'^([A-Za-z]+)(\d+)$', license_id)
            if _alpha_space and _m157_alpha:
                _p157_cands.append(
                    (f"{_m157_alpha.group(1)} {_m157_alpha.group(2)}", "alpha-space-insert"))
            # Zero-pad digit portion for inputs that already have a letter prefix (e.g. "D63352" → "D0063352")
            if _digit_pad and _m157_alpha:
                _pfx_letters = _m157_alpha.group(1)
                _pfx_digits = _m157_alpha.group(2)
                if len(_pfx_digits) < _digit_pad:
                    _p157_cands.append(
                        (f"{_pfx_letters}{_pfx_digits.zfill(_digit_pad)}", f"digit-pad-{_digit_pad}"))
            # Zero-pad purely numeric license (e.g. "1681" → "01681" with digit_pad=5)
            if _digit_pad and not _m157_alpha and re.match(r'^\d+$', license_id) and len(license_id) < _digit_pad:
                _p157_cands.append((license_id.zfill(_digit_pad), f"zfill-{_digit_pad}"))
            if _digit_pfxs and re.match(r'^\d+$', license_id):
                for _pfx157 in _digit_pfxs:
                    if _digit_pad:
                        # No space; zero-pad digit portion (e.g. MD: "D" + "63352".zfill(7) → "D0063352")
                        _p157_cands.append((f"{_pfx157}{license_id.zfill(_digit_pad)}", f"prefix-{_pfx157}-pad"))
                    else:
                        _p157_cands.append((f"{_pfx157} {license_id}", f"prefix-{_pfx157}"))
            # Case 4: mixed-format input (e.g. "CDRH.0071196") — a dot or other separator
            # breaks both the alpha-prefix and pure-digit regexes above.  Extract just the
            # digit run and apply the prefix+pad combos so "D0071196" etc. are still tried.
            if _digit_pad and _digit_pfxs and not _m157_alpha and not re.match(r'^\d+$', license_id):
                _mixed_digits = re.sub(r'\D', '', license_id)
                if _mixed_digits and len(_mixed_digits) <= _digit_pad:
                    for _pfx157 in _digit_pfxs:
                        _p157_cands.append(
                            (f"{_pfx157}{_mixed_digits.zfill(_digit_pad)}", f"prefix-{_pfx157}-pad"))
            for _cand157, _lbl157 in _p157_cands:
                if _cand157 == license_id:
                    continue
                log.info("[%s] Trying formatted license (%s): '%s'", src_id, _lbl157, _cand157)
                q157 = SearchQuery(
                    mode="license_number",
                    query=_cand157,
                    license_number=_cand157,
                    first_name=first_name or None,
                    middle_name=middle_name or None,
                    last_name=last_name or None,
                    **_type_kwargs,
                )
                records = await (_try_search_psv(psv_b, q157, timeout) if is_browser
                                 else _try_search_api(api_cfg, q157, timeout))
                if records:
                    both, name_hits, lic_hits = _match_analysis(
                        records, last_name, first_name, _cand157)
                    if both:
                        expiry = _get_expiry(both[0])
                        if not expiry and is_browser and psv_b:
                            expiry = await _fetch_detail_expiry(psv_b, both[0], timeout)
                        _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                        return "Pass", f"Verified via {src_id} ({_lbl157}){_gn}", expiry
                    if name_hits and not lic_hits:
                        last_fail_reason = f"name found but {_lbl157} license not matched — {src_id}"
                        continue
                    if lic_hits:
                        last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                        if last_only:
                            return "Pass", f"Verified via {src_id} ({_lbl157} + last name)", _get_expiry(last_only[0])
                        return "Pass", f"Verified via {src_id} ({_lbl157} — name differs)", _get_expiry(lic_hits[0])

        # --- Pass 1.6: try prefix-dash license when board specifies license_prefix_dash ---
        # Handles licenses like "L301745" → "L-301745" for IBCLC_COMMISSION.
        _prefix_dash = getattr(getattr(_board_cfg, "search", None), "license_prefix_dash", False)
        if not lic_search_garbage and _prefix_dash:
            _m = re.match(r'^([A-Za-z]+)(\d+)$', license_id)
            if _m:
                _prefixed_id = f"{_m.group(1)}-{_m.group(2)}"
                log.info("[%s] Trying prefix-dash license: '%s'", src_id, _prefixed_id)
                q_pfx = SearchQuery(
                    mode="license_number",
                    query=_prefixed_id,
                    license_number=_prefixed_id,
                    first_name=first_name or None,
                    middle_name=middle_name or None,
                    last_name=last_name or None,
                    **_type_kwargs,
                )
                records = await (_try_search_psv(psv_b, q_pfx, timeout) if is_browser
                                 else _try_search_api(api_cfg, q_pfx, timeout))
                if records:
                    both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, _prefixed_id)
                    if both:
                        _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                        return "Pass", f"Verified via {src_id} (prefix-dash license){_gn}", _get_expiry(both[0])
                    if name_hits and not lic_hits:
                        last_fail_reason = f"name found but prefix-dash license not matched — {src_id}"
                        continue
                    if lic_hits and not name_hits:
                        last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                        if last_only:
                            return "Pass", f"Verified via {src_id} (prefix-dash license + last name)", _get_expiry(last_only[0])
                        return "Pass", f"Verified via {src_id} (prefix-dash license match — name differs)", _get_expiry(lic_hits[0])
                    lic_search_garbage = True

        # --- Pass 1.7: try license_and_last combo when board supports it ---
        # Needed for boards like IBCLC_COMMISSION where credential_number is a separate field
        # from the name search field; license_number mode incorrectly fills the name field.
        # Try each candidate license format (raw, prefix-dashed) paired with last_name.
        if is_browser and last_name and not lic_search_garbage:
            _has_lic_and_last = any(
                getattr(m, "mode", None) == "license_and_last"
                for m in (getattr(getattr(psv_b.config, "search", None), "modes", None) or [])
            )
            if _has_lic_and_last:
                _lic_candidates = [license_id]
                _m17 = re.match(r'^([A-Za-z]+)(\d+)$', license_id)
                if _m17:
                    _lic_candidates.append(f"{_m17.group(1)}-{_m17.group(2)}")
                for _lic_cand in _lic_candidates:
                    log.info("[%s] Trying license_and_last combo: name=%s cred=%s", src_id, last_name, _lic_cand)
                    q_combo = SearchQuery(
                        mode="license_and_last",
                        query=f"{last_name} {_lic_cand}",
                        license_number=_lic_cand,
                        first_name=first_name or None,
                        middle_name=middle_name or None,
                        last_name=last_name or None,
                        **_type_kwargs,
                    )
                    records = await _try_search_psv(psv_b, q_combo, timeout)
                    if records:
                        both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, _lic_cand)
                        if both:
                            _gn = _epdb_name_gate_note(both[0], last_name, first_name)
                            return "Pass", f"Verified via {src_id} (license_and_last combo){_gn}", _get_expiry(both[0])
                        if lic_hits and not name_hits:
                            last_only = [r for r in lic_hits if _name_matches(r, last_name, "")]
                            if last_only:
                                return "Pass", f"Verified via {src_id} (license_and_last + last name)", _get_expiry(last_only[0])
                            return "Pass", f"Verified via {src_id} (license_and_last — name differs)", _get_expiry(lic_hits[0])

        # --- Pass 2: fall back to last_name search if license searches returned nothing ---
        if not last_name:
            continue
        log.info("[%s] License search empty, trying last_name fallback for %s", src_id, license_id)
        # Pass 2 is a genuine name-only search: omit license_number so the Socrata
        # combo URL doesn't repeat the same AND clause that already returned 0 results.
        # _type_kwargs (board-level type filter) is kept to narrow name results.
        q_name = SearchQuery(
            mode="last_name",
            query=last_name,
            first_name=first_name or None,
            middle_name=middle_name or None,
            last_name=last_name or None,
            **_type_kwargs,
        )

        if is_browser:
            records = await _try_search_psv(psv_b, q_name, timeout)
        else:
            records = await _try_search_api(api_cfg, q_name, timeout)

        _do_swap = is_browser and bool(getattr(
            (psv_b.config.identity if psv_b else None), "try_name_swap", False))

        if not records:
            # Pass 2.5a: last_name returned nothing — try first_and_last if board supports it
            if is_browser and first_name:
                fal = await _try_first_and_last_psv(
                    psv_b, first_name, last_name, middle_name, license_id, src_id, timeout,
                    type_kwargs=_type_kwargs, try_swap=_do_swap,
                )
                if fal:
                    return fal
            # Pass 2.5a2: compound last name with space — retry with each token.
            # Try last token first (e.g. "Campos Papaioannou" → "Papaioannou"), then first
            # token (e.g. "Marvel Massey" → "Marvel") for boards that index only one component.
            if is_browser and " " in last_name:
                _first_token = last_name.split(" ", 1)[0]
                _last_token = last_name.rsplit(" ", 1)[-1]
                for _tok_label, _search_tok in [("last token", _last_token), ("first token", _first_token)]:
                    log.info("[%s] Compound last name — retrying with %s '%s'",
                             src_id, _tok_label, _search_tok)
                    _q_token = SearchQuery(
                        mode="last_name",
                        query=_search_tok,
                        license_number=license_id,
                        first_name=first_name or None,
                        middle_name=middle_name or None,
                        last_name=last_name or None,
                        **_type_kwargs,
                    )
                    _records_tok = await _try_search_psv(psv_b, _q_token, timeout)
                    if _records_tok:
                        # Try full compound name match first; fall back to single-token match
                        _both_t, _, _ = _match_analysis(_records_tok, last_name, first_name, license_id)
                        if not _both_t:
                            _both_t, _, _ = _match_analysis(_records_tok, _search_tok, first_name, license_id)
                        if _both_t:
                            _exp_t = _get_expiry(_both_t[0])
                            if not _exp_t:
                                _exp_t = await _fetch_detail_expiry(psv_b, _both_t[0], timeout)
                            _gn_t = _epdb_name_gate_note(_both_t[0], last_name, first_name)
                            return "Pass", f"Verified via {src_id} (compound last name — {_tok_label}){_gn_t}", _exp_t
            # Pass 2.6: all type-filtered name searches returned 0 — try fallback license types.
            # Handles career progression (e.g. PN/LPN→RN): the board reclassified the person
            # under a different license type while the input still shows the original prov_type.
            _fallback_types = _BOARD_TYPE_FALLBACKS.get((src_id, prov_type), [])
            if is_browser and _type_kwargs and _fallback_types and last_name:
                for _fb_type in _fallback_types:
                    log.info("[%s] Type-filtered name search empty — retrying with licenseType=%s",
                             src_id, _fb_type)
                    _q_fb = SearchQuery(
                        mode="last_name",
                        query=last_name,
                        license_number=license_id,
                        first_name=first_name or None,
                        middle_name=middle_name or None,
                        last_name=last_name or None,
                        license_type=_fb_type,
                    )
                    _recs_fb = await _try_search_psv(psv_b, _q_fb, timeout)
                    if _recs_fb:
                        _both_fb, _nm_fb, _lh_fb = _match_analysis(
                            _recs_fb, last_name, first_name, license_id)
                        if _both_fb:
                            _exp_fb = _get_expiry(_both_fb[0])
                            if not _exp_fb:
                                _exp_fb = await _fetch_detail_expiry(psv_b, _both_fb[0], timeout)
                            _gn_fb = _epdb_name_gate_note(_both_fb[0], last_name, first_name)
                            return "Pass", (
                                f"Verified via {src_id} (fallback licenseType={_fb_type}){_gn_fb}"
                            ), _exp_fb
                        if _nm_fb and not _lh_fb and len(_nm_fb) == 1 and first_name:
                            # Name match but license class differs — store as single-name fallback
                            if _single_name_fallback is None:
                                _single_name_fallback = (src_id, _nm_fb[0])
                        break  # stop after first fallback type that returns any records
            continue

        both, name_hits, lic_hits = _match_analysis(records, last_name, first_name, license_id)
        if both:
            expiry = _get_expiry(both[0])
            if not expiry and is_browser and psv_b:
                expiry = await _fetch_detail_expiry(psv_b, both[0], timeout)
            _gn = _epdb_name_gate_note(both[0], last_name, first_name)
            return "Pass", f"Verified via {src_id} (name search){_gn}", expiry
        if name_hits and not lic_hits:
            if license_id.upper().startswith("TC"):
                return "Pass", f"Verified via {src_id} (TC temp cert — name match only)", _get_expiry(name_hits[0])
            # Track as last-resort fallback when exactly 1 first+last match — PSV may store
            # a license from a different class (e.g. TL- telemedicine, COM- combined) that
            # does not appear in this board's search type.
            if len(name_hits) == 1 and first_name and _single_name_fallback is None:
                _single_name_fallback = (src_id, name_hits[0])
            # Pass 2.5b: name found but license not matched and results are truncated
            if is_browser and len(records) >= 20 and first_name:
                fal = await _try_first_and_last_psv(
                    psv_b, first_name, last_name, middle_name, license_id, src_id, timeout,
                    type_kwargs=_type_kwargs, try_swap=_do_swap,
                )
                if fal:
                    return fal
            last_fail_reason = f"name found but license not matched in name search ({len(name_hits)} records) — {src_id}"
            continue
        if lic_hits and not name_hits:
            last_fail_reason = f"license matched but name mismatch in name search — {src_id}"
            continue
        # Pass 2.5c: some records returned but nothing matched — try first_and_last when truncated
        if is_browser and len(records) >= 20 and first_name:
            fal = await _try_first_and_last_psv(
                psv_b, first_name, last_name, middle_name, license_id, src_id, timeout,
                type_kwargs=_type_kwargs, try_swap=_do_swap,
            )
            if fal:
                return fal
        # Pass 2.5d: small result set — try last-name-only + license for first-name spelling variants
        # (e.g. "Dierdre" vs "Deirdre").  Only safe when results are few (≤5) to avoid false positives.
        if len(records) <= 5:
            last_lic = [r for r in records if _license_matches(r, license_id) and _name_matches(r, last_name, "")]
            if last_lic:
                expiry = _get_expiry(last_lic[0])
                if not expiry and is_browser and psv_b:
                    expiry = await _fetch_detail_expiry(psv_b, last_lic[0], timeout)
                return "Pass", f"Verified via {src_id} (last name + license match)", expiry
        last_fail_reason = f"{len(records)} record(s) from name search but no license+name match — {src_id}"

    if _single_name_fallback:
        s_id, s_rec = _single_name_fallback
        expiry = _get_expiry(s_rec)
        if not expiry and s_id in psv_browsers:
            expiry = await _fetch_detail_expiry(psv_browsers[s_id], s_rec, timeout)
        return "Pass", f"Verified via {s_id} (name match — PSV license class not in this search type)", expiry
    # --- PSYPACT secondary check for NC CP ---
    # NC Clinical Psychologists may hold PSYPACT E.Passports. When the primary
    # NC board ladder fails, search the PSYPACT national registry as a fallback.
    if lic_state == "NC" and prov_type == "CP":
        _trace_stub = types.SimpleNamespace(final_outcome=None, final_reason=None)
        _psypact_lr = await _verify_psypact_row(row_data, _trace_stub)
        if _psypact_lr.status == "Pass" and getattr(_psypact_lr, "best_record", None):
            _exp = getattr(_psypact_lr.best_record, "expiration_date", None)
            _exp_str = _exp.isoformat() if hasattr(_exp, "isoformat") else (str(_exp) if _exp else "")
            return "Pass", "Verified via PSYPACT (secondary — NC CP)", _exp_str
    # Post-scrape N/A override: board was visited but state does not license this type.
    if _post_na := NA_PROV_TYPES.get((lic_state, prov_type)):
        return "N/A", _post_na, ""
    return "Fail", last_fail_reason, ""


def _write_remove_license(rows: list[dict], output_dir: Path) -> None:
    """Write rows whose License State is not in Service Location State to
    <output_dir>/RemoveLicense/RemoveLicense.xlsx.

    rows may be pre-filtered (orchestrated path) or full result dicts
    (legacy path — filtered internally by status + reason).
    """
    # Legacy path: rows are full result dicts (have "status"/"reason" keys).
    # Orchestrated path: rows are raw master_row dicts, already pre-filtered.
    if rows and "status" in rows[0]:
        svc_na_rows = [
            r for r in rows
            if r.get("status") == "N/A"
            and "not in Service Location State" in r.get("reason", "")
        ]
    else:
        svc_na_rows = rows

    if not svc_na_rows:
        return

    remove_dir = output_dir / "RemoveLicense"
    remove_dir.mkdir(parents=True, exist_ok=True)
    remove_path = remove_dir / "RemoveLicense.xlsx"

    if remove_path.exists():
        wb = openpyxl.load_workbook(str(remove_path))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RemoveLicense"
        ws.append(["PIN", "State", "LicenseNumber", "LicenseType"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for r in svc_na_rows:
        ws.append([
            r.get("epdb_pin", ""),
            r.get("lic_state", ""),
            r.get("license_id", ""),
            r.get("lic_type", ""),
        ])

    wb.save(str(remove_path))
    log.info("RemoveLicense: wrote %d row(s) to %s", len(svc_na_rows), remove_path)


def write_results(results: list[dict], output_path: Path, append: bool) -> None:
    if append and output_path.exists():
        wb = openpyxl.load_workbook(str(output_path))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["First Name", "Middle Name", "Last Name",
                   "License State", "Prov Type", "License Type", "License ID",
                   "Status", "License Expiry", "Reason"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")

    for r in results:
        ws.append([
            r["first_name"], r["middle_name"], r["last_name"],
            r["lic_state"], r["prov_type"], r["lic_type"], r["license_id"],
            r["status"], r.get("expiry_date", ""),
            r["reason"] if r["status"] in ("Fail", "N/A") else "",
        ])
        if r["status"] == "Pass":
            fill = green
        elif r["status"] == "N/A":
            fill = yellow
        else:
            fill = red
        for col in range(1, 11):
            ws.cell(row=ws.max_row, column=col).fill = fill

    wb.save(str(output_path))
    _write_remove_license(results, output_path.parent)


def _cell_to_iso_date(val) -> str:
    """Convert an openpyxl cell value (datetime, date, or string) to YYYY-MM-DD, or ''."""
    if val is None:
        return ""
    if hasattr(val, "date"):          # datetime.datetime
        return val.date().isoformat()
    if hasattr(val, "isoformat"):     # datetime.date
        return val.isoformat()
    s = str(val).strip()
    if not s or s.lower() == "none":
        return ""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return ""   # unparseable — treat as absent


def load_input_rows(input_path: str, state_filter: str, sheet_name: str = "") -> list[dict]:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb["PSV Tab"] if "PSV Tab" in wb.sheetnames else wb.active

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        return []

    start = 0
    first = all_rows[0]
    v = str(first[C_LIC_STATE] or "").strip().upper() if len(first) > C_LIC_STATE else ""
    if v in ("LIC_STATE", "C_LIC_STATE", "LICENSE STATE", "STATE", "LICENSE STATE", ""):
        start = 1

    # Discover NPI_NO column index by header name (dynamic — may be absent)
    npi_col_idx: Optional[int] = None
    if start == 1:
        for i, cell in enumerate(first):
            if cell is None:
                continue
            header = str(cell).strip().upper().replace("_", " ").replace(".", "")
            for alias in _NPI_HEADER_ALIASES:
                if header == alias.upper().replace("_", " "):
                    npi_col_idx = i
                    break
            if npi_col_idx is not None:
                break

    rows_data = []
    for row in all_rows[start:]:
        if not row or all(v is None for v in row):
            continue

        def c(idx):
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None and str(val) != "None" else ""

        lic_state = c(C_LIC_STATE)
        if state_filter:
            allowed = {s.strip().upper() for s in state_filter.split(",")}
            if lic_state.upper() not in allowed:
                continue

        npi_val = c(npi_col_idx) if npi_col_idx is not None else ""
        # Strip non-digits and validate as 10-digit NPI.
        # Excel stores integer NPI values as floats ("1234567890.0") — convert to int
        # string first so the trailing ".0" doesn't inflate the digit count to 11.
        import re as _re
        try:
            _f = float(npi_val)
            if _f == int(_f):
                npi_val = str(int(_f))
        except (ValueError, TypeError):
            pass
        npi_clean = _re.sub(r"\D", "", npi_val)
        if len(npi_clean) != 10:
            npi_clean = ""

        rows_data.append({
            "first_name": c(C_FIRST_NAME),
            "middle_name": c(C_MIDDLE_NAME),
            "last_name": c(C_LAST_NAME),
            "epdb_pin": c(C_EPDB_PIN),
            "prov_type": c(C_PROV_TYPE),
            "maintained_by": c(C_MAINTAINED_BY),
            "npi_no": npi_clean,
            "lic_state": lic_state,
            "lic_type": c(C_LIC_TYPE),
            "license_id": c(C_LIC_ID).rstrip("_ "),
            "input_expiry": _cell_to_iso_date(
                row[C_LIC_EXPIRY] if C_LIC_EXPIRY < len(row) else None
            ),
            "svc_loc_state": c(C_SVC_LOC_STATE),
        })
    return rows_data


_SKIP_REASON_BY_SID: dict[str, str] = {}

_CAPTCHA_SKIP_KEYWORDS = ("captcha", "mcafee", "datadome", "cloudflare", "recaptcha", "waf", "blocked")


def _is_captcha_skip(reason: str) -> bool:
    return any(kw in reason.lower() for kw in _CAPTCHA_SKIP_KEYWORDS)


# BACB certification number format: TYPE_CODE(-LEVEL)-YY-NNNNNN
# TYPE_CODE: RBT, BCBA, BCaBA (or BCABA), 1 (=BCBA), 0 (=BCaBA)
# BCBA\d* handles variants like "BCBA1-24-70814" where a digit is appended to BCBA.
# BCaBA with IGNORECASE also matches BCABA (e.g. "BCABA-0-24-15058" 4-segment format
# where an extra level/type segment sits between the type code and the year).
# The bare YY-NNNNN alternative catches certs stored without a type-code prefix
# (e.g. "23-261946") — safe because this check is only called for prov_type=="ABA"
# and Indiana/IL/NC/NJ state ABA licenses use pure-numeric formats (no hyphens).
_BACB_LICENSE_RE = re.compile(
    r"^(RBT|BCBA\d*|BCaBA|1|0)(-\d{1,4}){1,2}-\d{4,7}$"  # TYPE(-LEVEL)-YY-NNNNN
    r"|^\d{2}-\d{4,7}$",  # bare YY-NNNNN (cert stored without type prefix, ABA-only)
    re.IGNORECASE,
)

def _is_bacb_license(license_id: str) -> bool:
    """Return True when the license ID matches BACB certification number format.

    Supported formats:
      RBT-YY-NNNNNN, BCBA-YY-NNNNNN, BCBA1-YY-NNNNNN,
      BCaBA-YY-NNNNNN (also BCABA, case-insensitive),
      BCaBA-LEVEL-YY-NNNNNN (4-segment, e.g. BCABA-0-24-15058),
      1-YY-NNNNNN (BCBA), 0-YY-NNNNNN (BCaBA),
      YY-NNNNNN (bare, no type prefix, e.g. 23-261946).
    """
    return bool(_BACB_LICENSE_RE.match((license_id or "").strip()))


# States where ABA is licensed via a state board AND BACB (captcha-blocked fallback).
# FL is excluded — FL ABA routes to BACB only (no state board).
_BACB_FALLBACK_STATES: frozenset[str] = frozenset({"IL", "IN", "NC", "NJ"})


def load_configs_by_source_ids(source_ids: set[str]) -> list:
    """Load configs for a specific set of source_ids (from routing table). Skips missing configs."""
    sites_dir = Path(__file__).parent / "sites"
    configs = []
    for sid in sorted(source_ids):
        config_path = sites_dir / sid / "config.yaml"
        if not config_path.exists():
            log.warning("No config.yaml for routed board %s — rows needing it will fail", sid)
            continue
        try:
            cfg = load_config(str(config_path))
            if getattr(cfg.identity, "skip", False):
                reason = getattr(cfg.identity, "skip_reason", "skip=true in config")
                _SKIP_REASON_BY_SID[sid] = reason
                log.info("Skipping board %s (skip=true): %s", sid, reason)
                continue
            configs.append(cfg)
            log.info("Loaded config: %s", sid)
        except Exception as exc:
            log.warning("Failed to load config %s: %s", sid, exc)
    return configs


def _board_proxy(cfg, resolved_proxy_cfg: Optional[dict]) -> Optional[dict]:
    """Return the proxy dict to use for a specific board config.

    Rules (mirrors the config.yaml proxy.enabled semantics):
      enabled: false  → always None  (board explicitly blocks proxy, e.g. NH_OPLC)
      enabled: true   → resolved_proxy_cfg (board requires proxy — comes from env or psv_config.yaml)
      enabled: None   → resolved_proxy_cfg if available, None otherwise (follow env)
    """
    if cfg.transport.proxy.enabled is False:
        return None
    return resolved_proxy_cfg


def _log_proxy_plan(state: str, configs: list, resolved_proxy_cfg: Optional[dict]) -> None:
    """Emit a single INFO line summarising which boards use proxy and which don't."""
    using_proxy = []
    no_proxy = []
    proxy_required_but_missing = []

    for cfg in configs:
        sid = cfg.identity.source_id
        enabled = cfg.transport.proxy.enabled
        if enabled is False:
            no_proxy.append(sid)
        elif enabled is True:
            if resolved_proxy_cfg:
                using_proxy.append(sid)
            else:
                proxy_required_but_missing.append(sid)
        else:
            # None — optional, use if available
            if resolved_proxy_cfg:
                using_proxy.append(sid)
            else:
                no_proxy.append(sid)

    server = resolved_proxy_cfg.get("server") if resolved_proxy_cfg else None
    if using_proxy:
        log.info("[%s] Proxy ON  (%s): %s", state, server, using_proxy)
    if no_proxy:
        log.info("[%s] Proxy OFF (config): %s", state, no_proxy)
    if proxy_required_but_missing:
        log.warning(
            "[%s] Boards require proxy but none is configured — they will likely fail: %s. "
            "Set PROXY=proxy:9119 or add proxy.server to psv_config.yaml.",
            state, proxy_required_but_missing,
        )


async def run_state(
    rows: list[dict],
    state: str,
    output_path: Path,
    append: bool = False,
    batch_size: int = 10,
    timeout: int = 45,
    sequential: bool = False,
) -> tuple[int, int]:
    """Run PSV verification for one state. Returns (passes, fails).

    Caller must have already called _load_routing() before invoking this.
    The routing table (_ROUTING) must be populated.
    """
    # --- Fail CAPTCHA-blocked prov_type rows immediately ---
    captcha_rows = [r for r in rows if (state, r["prov_type"].upper()) in CAPTCHA_PROV_TYPES]
    rows = [r for r in rows if (state, r["prov_type"].upper()) not in CAPTCHA_PROV_TYPES]
    fails = 0
    skips = 0
    if captcha_rows:
        log.warning(
            "[%s] %d row(s) with CAPTCHA-blocked prov_types written as Skip: %s",
            state, len(captcha_rows),
            sorted({r["prov_type"] for r in captcha_rows}),
        )
        captcha_results = [
            {**r, "status": "Skip",
             "reason": CAPTCHA_PROV_TYPES[(state, r["prov_type"].upper())],
             "expiry_date": ""}
            for r in captcha_rows
        ]
        write_results(captcha_results, output_path, append)
        append = True
        skips += len(captcha_rows)
    if not rows:
        return 0, fails, skips

    # Determine which boards are needed for the prov_types in this batch
    needed_sids: set[str] = set()
    no_routing_combos: set[tuple[str, str]] = set()
    for row in rows:
        key = (row["lic_state"].upper(), row["prov_type"].upper())
        sids = _ROUTING.get(key, [])
        if sids:
            needed_sids.update(sids)
        else:
            no_routing_combos.add(key)

    if no_routing_combos:
        log.warning(
            "[%s] %d (state,prov_type) combo(s) have no board routing — those rows will Fail: %s",
            state, len(no_routing_combos), sorted(no_routing_combos),
        )

    if not needed_sids:
        log.warning("[%s] No boards in routing table for any row — all %d rows will Fail", state, len(rows))
        results = [{**r, "status": "Fail", "reason": f"No board configured for prov_type '{r['prov_type']}'",
                    "expiry_date": ""} for r in rows]
        write_results(results, output_path, append)
        return 0, len(rows)

    log.info("[%s] Routing requires %d board(s): %s", state, len(needed_sids), sorted(needed_sids))
    configs = load_configs_by_source_ids(needed_sids)
    if not configs:
        log.error("[%s] No board configs could be loaded — check sites/<board>/config.yaml", state)
        results = [{**r, "status": "Fail", "reason": "Board config not found", "expiry_date": ""} for r in rows]
        write_results(results, output_path, append)
        return 0, len(rows)

    log.info("[%s] Loaded %d board config(s): %s", state,
             len(configs), [c.identity.source_id for c in configs])

    browser_configs = [c for c in configs if c.identity.archetype in _BROWSER_ARCHETYPES]
    api_configs_list = [c for c in configs if c.identity.archetype not in _BROWSER_ARCHETYPES]
    api_configs = {c.identity.source_id: c for c in api_configs_list}
    all_configs_by_sid = {c.identity.source_id: c for c in configs}

    log.info("[%s] Browser boards (%d): %s  API boards (%d): %s", state,
             len(browser_configs), [c.identity.source_id for c in browser_configs],
             len(api_configs_list), [c.identity.source_id for c in api_configs_list])

    # --- Proxy diagnostics ---
    # Resolve proxy once; boards with proxy.enabled: false will override to None.
    proxy_cfg = get_proxy_config()
    _log_proxy_plan(state, configs, proxy_cfg)

    total = len(rows)
    passes = fails = skips = 0

    async with async_playwright() as pw:
        browser: Browser | None = None
        psv_browsers: dict = {}

        if browser_configs:
            log.info("[%s] Launching shared browser ...", state)
            # Respect transport settings from configs — non-headless/channel wins for any board
            _headless = all(cfg.transport.headless for cfg in browser_configs)
            _channel = next((cfg.transport.channel for cfg in browser_configs if cfg.transport.channel), None)
            _launch_kw: dict = {"headless": _headless, "args": _STEALTH_ARGS}
            if _channel:
                _launch_kw["channel"] = _channel
            browser = await pw.chromium.launch(**_launch_kw)
            for cfg in browser_configs:
                board_proxy = _board_proxy(cfg, proxy_cfg)
                psv_browsers[cfg.identity.source_id] = PsvBrowser(cfg, browser, board_proxy)
            log.info("[%s] Browser ready. Boards: %s", state, list(psv_browsers.keys()))

        try:
            for batch_start in range(0, total, batch_size):
                batch = rows[batch_start: batch_start + batch_size]
                batch_num = batch_start // batch_size + 1
                total_batches = (total + batch_size - 1) // batch_size
                log.info("[%s] === Batch %d/%d (rows %d-%d of %d) ===",
                         state, batch_num, total_batches,
                         batch_start + 1, batch_start + len(batch), total)

                if sequential:
                    outcomes = []
                    for r in batch:
                        outcomes.append(await run_row(r, psv_browsers, api_configs, all_configs_by_sid, timeout))
                else:
                    outcomes = list(await asyncio.gather(*[
                        run_row(r, psv_browsers, api_configs, all_configs_by_sid, timeout)
                        for r in batch
                    ]))

                batch_results = []
                for row, (status, reason, expiry_date) in zip(batch, outcomes):
                    log.info("  [%s] %s %s %s %s → %s | %s%s",
                             row["lic_state"], row["prov_type"], row["last_name"], row["first_name"],
                             row["license_id"], status,
                             f"expiry={expiry_date} | " if expiry_date else "", reason)
                    if status == "Pass":
                        passes += 1
                    else:
                        fails += 1
                    batch_results.append({**row, "status": status, "reason": reason,
                                          "expiry_date": expiry_date})

                write_results(batch_results, output_path, append)
                append = True
                log.info("[%s] Saved batch. Running totals: %d Pass / %d Fail / %d done of %d",
                         state, passes, fails, batch_start + len(batch), total)

        finally:
            if browser:
                log.info("[%s] Closing shared browser...", state)
                await browser.close()

    log.info("[%s] State complete: %d Pass / %d Fail / %d Total", state, passes, fails, total)
    return passes, fails, 0


async def run_state_orchestrated(
    rows: list[dict],
    state: str,
    emitter,                      # orchestrator.output_emitter.OutputEmitter
    run_id: str,
    enable_nppes: bool = True,
    enable_ai: bool = True,
    force_ai: bool = False,
    timeout: int = 45,
) -> tuple[int, int]:
    """Orchestrated per-state run. For every row:
       1. NPPES universal fetch
       2. Rule-based ladder (master)
       3. NPPES targeted retry (if discrepancy)
       4. AI agent (when ladders escalate, unless --no-ai)
       5. Emit to 4 channels via the shared OutputEmitter

    The emitter accumulates rows across states for one combined per-channel
    file at flush time.
    """
    from orchestrator import ladder as ladder_mod
    from orchestrator import nppes_client as nppes_mod
    from orchestrator import ai_agent as ai_mod
    from orchestrator import disambiguator as disamb_mod
    from orchestrator.output_emitter import RowOutcome
    from orchestrator.trace import RowTrace, make_master_row_id, REASON_PROVIDER_TYPE_MISMATCH

    # Reset the AI circuit breaker so a previous run's connection error doesn't
    # carry over and block the entire current run.
    ai_mod.reset_circuit_breaker()

    if not _ROUTING:
        _load_routing()

    # Resolve routing for every row
    needed_sids: set[str] = set()
    for row in rows:
        key = (row["lic_state"].upper(), row["prov_type"].upper())
        for sid in _ROUTING.get(key, []):
            needed_sids.add(sid)

    configs_list = load_configs_by_source_ids(needed_sids) if needed_sids else []
    cfg_by_sid = {c.identity.source_id: c for c in configs_list}

    browser_configs = [c for c in configs_list if c.identity.archetype in _BROWSER_ARCHETYPES]
    api_configs_list = [c for c in configs_list if c.identity.archetype not in _BROWSER_ARCHETYPES]
    api_cfg_by_sid = {c.identity.source_id: c for c in api_configs_list}

    proxy_cfg = get_proxy_config()
    _log_proxy_plan(state, configs_list, proxy_cfg)

    passes = fails = skips = 0
    _svc_loc_na_rows: list[dict] = []

    async with async_playwright() as pw:
        browser = None
        psv_browsers: dict = {}
        if browser_configs:
            log.info("[%s] Launching shared browser ...", state)
            _headless = all(cfg.transport.headless for cfg in browser_configs)
            _channel = next((cfg.transport.channel for cfg in browser_configs if cfg.transport.channel), None)
            _launch_kw2: dict = {"headless": _headless, "args": _STEALTH_ARGS}
            if _channel:
                _launch_kw2["channel"] = _channel
            browser = await pw.chromium.launch(**_launch_kw2)
            for cfg_obj in browser_configs:
                board_proxy = _board_proxy(cfg_obj, proxy_cfg)
                psv_browsers[cfg_obj.identity.source_id] = PsvBrowser(cfg_obj, browser, board_proxy)

        async def executor(cfg_obj, query, run_id_arg):
            """SearchExecutor: route browser boards through PsvBrowser, others
            through verify_license."""
            sid = cfg_obj.identity.source_id
            eff_t = cfg_obj.transport.ladder_timeout_s or timeout
            if sid in psv_browsers:
                return await psv_browsers[sid].search(
                    query, timeout_ms=eff_t * 1000, run_id=run_id_arg,
                )
            return await asyncio.wait_for(
                verify_license(cfg_obj, query, db=None),
                timeout=float(eff_t),
            )

        try:
            for idx, row in enumerate(rows):
                master_row_id = make_master_row_id(idx, row.get("npi_no", ""))
                trace = RowTrace(
                    master_row_id=master_row_id,
                    run_id=run_id,
                    state=state,
                    prov_type=row.get("prov_type", ""),
                    npi_no=row.get("npi_no", ""),
                )

                # --- NPPES universal fetch ---
                nppes = None
                discrepancy = None
                if enable_nppes and row.get("npi_no"):
                    try:
                        nppes = await nppes_mod.fetch_provider(row["npi_no"])
                    except Exception as exc:
                        log.warning("[%s] NPPES fetch failed for npi=%s: %s",
                                    state, row.get("npi_no"), exc)
                    if nppes:
                        discrepancy = nppes_mod.diff_master_vs_nppes(row, nppes)
                        trace.nppes_discrepancy = discrepancy.to_dict()

                # --- PSYPACT E.Passport: verify via directory.psypact.gov ---
                # Recognise both "PSYPACT18696" and "APIT-18240" (Authority to
                # Practice Interjurisdictional Telepsychology) license formats.
                _lic_upper = (row.get("license_id") or "").upper()
                if _lic_upper.startswith("PSYPACT") or _lic_upper.startswith("APIT"):
                    ladder_result = await _verify_psypact_row(row, trace, run_id=run_id)
                    ai_result = None
                    outcome = RowOutcome(
                        master_row=row,
                        master_row_id=master_row_id,
                        trace=trace,
                        nppes=nppes,
                        discrepancy=discrepancy,
                        ladder_result=ladder_result,
                        ai_result=ai_result,
                    )
                    emitter.collect(outcome)
                    if outcome.status == "Pass":
                        passes += 1
                    else:
                        fails += 1
                    log.info("[%s] %s %s %s %s -> %s | psypact_directory",
                             state, row["prov_type"], row["last_name"],
                             row["first_name"], row["license_id"], outcome.status)
                    continue

                # --- Resolve routing for this row ---
                prov_type_upper = row.get("prov_type", "").upper()
                key = (row["lic_state"].upper(), prov_type_upper)
                routed_sids = _ROUTING.get(key, [])

                # IL + LC conditional routing: license starting with "L-" goes to
                # IBCLC_COMMISSION only; all other IL LC licenses go to IL_LICENSING only.
                if (row["lic_state"].upper() == "IL" and prov_type_upper == "LC"):
                    _lic_id = (row.get("license_id") or "").strip()
                    if _lic_id.upper().startswith("L-"):
                        routed_sids = [s for s in routed_sids if s == "IBCLC_COMMISSION"]
                    else:
                        routed_sids = [s for s in routed_sids if s == "IL_LICENSING"]

                # CO + LC conditional routing: IBCLC credentials always start with "L-";
                # route L- licenses directly to IBCLC_COMMISSION (skip CO_DORA entirely to
                # prevent NPI-substituted RN records from blocking the IBLC lookup).
                # Non-L- licenses (e.g. RN., APN.) belong to CO_DORA only.
                if (row["lic_state"].upper() == "CO" and prov_type_upper == "LC"):
                    _lic_id = (row.get("license_id") or "").strip()
                    if _lic_id.upper().startswith("L-"):
                        routed_sids = [s for s in routed_sids if s == "IBCLC_COMMISSION"]
                    else:
                        routed_sids = [s for s in routed_sids if s == "CO_DORA"]

                # NC LPC conditional routing: LCAS-prefix → NC_DAC only; others → NC_MENTAL_HEALTH only.
                if row["lic_state"].upper() == "NC" and prov_type_upper == "LPC":
                    _lic_id = (row.get("license_id") or "").strip()
                    if re.sub(r"[^A-Z]", "", _lic_id.upper()).startswith("LCAS"):
                        routed_sids = [s for s in routed_sids if s == "NC_DAC"]
                    else:
                        routed_sids = [s for s in routed_sids if s == "NC_MENTAL_HEALTH"]

                routed_configs = [cfg_by_sid[s] for s in routed_sids if s in cfg_by_sid]

                # Build per-board license_type map so {type} template in extra_selects
                # resolves to the board's specific dropdown value (e.g. "NP", "LPN").
                board_lt_map: dict[str, str] = {}
                for sid in routed_sids:
                    lt_val = _SOCRATA_TYPE_MAP.get((sid, prov_type_upper))
                    if lt_val:
                        board_lt_map[sid] = lt_val

                ladder_result = None
                ai_result = None

                # Cap: License State must appear in Service Location State
                _svc_raw = row.get("svc_loc_state", "")
                _svc_states = [s.strip().upper() for s in _svc_raw.split(",") if s.strip()]
                if _svc_states and row["lic_state"].upper() not in _svc_states:
                    trace.final_outcome = "N/A"
                    trace.final_reason = (
                        f"License State ({row['lic_state'].upper()}) not in Service Location State "
                        f"({_svc_raw}) — PSV step N/A"
                    )
                    _svc_loc_na_rows.append(row)
                # CAPTCHA-blocked prov_type: skip immediately, skip all board calls
                elif _captcha_reason := CAPTCHA_PROV_TYPES.get((state, prov_type_upper)):
                    trace.final_outcome = "Skip"
                    trace.final_reason = "prov_type_captcha_blocked"
                # State does not license this profession (e.g. CO DT/NUT/ABA, MI DT/NUT)
                # — Skip immediately with the stated reason, no board queries needed.
                elif (state, prov_type_upper) in SKIP_UNLICENSED_PROV_TYPES:
                    trace.final_outcome = "Skip"
                    trace.final_reason = NA_PROV_TYPES.get((state, prov_type_upper), "")
                # ABA rows whose license_id is a BACB certification number → always Skip.
                elif (prov_type_upper == "ABA" and _is_bacb_license(row.get("license_id", ""))):
                    trace.final_outcome = "Skip"
                    trace.final_reason = "board_skip_captcha"
                elif not routed_configs:
                    # Any skip:true board → Skip regardless of whether the skip reason
                    # mentions captcha keywords (e.g. BACB "Registry Down" is still a Skip).
                    # Only fall to Fail when no routing at all (empty routed_sids).
                    _has_skip_board = bool(routed_sids) and any(
                        s in _SKIP_REASON_BY_SID for s in routed_sids
                    )
                    if _has_skip_board:
                        _is_captcha = any(
                            _is_captcha_skip(_SKIP_REASON_BY_SID.get(s, ""))
                            for s in routed_sids if s in _SKIP_REASON_BY_SID
                        )
                        trace.final_outcome = "Skip"
                        trace.final_reason = "board_skip_captcha" if _is_captcha else "board_skipped"
                        if not _is_captcha:
                            trace.skip_reason_text = next(
                                (_SKIP_REASON_BY_SID[s] for s in routed_sids if s in _SKIP_REASON_BY_SID),
                                "",
                            )
                    else:
                        trace.final_outcome = "Fail"
                        trace.final_reason = "no_routing"
                else:
                    # --- Run rule-based ladder ---
                    ladder_result = await ladder_mod.run_ladder(
                        routed_configs=routed_configs,
                        master_row=row,
                        nppes_record=nppes,
                        discrepancy=discrepancy,
                        trace=trace,
                        executor=executor,
                        timeout_s=timeout,
                        board_license_type_map=board_lt_map,
                    )

                    # --- Post-license name gate ---
                    # Validates board name vs NPPES name (first) then EPDB name (second)
                    # after cleanup.  Forces AI for [0.70, 0.80) gray-zone scores.
                    # Score computation always runs for Pass rows (for output columns);
                    # AI routing only fires when enable_ai=True.
                    _force_name_gate_ai = False
                    if (ladder_result.status == "Pass"
                            and ladder_result.best_record is not None):
                        from orchestrator.name_gate import evaluate_name_gate  # noqa: PLC0415
                        _ng = evaluate_name_gate(
                            master_row=row,
                            board_record=ladder_result.best_record,
                            nppes=nppes,
                        )
                        trace.epdb_name_score = _ng.epdb_score
                        trace.nppes_name_score = _ng.nppes_score
                        if not _ng.skipped:
                            log.info(
                                "[%s] name_gate: epdb=%s nppes=%s max=%.3f verdict=%s",
                                state,
                                f"{_ng.epdb_score:.3f}" if _ng.epdb_score is not None else "N/A",
                                f"{_ng.nppes_score:.3f}" if _ng.nppes_score is not None else "N/A",
                                _ng.max_score, _ng.verdict,
                            )
                        if _ng.verdict == "ai_review" and enable_ai:
                            _force_name_gate_ai = True
                            trace.name_gate_reason = "name_gate_ai_review"
                        elif _ng.verdict == "manual":
                            trace.name_gate_reason = "name_gate_manual"

                    # --- Cross-state routing: if NPPES shows the license belongs to a
                    # different state, add that state's boards to the AI's routing.
                    # Example: master lic_state=IN but NPPES identifier state=IL.
                    _ai_routed_configs = list(routed_configs)
                    if nppes and nppes.license_numbers:
                        _master_lic_digits = re.sub(r"\D", "", (row.get("license_id") or ""))
                        for _nl in nppes.license_numbers:
                            _nl_state = (_nl.get("state") or "").upper().strip()
                            _nl_digits = re.sub(r"\D", "", str(_nl.get("number") or ""))
                            if (
                                _nl_state
                                and _nl_state != row.get("lic_state", "").upper()
                                and _master_lic_digits
                                and _nl_digits == _master_lic_digits
                            ):
                                _cross_key = (_nl_state, prov_type_upper)
                                _cross_sids = _ROUTING.get(_cross_key, [])
                                _existing_sids = {c.identity.source_id for c in _ai_routed_configs}
                                for _csid in _cross_sids:
                                    if _csid not in _existing_sids and _csid in psv_browsers:
                                        _ai_routed_configs.append(psv_browsers[_csid].config)
                                        log.info(
                                            "[%s] Cross-state routing: adding %s (NPPES "
                                            "license state=%s differs from lic_state=%s)",
                                            state, _csid, _nl_state, row.get("lic_state"),
                                        )
                                break

                    # --- AI agent fallback ---
                    if enable_ai and (
                        ladder_result.status == "EscalateAi" or force_ai or _force_name_gate_ai
                    ):
                        candidate_cache: dict[str, list] = {}
                        for _a in trace.attempts:
                            if _a.candidates:
                                candidate_cache.setdefault(_a.source_id, []).extend(_a.candidates)
                        ai_result = await ai_mod.run_ai_agent(
                            master_row=row,
                            nppes=nppes,
                            discrepancy=discrepancy,
                            routed_configs=_ai_routed_configs,
                            trace=trace,
                            executor=executor,
                            candidate_cache=candidate_cache,
                            timeout_s=timeout,
                            drift_dir=emitter.dirs.get("drift"),
                        )
                        if ai_result.outcome == "resolved":
                            # If AI resolved a candidate but the license number from the
                            # board does not match the input license ID, treat as Fail and
                            # route to Manual — the AI found the wrong person or the board
                            # stores a different license format we cannot reconcile.
                            _ai_bd = ai_result.chosen_breakdown
                            _input_lic_ai = (row.get("license_id") or "").strip()
                            _board_lic_ai = (
                                getattr(ai_result.chosen_candidate, "license_number", "") or ""
                            ).strip()
                            if (
                                _ai_bd is not None
                                and _ai_bd.license_numerics == 0.0
                                and _input_lic_ai
                                and _board_lic_ai
                            ):
                                trace.final_outcome = "Fail"
                                trace.final_reason = "AI found License ID mismatched"
                                ai_result.outcome = "gave_up"
                            else:
                                trace.final_outcome = "Pass"
                                # Back-fill expiry from detail page when the AI chose a
                                # record from a multi-row name search (detail not yet visited).
                                _ai_cand = ai_result.chosen_candidate
                                _ai_src = ai_result.chosen_source_id or ""
                                if _ai_cand is not None and not _get_expiry(_ai_cand) and _ai_src in psv_browsers:
                                    _detail_exp = await _fetch_detail_expiry(
                                        psv_browsers[_ai_src], _ai_cand, timeout
                                    )
                                    if _detail_exp:
                                        try:
                                            from datetime import date as _dt
                                            _ai_cand.expiration_date = _dt.fromisoformat(_detail_exp[:10])
                                        except Exception:
                                            pass
                        elif (
                            ai_result.outcome in ("skipped", "errored")
                            and _force_name_gate_ai
                            and not (ladder_result.status == "EscalateAi" or force_ai)
                        ):
                            # AI was triggered only for name-gate borderline review, not
                            # because the license lookup failed. The ladder already confirmed
                            # the license on the board. When AI is unavailable (circuit-breaker
                            # open or transient error), trust the license match rather than
                            # failing the entire record.
                            log.warning(
                                "[%s] AI unavailable for name-gate review (%s) — "
                                "falling back to ladder Pass (license confirmed on board)",
                                state, ai_result.outcome,
                            )
                            trace.final_outcome = "Pass"
                            trace.name_gate_reason = f"name_gate_ai_fallback:{ai_result.outcome}"
                        elif ai_result.outcome == "errored":
                            # AI API errored (e.g. transient failure on last turn).
                            # Before failing, check if the ladder already accumulated a
                            # candidate with an exact license match — if so, use it.
                            _err_input_lic = (row.get("license_id") or "").strip()
                            _err_best = None
                            _err_bd = None
                            for _a in trace.attempts:
                                for _cand in (_a.candidates or []):
                                    _cand_lic = (getattr(_cand, "license_number", "") or "").strip()
                                    if (_err_input_lic and _cand_lic
                                            and disamb_mod.license_numerics_match(
                                                _err_input_lic, _cand_lic)):
                                        _bd_try = disamb_mod.score_candidate(_cand, row)
                                        if _bd_try.gate_passed:
                                            _err_best = _cand
                                            _err_bd = _bd_try
                                            break
                                if _err_best:
                                    break
                            if _err_best is not None:
                                log.warning(
                                    "[%s] AI errored (%s) — falling back to ladder "
                                    "candidate with exact license match",
                                    state, ai_result.reason,
                                )
                                trace.final_outcome = "Pass"
                                ai_result.outcome = "resolved"
                                ai_result.chosen_candidate = _err_best
                                ai_result.chosen_breakdown = _err_bd
                                ai_result.reason = "ai_error_ladder_fallback"
                            else:
                                trace.final_outcome = "Fail"
                                trace.final_reason = ai_result.reason
                        else:
                            trace.final_outcome = "Fail"
                            trace.final_reason = ai_result.reason

                    # --- PSYPACT secondary check for CP prov_type ---
                    # CP (Clinical Psychologist) providers may hold PSYPACT E.Passports
                    # with mobility numbers matching their state license ID. When the
                    # primary board ladder does not produce a Pass, check PSYPACT as a
                    # secondary source regardless of license-ID prefix.
                    if (prov_type_upper == "CP"
                            and ladder_result is not None
                            and ladder_result.status != "Pass"
                            and trace.final_outcome != "Pass"):
                        _psypact_lr = await _verify_psypact_row(row, trace, run_id=run_id)
                        if _psypact_lr.status == "Pass":
                            ladder_result = _psypact_lr
                            ai_result = None

                # Post-scrape N/A override: board was visited but state does not license this type.
                # The board result (Pass/Fail) is kept; only the reason is overridden so the
                # output explains why licensure was not expected rather than showing "no_records".
                if _post_na := NA_PROV_TYPES.get((state, prov_type_upper)):
                    trace.final_reason = _post_na
                    trace.no_licensure_required = True
                    if ladder_result is not None:
                        ladder_result.reason = ""  # clear so trace.final_reason surfaces

                # nppes_used: True only when NPPES data actually drove the resolution,
                # not merely when it was fetched.
                trace.nppes_used = bool(
                    (ladder_result is not None and ladder_result.npi_substituted)
                    or (ai_result is not None and ai_result.outcome == "resolved"
                        and nppes is not None)
                    or (getattr(trace, "name_gate_reason", None) in (
                            "name_gate_manual", "name_gate_ai_review")
                        and nppes is not None)
                )

                outcome = RowOutcome(
                    master_row=row,
                    master_row_id=master_row_id,
                    trace=trace,
                    nppes=nppes,
                    discrepancy=discrepancy,
                    ladder_result=ladder_result,
                    ai_result=ai_result,
                )
                emitter.collect(outcome)

                _row_display_status = trace.final_outcome or outcome.status
                if outcome.status == "Pass":
                    passes += 1
                elif _row_display_status == "Skip":
                    skips += 1
                else:
                    fails += 1
                log.info("[%s] %s %s %s %s -> %s | %s",
                         state, row["prov_type"], row["last_name"],
                         row["first_name"], row["license_id"],
                         _row_display_status, outcome.reason or "ok")

        finally:
            if browser:
                await browser.close()
    if _svc_loc_na_rows:
        _remove_output_dir = emitter.dirs.get("standard", Path(".")).parent
        _write_remove_license(_svc_loc_na_rows, _remove_output_dir)

    log.info("[%s] State complete: %d Pass / %d Fail / %d Skip / %d Total",
             state, passes, fails, skips, len(rows))
    return passes, fails, skips


async def main_async(args: argparse.Namespace) -> None:
    _load_routing()

    state = args.state.upper()
    output_path = Path(args.output)

    log.info("Loading input rows from %s", args.input)
    rows = load_input_rows(args.input, state, getattr(args, "sheet", ""))
    log.info("Found %d rows for state %s", len(rows), state)
    if not rows:
        log.error("No rows found — check --state and input file")
        sys.exit(1)

    if args.skip_rows and args.skip_rows > 0:
        rows = rows[args.skip_rows:]
        log.info("Skipping first %d rows (--skip-rows), %d remaining", args.skip_rows, len(rows))

    if args.max_rows and args.max_rows < len(rows):
        rows = rows[: args.max_rows]
        log.info("Limiting to %d rows (--max-rows)", args.max_rows)

    passes, fails, *_ = await run_state(
        rows=rows,
        state=state,
        output_path=output_path,
        append=False,
        batch_size=args.batch_size,
        timeout=args.timeout,
        sequential=args.sequential,
    )
    log.info("=== COMPLETE: %d Pass / %d Fail / %d Total → %s ===",
             passes, fails, passes + fails, output_path)


def main() -> None:
    p = argparse.ArgumentParser(description="PSV batch tester")
    p.add_argument("--input", required=True, help="Input Excel path (PSV Tab sheet)")
    p.add_argument("--output", required=True, help="Output Excel path")
    p.add_argument("--state", required=True, help="State abbreviation (e.g. MD, WY)")
    p.add_argument("--batch-size", type=int, default=10,
                   help="Rows to process per batch before writing (default: 10)")
    p.add_argument("--timeout", type=int, default=120,
                   help="Per-board per-mode timeout in seconds (default: 120)")
    p.add_argument("--skip-rows", type=int, default=0,
                   help="Skip the first N rows of the filtered input (default: 0)")
    p.add_argument("--max-rows", type=int, default=0,
                   help="Stop after processing this many rows (0 = all)")
    p.add_argument("--sequential", action="store_true", default=False,
                   help="Process rows one at a time (avoids concurrent board overload)")
    p.add_argument("--sheet", default="",
                   help="Sheet name to read from (default: 'PSV Tab')")
    args = p.parse_args()
    asyncio.run(main_async(args))


if __name__ == "__main__":
    main()
