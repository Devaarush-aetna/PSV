"""Generate OH failing-cases analysis Excel sheet."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
from engine.csv_extractor import load_csv, search_by_license_number
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH = Path(__file__).parent / "../../../PSV/CSVS/OH_PROVIDERS_INDIVIDUAL_20260727_1946.csv"
OUT_PATH = r"C:\Users\n676150\Downloads\OH_Analysis_Failing_Cases.xlsx"

print("Loading CSV...")
df = load_csv(CSV_PATH, "utf-8-sig", 0, ",")
print(f"Loaded {len(df):,} rows.")

cols = ["LICENSE_NUMBER"] + [f"ENDORSEMENT_NUMBER_{i}" for i in range(1, 8)]

# (first, middle, last, state, prov_type, lic_type, license_id, npi, orig_status)
cases = [
    ("Stawa",    "Rajab",     "Blanco",       "OH","NP",  "OPERATING",    "APRN.CNP.0026867",  "1073126314","Fail"),
    ("Nicole",   "Fritz",     "FRITZ",         "OH","LPC", "OPERATING",    "E.2404431",         "1326534975","Fail"),
    ("Riley",    "James",     "Miller",        "OH","NP",  "OPERATING",    "APRN.CNP.0039506",  "1508750746","Fail"),
    ("Erin",     "",          "Meckley",       "OH","NP",  "OPERATING",    "APRN.CNP.0037578",  "1104298249","Fail"),
    ("Nekia",    "",          "Jackson",       "OH","NP",  "OPERATING",    "APRN.CNP.0033268",  "1821606658","Fail"),
    ("Matthew",  "Edward",    "Raymond",       "OH","RNA", "OPERATING",    "APRN.CRNA.0021407", "1710561147","Fail"),
    ("Laura",    "",          "Babbitt",       "OH","NP",  "OPERATING",    "APRN.CNP.14484",    "1326428715","Fail"),
    ("Kenneth",  "L",         "Moore",         "OH","NP",  "OPERATING",    "APRN.CNP.14566",    "1013995323","Fail"),
    ("Ciara",    "",          "Haas",          "OH","PN",  "OPERATING",    "0039670",           "1568059715","Fail"),
    ("Joanna",   "",          "Overholt",      "OH","NP",  "OPERATING",    "APRN.CNP.0030764",  "1194161687","Fail"),
    ("Sarah",    "",          "Jernigan",      "OH","DT",  "OPERATING",    "86050463",          "1427714013","Fail"),
    ("Nadia",    "M.",        "Robinson",      "OH","MW",  "OPERATING",    "APRN.CNP.0030218",  "1407460488","Fail"),
    ("Tia",      "",          "King",          "OH","NP",  "OPERATING",    "APRN.CNP.0036979",  "1841740206","Fail"),
    ("Carol",    "",          "Green",         "OH","NP",  "OPERATING",    "APRN.CNP.0028065",  "1568026011","Fail"),
    ("Leanne",   "",          "Carman",        "OH","NP",  "OPERATING",    "APRN.CNP.0039312",  "1174315592","Fail"),
    ("Gregory",  "Michael",   "McDonnell",     "OH","DC",  "OPERATING",    "ACUP-00177",        "1285174151","Fail"),
    ("Douglas",  "S",         "Larner",        "OH","NP",  "OPERATING",    "APRN.CNP.0037279",  "1467447896","Fail"),
    ("McKenna",  "",          "Waltenbaugh",   "OH","NP",  "OPERATING",    "APRN.CNP.0036725",  "1902650518","Fail"),
    ("Rachel",   "L.",        "Singer",        "OH","NP",  "OPERATING",    "APRN.CNP.13339",    "1093073546","Fail"),
    ("Maria",    "",          "Surmachevska",  "OH","NP",  "OPERATING",    "APRN.CNP.0039161",  "1730897554","Fail"),
    ("Tiffany",  "",          "Colston",       "OH","NP",  "OPERATING",    "APRN.CNP.0028286",  "1083203665","Fail"),
    ("Cierra",   "Christine", "Ramey",         "OH","NP",  "OPERATING",    "APRN.CNP.026234",   "1265078307","Fail"),
    ("Lucille",  "Elizabeth", "Nathwani",      "OH","NP",  "OPERATING",    "APRN.CNP.0032280",  "1235740556","Fail"),
    ("April",    "",          "Kline",         "OH","LC",  "OPERATING",    "10030039",          "1164752457","Fail"),
    ("Angela",   "",          "Lis",           "OH","NP",  "OPERATING",    "APRN.CNP.0034610",  "1245758978","Fail"),
    ("Carol",    "",          "Green",         "OH","NP",  "OPERATING",    "APRN.CNP.0028065",  "1023791449","Fail"),
    ("Stacey",   "L",         "Childress",     "OH","NP",  "OPERATING",    "COA.12146-NP",      "1083018030","Fail"),
    ("Christina","",          "Serger",        "OH","NP",  "OPERATING",    "APRN.CNP.0038559",  "1295466084","Fail"),
    ("Kayla",    "",          "McAdam",        "OH","NP",  "OPERATING",    "APRN.CNP.0035678",  "1457396368","Fail"),
    ("Kayla",    "",          "McAdam",        "OH","NP",  "OPERATING",    "APRN.CNP.0035678",  "1457113144","Fail"),
    ("Marco",    "",          "More",          "OH","NP",  "OPERATING",    "APRN.CNP.0037287",  "1659882470","Fail"),
    ("Veronica", "",          "Galaszewski",   "OH","NP",  "OPERATING",    "APRN.CNP.0039555",  "1194619700","Fail"),
    ("Maria",    "E",         "Carico",        "OH","NP",  "OPERATING",    "APRN.CNP.17686",    "1265813141","Fail"),
    ("Sylvie",   "",          "Riley",         "OH","NP",  "OPERATING",    "APRN.CNP.0037507",  "1285469361","Fail"),
    ("Jennifer", "L",         "Lakeberg",      "OH","NP",  "OPERATING",    "APRN.CNP.14477",    "1295168409","Fail"),
    ("Aissata",  "",          "Diallo",        "OH","NP",  "OPERATING",    "APRN.CNP.0041396",  "1336993351","Fail"),
    ("Sydney",   "",          "Waldon",        "OH","NP",  "OPERATING",    "APRN.CNP.0036240",  "1760952873","Fail"),
    ("Sarah",    "Michelle",  "Foltz",         "OH","NP",  "OPERATING",    "APRN.CNP.022942",   "1306315007","Fail"),
    ("Vickie",   "",          "Knueven",       "OH","NP",  "OPERATING",    "APRN.CNP.0039311",  "1083406516","Fail"),
    ("Umera",    "",          "Paracha",       "OH","PH",  "STATE MEDICAL","25MA11601000",      "1558717850","Fail"),
    # Pass cases (already working or fixed)
    ("Michelle", "",          "Lepsesty",      "OH","NP",  "OPERATING",    "APRN.CNP.0032579",  "1811604556","Pass"),
    ("Katelyn",  "Emily",     "Dudenhoeffer",  "OH","SH",  "OPERATING",    "SP/11765",          "1689073421","Pass"),
]

rows = []
for (first, mid, last, state, prov, lic_type, lic, npi, orig_status) in cases:
    results = search_by_license_number(df, cols, lic)
    if not results:
        rows.append({
            "first_name": first, "middle_name": mid, "last_name": last,
            "lic_state": state, "prov_type": prov, "lic_type": lic_type,
            "license_id": lic, "npi_no": npi, "original_status": orig_status,
            "search_result": "NOT FOUND",
            "matched_column": "",
            "csv_name": "",
            "csv_license_number": "",
            "endorsement_status": "",
            "endorsement_expiry": "",
            "main_license_status": "",
            "main_license_expiry": "",
            "finding": "License not present in OH Individual CSV — wrong license number or different board",
            "expected_outcome": "Needs Investigation",
        })
        continue

    r = results[0]
    csv_name = r.get("LICENSEE_NAME", "")
    main_status = r.get("STATUS", "")
    main_exp = r.get("EXPIRATION_DATE", "")
    matched_col = "LICENSE_NUMBER"
    end_status = main_status
    end_exp = main_exp

    for i in range(1, 8):
        ec = f"ENDORSEMENT_NUMBER_{i}"
        if r.get(ec, "").strip().upper() == lic.strip().upper():
            matched_col = ec
            end_status = r.get(f"ENDORSEMENT_STATUS_{i}", main_status)
            end_exp = r.get(f"ENDORSEMENT_EXPIRATION_DATE_{i}", main_exp)
            break

    last_in_csv = last.upper() in csv_name.upper()

    if not last_in_csv:
        finding = f"WRONG PERSON — input last={last}, CSV name={csv_name}"
        outcome = "Investigate / Data Discrepancy"
    elif end_status.lower() in ("inactive", "expired", "revoked", "suspended"):
        finding = f"Found via {matched_col}. Endorsement status={end_status} — correctly FAIL"
        outcome = "Correct Fail (Inactive/Expired Endorsement)"
    elif matched_col == "LICENSE_NUMBER":
        csv_lic = r.get("LICENSE_NUMBER", "")
        if lic.upper() != csv_lic.upper():
            finding = f"Substring match in LICENSE_NUMBER (CSV={csv_lic}). Status={end_status}"
        else:
            finding = f"Exact match in LICENSE_NUMBER. Status={end_status}"
        outcome = "Pass after fix"
    else:
        finding = f"Found via {matched_col} (endorsement). Status={end_status}, Expiry={end_exp}"
        outcome = "Pass after endorsement fix"

    rows.append({
        "first_name": first, "middle_name": mid, "last_name": last,
        "lic_state": state, "prov_type": prov, "lic_type": lic_type,
        "license_id": lic, "npi_no": npi, "original_status": orig_status,
        "search_result": "FOUND",
        "matched_column": matched_col,
        "csv_name": csv_name,
        "csv_license_number": r.get("LICENSE_NUMBER", ""),
        "endorsement_status": end_status,
        "endorsement_expiry": end_exp,
        "main_license_status": main_status,
        "main_license_expiry": main_exp,
        "finding": finding,
        "expected_outcome": outcome,
    })

out_df = pd.DataFrame(rows)
print(f"Total rows: {len(rows)}")

# ── Build formatted Excel ──────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "OH Analysis"

hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = list(out_df.columns)
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

green  = PatternFill("solid", fgColor="C6EFCE")
red    = PatternFill("solid", fgColor="FFC7CE")
yellow = PatternFill("solid", fgColor="FFEB9C")
orange = PatternFill("solid", fgColor="F4B942")
blue   = PatternFill("solid", fgColor="DDEBF7")
grey   = PatternFill("solid", fgColor="EDEDED")

finding_ci = headers.index("finding") + 1

for ri, row in enumerate(out_df.itertuples(index=False), 2):
    vals = list(row)
    outcome = vals[headers.index("expected_outcome")]
    found   = vals[headers.index("search_result")]
    orig    = vals[headers.index("original_status")]

    if "Pass after" in outcome:            row_fill = green
    elif "Correct Fail" in outcome:        row_fill = yellow
    elif "Data Discrepancy" in outcome:    row_fill = orange
    elif "NOT FOUND" == found:             row_fill = red
    elif orig == "Pass":                   row_fill = blue
    else:                                  row_fill = grey

    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = row_fill
        c.alignment = Alignment(
            vertical="center",
            wrap_text=(ci == finding_ci),
        )
        c.font = Font(size=9)
        c.border = border

widths = {
    "first_name": 14, "middle_name": 11, "last_name": 15, "lic_state": 7,
    "prov_type": 8, "lic_type": 13, "license_id": 22, "npi_no": 13,
    "original_status": 11, "search_result": 11, "matched_column": 24,
    "csv_name": 28, "csv_license_number": 19, "endorsement_status": 15,
    "endorsement_expiry": 15, "main_license_status": 15, "main_license_expiry": 15,
    "finding": 55, "expected_outcome": 32,
}
for ci, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)

ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

# ── Summary sheet ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
total     = len(rows)
pass_fix  = sum(1 for r in rows if "Pass after" in r["expected_outcome"])
corr_fail = sum(1 for r in rows if "Correct Fail" in r["expected_outcome"])
investig  = sum(1 for r in rows if "Investigate" in r["expected_outcome"])
not_found = sum(1 for r in rows if r["search_result"] == "NOT FOUND")
orig_pass = sum(1 for r in rows if r["original_status"] == "Pass")

summary = [
    ("Category", "Count", "Notes"),
    ("Total OH cases in sheet", total, "42 fail + 2 already-pass"),
    ("Will PASS after endorsement fix", pass_fix,
     "Found in ENDORSEMENT_NUMBER_N col, name ok, Active status"),
    ("Correctly FAIL (inactive/expired endorsement)", corr_fail,
     "Erin Meckley CNP.0037578, Nadia Robinson CNP.0030218, Gregory McDonnell ACUP-00177"),
    ("NOT FOUND in OH Individual CSV", not_found,
     "86050463 Jernigan, 10030039 Kline, COA.12146-NP Childress, 25MA11601000 Paracha"),
    ("Wrong person / Data Discrepancy", investig,
     "0039670 Haas (CSV=SIZEMORE), E.2404431 Fritz (CSV=E.2404431-SUPV diff person)"),
    ("Already passing (reference cases)", orig_pass,
     "Michelle Lepsesty, Katelyn Dudenhoeffer"),
]

hdr_fill2 = PatternFill("solid", fgColor="1F4E79")
for ri, (a, b, c) in enumerate(summary, 1):
    ca = ws2.cell(ri, 1, a)
    cb = ws2.cell(ri, 2, b)
    cc = ws2.cell(ri, 3, c)
    if ri == 1:
        for cell in [ca, cb, cc]:
            cell.fill = hdr_fill2
            cell.font = Font(bold=True, color="FFFFFF", size=10)
    else:
        ca.font = Font(bold=True, size=10)
        cb.font = Font(size=10)
        cc.font = Font(size=9)
        cc.alignment = Alignment(wrap_text=True)

ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 8
ws2.column_dimensions["C"].width = 65
for ri in range(1, len(summary) + 1):
    ws2.row_dimensions[ri].height = 22 if ri == 1 else 18

ws2.freeze_panes = "A2"

# ── Legend sheet ───────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Legend")
legend = [
    ("Colour", "Meaning"),
    ("Green",  "Will PASS after endorsement-column fix (Fix 1)"),
    ("Yellow", "Correctly FAIL — endorsement is Inactive/Expired (system working correctly)"),
    ("Red",    "NOT FOUND in OH Individual CSV — wrong license or different board"),
    ("Orange", "Found but wrong person — data discrepancy in input"),
    ("Blue",   "Already passing reference cases"),
]
for ri, (a, b) in enumerate(legend, 1):
    ws3.cell(ri, 1, a).font = Font(bold=(ri==1), size=10)
    ws3.cell(ri, 2, b).font = Font(size=10)
    fills = [None, green, yellow, red, orange, blue]
    if ri > 1:
        ws3.cell(ri, 1).fill = fills[ri - 1]
ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 60

wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
print(f"  Pass after fix  : {pass_fix}")
print(f"  Correct fail    : {corr_fail}")
print(f"  Not found       : {not_found}")
print(f"  Data discrepancy: {investig}")
print(f"  Orig pass refs  : {orig_pass}")
