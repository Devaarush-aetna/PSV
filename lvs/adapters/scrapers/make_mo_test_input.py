"""Create a small MO test input xlsx for verifying the mopro_zip dedup fix.

Providers are taken from the known MO_HEALING_ARTS CSV where the same
(lic_number, lic_profession) triplicate pattern causes AI escalation.

Column layout matches psv_test.py constants:
  0  FIRST_NAME
  1  MIDDLE_NAME
  2  LAST_NAME
  3  EPDB_PIN
  4  PROV_TYPE
  5  (blank)
  6  (blank)
  7  MAINTAINED_BY
  8  (blank)
  9  LIC_STATE
  10 LIC_TYPE
  11 LIC_ID
  12 LIC_EXPIRY
  13 (blank)
  14 (blank)
  15 SVC_LOC_STATE
  (dynamic) NPI_NO  -- header detected by name
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

HEADER = [
    "FIRST_NAME", "MIDDLE_NAME", "LAST_NAME", "EPDB_PIN",
    "PROV_TYPE", "", "", "MAINTAINED_BY", "",
    "LIC_STATE", "LIC_TYPE", "LIC_ID", "LIC_EXPIRY",
    "", "", "SVC_LOC_STATE", "NPI_NO",
]

# 5 real providers from MO_HEALING_ARTS CSV — all known triplicates in the
# unpatched CSV (same lic_number+lic_profession appearing 3x from 3 ZIPs).
# NPI_NO left empty — NPPES enrichment will be skipped; board lookup still runs.
ROWS = [
    # FIRST      MID  LAST       PIN  TYPE  _  _  MAINT  _  STATE  LIC_TYPE  LIC_ID        EXPIRY       _  _  SVC_STATE  NPI
    ["Shilpa",   "",  "Saxena",  "",  "MD", "", "", "",  "", "MO", "MED", "2019033203", "01/31/2027", "", "", "MO",     ""],
    ["Jeffrey",  "",  "Markey",  "",  "MD", "", "", "",  "", "MO", "MED", "2019033208", "01/31/2027", "", "", "MO",     ""],
    ["Mark",     "",  "Novick",  "",  "MD", "", "", "",  "", "MO", "MED", "2011026857", "01/31/2027", "", "", "MO",     ""],
    ["Michael",  "",  "Camp",    "",  "MD", "", "", "",  "", "MO", "MED", "2010027315", "01/31/2027", "", "", "MO",     ""],
    ["Alisha",   "",  "White",   "",  "MD", "", "", "",  "", "MO", "MED", "2010021889", "01/31/2027", "", "", "MO",     ""],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PSV Tab"

header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
for col, val in enumerate(HEADER, 1):
    cell = ws.cell(row=1, column=col, value=val)
    cell.fill = header_fill
    cell.font = header_font

for r, row in enumerate(ROWS, 2):
    for col, val in enumerate(row, 1):
        ws.cell(row=r, column=col, value=val)

out = "Input_MO_dedup_test.xlsx"
wb.save(out)
print(f"Saved {out} with {len(ROWS)} MO rows")
print("Run with:")
print(f"  python psv_test.py --input {out} --output Output_MO_dedup_test.xlsx --state MO --batch-size 5 --sequential")
