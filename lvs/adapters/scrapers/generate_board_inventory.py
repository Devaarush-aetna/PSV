"""
Generate board_inventory.xlsx from all sites/*/config.yaml files.

Run from lvs/adapters/scrapers/:
    python generate_board_inventory.py

Re-run after adding any new board to pick it up automatically.
Output: board_inventory.xlsx (same directory)
"""

import glob
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import yaml

SCRIPT_DIR = Path(__file__).parent
SITES_DIR = SCRIPT_DIR / "sites"
OUTPUT_FILE = SCRIPT_DIR / "board_inventory.xlsx"

ARCHETYPE_LABELS = {
    "thentia_cloud":    "Web Scraping – Thentia Cloud SPA",
    "classic_html_form": "Web Scraping – HTML Form",
    "ag_grid_spa":      "Web Scraping – AG Grid SPA",
    "socrata_api":      "API – Socrata JSON (direct HTTP)",
    "socrata_bulk_csv": "API – Socrata JSON (browser-based)",
    "pdf_bulk":         "PDF Download",
}

STATUS_COLORS = {
    "PASS":    "C6EFCE",  # green
    "SKIP":    "FFEB9C",  # yellow
    "PARTIAL": "FFC7CE",  # pink/red
    "MISSING": "D9D9D9",  # grey
    "READY":   "DDEBF7",  # blue
}

HEADERS = [
    "#",
    "State",
    "Source ID",
    "Board Name",
    "Profession Codes",
    "Source URL",
    "Ingestion Type",
    "Archetype",
    "Smoke Test Status",
    "Skip Reason / Notes",
]

COL_WIDTHS = [5, 8, 18, 50, 20, 55, 38, 22, 22, 45]


def load_board(config_path: Path) -> dict:
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    identity = cfg.get("identity", {})
    smoke = cfg.get("smoke_test", None)

    archetype = identity.get("archetype", "")
    ingestion_type = ARCHETYPE_LABELS.get(archetype, archetype)

    if smoke is None:
        status = "MISSING"
        skip_reason = "No smoke_test block in config"
    elif smoke.get("skip", False):
        status = "SKIP"
        skip_reason = smoke.get("skip_reason", "")
    else:
        status = "READY"
        skip_reason = ""

    professions = identity.get("profession_codes", [])
    profession_str = ", ".join(professions) if professions else ""

    return {
        "state": identity.get("state", ""),
        "source_id": identity.get("source_id", ""),
        "board_name": identity.get("board_name", ""),
        "professions": profession_str,
        "base_url": identity.get("base_url", ""),
        "ingestion_type": ingestion_type,
        "archetype": archetype,
        "status": status,
        "skip_reason": skip_reason,
    }


def build_excel(boards: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Board Inventory"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # State group fill alternation
    state_fills = {}
    alt_colors = ["EBF4FF", "FFFFFF"]
    color_idx = 0

    # Data rows
    row_num = 2
    for seq, board in enumerate(boards, start=1):
        state = board["state"]
        if state not in state_fills:
            state_fills[state] = alt_colors[color_idx % 2]
            color_idx += 1
        row_fill_color = state_fills[state]

        status = board["status"]
        status_color = STATUS_COLORS.get(status, "FFFFFF")

        row_data = [
            seq,
            board["state"],
            board["source_id"],
            board["board_name"],
            board["professions"],
            board["base_url"],
            board["ingestion_type"],
            board["archetype"],
            board["status"],
            board["skip_reason"],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (4, 6, 10)))

            # Light row background (alternating by state)
            if col_idx == 9:  # Status column gets status color
                cell.fill = PatternFill("solid", fgColor=status_color)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = PatternFill("solid", fgColor=row_fill_color)

        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Boards: {len(boards)}")


def main():
    config_paths = sorted(SITES_DIR.glob("*/config.yaml"))
    if not config_paths:
        print(f"No config.yaml files found under {SITES_DIR}")
        return

    boards = []
    for path in config_paths:
        try:
            boards.append(load_board(path))
        except Exception as e:
            print(f"  WARN: skipped {path.parent.name}: {e}")

    # Sort by state, then board name
    boards.sort(key=lambda b: (b["state"], b["board_name"]))

    build_excel(boards)

    # Print summary table to console
    print(f"\n{'#':>3}  {'State':<6} {'Source ID':<20} {'Status':<8}  Board Name")
    print("-" * 90)
    for i, b in enumerate(boards, 1):
        print(f"{i:>3}  {b['state']:<6} {b['source_id']:<20} {b['status']:<8}  {b['board_name']}")


if __name__ == "__main__":
    main()
