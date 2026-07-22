"""PSV Verification Runner — Streamlit UI

Professional parallel orchestration interface for CVS Health PSV batch runs.
One-click launcher: double-click launch_psv_ui.bat

Pages:
  Run Manager  — scan RunQueue/, start parallel runs, live progress tracking
  Results Viewer — browse Standard output, filter by status/state/prov_type, export
"""
from __future__ import annotations

import re
import shutil
import subprocess
import sys
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Constants — file lives at lvs/adapters/scrapers/engine/psv_ui.py
# ---------------------------------------------------------------------------
PSV_DEV     = Path(__file__).parents[4]   # 4 levels up → PSV_DEV root
RUN_QUEUE   = PSV_DEV / "RunQueue"
ASSETS      = PSV_DEV / "assets"
OUTPUT_ROOT = PSV_DEV / "Output"
LOG_DIR     = Path("c:/tmp")
MAX_FILES   = 10
REFRESH_S   = 1   # seconds between live progress refreshes

# ---------------------------------------------------------------------------
# One-time setup (runs on every cold start, fast if already done)
# ---------------------------------------------------------------------------
RUN_QUEUE.mkdir(exist_ok=True)
ASSETS.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)

_LOGO = ASSETS / "cvs_logo.jpeg"
for _src in [
    Path("C:/Users/n661685/Downloads/CVS.jpg"),
    Path("C:/Users/n661685/Downloads/CVS.jpeg"),
]:
    if _src.exists() and not _LOGO.exists():
        shutil.copy(_src, _LOGO)

# ---------------------------------------------------------------------------
# Streamlit page config (must be first st call)
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="PSV Verification Runner",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    /* Hide Streamlit top bar decoration */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }

    /* Red progress bars */
    div[data-testid="stProgressBar"] > div > div > div {
        background-color: #CC0000 !important;
    }

    /* Compact metric labels */
    div[data-testid="stMetricLabel"] p { font-size: .8rem; }
    div[data-testid="stMetricValue"]  { font-size: 1.5rem; font-weight: 700; }

    /* Run card hover effect */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        transition: box-shadow .15s;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:hover {
        box-shadow: 0 2px 12px rgba(204,0,0,.15);
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_US_STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY","DC","PR","VI","GU","MP","AS",
}

def _detect_state(path: Path) -> str:
    """Extract 2-letter state abbreviation from filename or cell 9 row 1."""
    # Scan all regex matches; return the first one that is a real US state code.
    # This prevents non-state tokens (e.g. "TP" in "Input_TP.xlsx") from being
    # returned prematurely and causing run_psv.py to filter 0 rows.
    for m in re.finditer(r'[_\-]([A-Z]{2})[_\-\.]', path.stem + "."):
        candidate = m.group(1)
        if candidate in _US_STATES:
            return candidate
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        wb.close()
        if row2 and len(row2) > 9:
            v = str(row2[9] or "").strip().upper()
            if re.fullmatch(r"[A-Z]{2}", v):
                return v
    except Exception:
        pass
    return "??"


def _parse_log(log_path: str) -> dict:
    """Read entire log, return latest progress snapshot.

    Per-row log format (from psv_test.py):
        [STATE] prov_type last_name first_name lic_id -> Pass|Fail|Skip | reason

    State complete line:
        [STATE] State complete: N Pass / N Fail / N Total
        (Fail count includes Skip rows — use per-row counts for accurate breakdown)
    """
    result = {"done": 0, "total": 0, "passes": 0, "fails": 0, "skips": 0, "finished": False}
    try:
        with open(log_path, "r", encoding="utf-8", errors="replace") as fh:
            content = fh.read()

        # Batch mode logs "Running totals: N Pass / N Fail / N done of M" after each
        # batch — use the last such line to get a live total row count before the
        # State complete line fires.
        totals_m = re.findall(
            r"Running totals: \d+ Pass / \d+ Fail / \d+ done of (\d+)", content
        )
        if totals_m:
            result["total"] = int(totals_m[-1])

        # Per-row outcomes — two formats exist:
        #   batch mode:        [STATE] ... → Pass|Fail|Skip | [expiry=...] reason
        #   single-row mode:   [STATE] ... -> Pass|Fail|Skip | reason
        # Capture optional reason token to reclassify Fail→Skip for blocked boards.
        _LOG_FAIL_AS_SKIP = frozenset(["no_routing", "board_skip_captcha", "board_skipped"])
        rows = re.findall(r"\[(\w+)\] .* (?:->|→) (Pass|Fail|Skip)(?:\s*\|\s*(\S+))?", content)
        result["passes"] = sum(1 for _, o, _r in rows if o == "Pass")
        result["skips"]  = sum(
            1 for _, o, r in rows
            if o == "Skip" or (o == "Fail" and r in _LOG_FAIL_AS_SKIP)
        )
        result["fails"]  = sum(
            1 for _, o, r in rows
            if o == "Fail" and r not in _LOG_FAIL_AS_SKIP
        )
        result["done"]   = result["passes"] + result["fails"] + result["skips"]

        # State complete line signals run finished; use pass count from it (authoritative)
        # Fail count from State complete lumps Skips in — keep per-row breakdown instead
        m2 = re.search(
            r"State complete: (\d+) Pass / (\d+) Fail / (\d+) Total", content
        )
        if m2:
            result["passes"]   = int(m2.group(1))
            result["total"]    = int(m2.group(3))
            # keep per-row fails/skips — more accurate than lumped State complete fails
            result["done"]     = result["total"]
            result["finished"] = True
    except Exception:
        pass
    return result


_SKIP_REASONS = frozenset([
    "prov_type_captcha_blocked",   # (state, prov_type) hard-coded captcha block
    "no_routing",                  # no board routing configured for this prov_type
    "board_skip_captcha",          # board config has skip:true + captcha keyword
    "board_skipped",               # board config has skip:true (non-captcha, e.g. registry down)
])

def _derive_display_status(df: pd.DataFrame) -> pd.Series:
    """Return a Series with Pass / Fail / Skip derived from status + reason columns."""
    if "status" not in df.columns:
        return pd.Series([""] * len(df), index=df.index)
    result = df["status"].copy()
    if "reason" in df.columns:
        is_skip = (result == "Fail") & df["reason"].isin(_SKIP_REASONS)
        result = result.where(~is_skip, other="Skip")
    return result


def _fmt_dur(seconds: float) -> str:
    if seconds < 60:
        return f"{int(seconds)}s"
    m, s = divmod(int(seconds), 60)
    return f"{m}m {s}s"


def _fmt_eta(run: dict) -> str:
    started = run.get("started_at")
    if not started or run["done"] == 0:
        return "Calculating..."
    elapsed = time.time() - started
    rate = run["done"] / elapsed if elapsed > 0 else 0
    if rate == 0 or run["total"] == 0:
        return "..."
    return _fmt_dur((run["total"] - run["done"]) / rate)


def _disk_free_mb(drive: str = "C:/") -> float:
    """Return free megabytes on the given drive."""
    try:
        return shutil.disk_usage(drive).free / 1024 ** 2
    except Exception:
        return float("inf")


def _scan_queue() -> tuple[list[Path], list[str], str | None]:
    """Scan RunQueue/. Returns (xlsx_files, skipped_names, error_msg|None)."""
    xlsx, skipped = [], []
    if not RUN_QUEUE.exists():
        return xlsx, skipped, None
    for f in sorted(RUN_QUEUE.iterdir()):
        if not f.is_file():
            continue
        # Skip Windows Excel lock files (~$filename.xlsx) and hidden dot-files
        if f.name.startswith("~$") or f.name.startswith("."):
            continue
        if f.suffix.lower() == ".xlsx":
            xlsx.append(f)
        else:
            skipped.append(f.name)
    if len(xlsx) > MAX_FILES:
        return xlsx, skipped, (
            f"⛔ **{len(xlsx)} .xlsx files** found in RunQueue/ — maximum is **{MAX_FILES}**. "
            f"Please remove files until {MAX_FILES} or fewer remain, then click **Refresh**."
        )
    return xlsx, skipped, None


def _launch_run(run: dict) -> None:
    """Start run_psv.py subprocess for a single run entry."""
    script   = PSV_DEV / "lvs/adapters/scrapers/run_psv.py"
    log_path = str(LOG_DIR / f"orch_{run['state']}_{run['run_id']}.log")
    run["log_path"]   = log_path
    run["started_at"] = time.time()
    run["status"]     = "running"
    cmd = [
        sys.executable, str(script),
        "--input",  run["input_file"],
        "--states", run["state"],
        "--run-id", run["run_id"],
        "--sheet",  "PSV Tab",
    ]
    with open(log_path, "w", encoding="utf-8") as fh:
        proc = subprocess.Popen(cmd, cwd=str(PSV_DEV), stdout=fh, stderr=fh)
    run["_proc"] = proc


def _scheduler(runs: list[dict], max_workers: int) -> None:
    """Background thread: worker pool. Completed slots auto-pick next queued job."""
    while True:
        active = [r for r in runs if r["status"] == "running"]
        queued = [r for r in runs if r["status"] == "queued"]

        for r in active:
            proc: subprocess.Popen | None = r.get("_proc")
            if proc and proc.poll() is not None:
                if not r.get("finished_at"):
                    r["finished_at"] = time.time()
                r["status"] = (
                    "stopped"  if r.get("_stop_requested") else
                    "complete" if proc.returncode == 0     else
                    "error"
                )

        active = [r for r in runs if r["status"] == "running"]
        queued = [r for r in runs if r["status"] == "queued"]

        while len(active) < max_workers and queued:
            nxt = queued.pop(0)
            _launch_run(nxt)
            active.append(nxt)

        if not active and not queued:
            break
        time.sleep(1)

# ---------------------------------------------------------------------------
# Shared header
# ---------------------------------------------------------------------------

def _header() -> None:
    col_logo, col_title = st.columns([1, 11])
    with col_logo:
        if _LOGO.exists():
            st.image(str(_LOGO), width=80)
        else:
            st.markdown(
                "<div style='width:80px;height:50px;background:#CC0000;"
                "border-radius:6px;display:flex;align-items:center;"
                "justify-content:center;color:white;font-weight:700;"
                "font-size:1.1rem;'>CVS</div>",
                unsafe_allow_html=True,
            )
    with col_title:
        st.markdown(
            "<h2 style='color:#CC0000;margin:0;padding-top:6px;'>"
            "PSV Verification Runner</h2>"
            "<p style='color:#666;margin:0;font-size:.9rem;'>"
            "CVS Health &nbsp;|&nbsp; Provider Specialty Verification "
            "&nbsp;&mdash;&nbsp; Parallel Orchestration</p>",
            unsafe_allow_html=True,
        )
    st.divider()

# ---------------------------------------------------------------------------
# Run card
# ---------------------------------------------------------------------------

_STATUS_ICON  = {"queued": "🕐 Queued", "running": "⏳ Running",
                 "complete": "✅ Complete", "error": "❌ Error", "stopped": "⏹ Stopped"}
_STATUS_COLOR = {"queued": "#757575", "running": "#1565C0",
                 "complete": "#2E7D32", "error": "#C62828", "stopped": "#E65100"}


def _run_card(run: dict, all_runs: list[dict]) -> None:
    s     = run["status"]
    done  = run["done"]
    total = run["total"]
    pct   = done / total if total > 0 else 0.0

    passes    = run["passes"]
    fails     = run["fails"]
    skips     = run["skips"]
    processed = passes + fails          # excludes skips — matches Results Viewer formula
    rate      = passes / processed * 100 if processed > 0 else 0.0

    with st.container(border=True):
        c1, c2, c3 = st.columns([4, 3, 3])

        with c1:
            st.markdown(f"**{run['filename']}**")
            st.caption(f"State: `{run['state']}`  •  Run ID: `{run['run_id']}`")

        with c2:
            st.markdown(
                f"<span style='color:{_STATUS_COLOR[s]};font-weight:600;'>"
                f"{_STATUS_ICON[s]}</span>",
                unsafe_allow_html=True,
            )
            if total > 0:
                st.markdown(f"**Rows: {done:,} / {total:,}**")

        with c3:
            if s == "running":
                if st.button("⏹ Stop", key=f"_stop_{run['run_id']}", type="secondary"):
                    proc = run.get("_proc")
                    if proc:
                        try:
                            proc.terminate()
                        except Exception:
                            pass
                    run["_stop_requested"] = True
                    run["status"] = "stopped"
                    if not run.get("finished_at"):
                        run["finished_at"] = time.time()
                    st.rerun()
                st.caption(f"ETA: {_fmt_eta(run)}")
            elif s in ("complete", "stopped") and run.get("started_at") and run.get("finished_at"):
                st.caption(f"Duration: {_fmt_dur(run['finished_at'] - run['started_at'])}")
            elif s == "error" and run.get("started_at") and run.get("finished_at"):
                st.caption(f"Duration: {_fmt_dur(run['finished_at'] - run['started_at'])}")
            elif s == "queued":
                q_list = [r for r in all_runs if r["status"] == "queued"]
                pos = next((i + 1 for i, r in enumerate(q_list) if r is run), "—")
                st.caption(f"Queue position: {pos}")

        if s == "error" and run.get("log_path"):
            # Show last few lines of log to help diagnose crash
            try:
                with open(run["log_path"], "r", encoding="utf-8", errors="replace") as _f:
                    _lines = _f.readlines()
                _tail = "".join(_lines[-10:]).strip()
                if _tail:
                    with st.expander("🔍 Error details — last log lines"):
                        st.code(_tail, language=None)
                        st.caption(f"Full log: `{run['log_path']}`")
                        _free = _disk_free_mb()
                        if _free < 500:
                            st.error(
                                f"C: drive has only **{_free:.0f} MB free** — "
                                f"free up disk space and retry."
                            )
            except Exception:
                pass

        if s in ("running", "complete", "stopped"):
            st.progress(pct, text=f"{pct*100:.0f}%")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("✅ Pass", f"{passes:,}")
            m2.metric("❌ Fail", f"{fails:,}")
            m3.metric("⏭️ Skip", f"{skips:,}")
            m4.metric(
                "Pass Rate",
                f"{rate:.1f}%",
                help=f"Pass / (Pass + Fail) = {passes:,} / {processed:,} — Skips excluded",
            )

# ---------------------------------------------------------------------------
# Page 1 — Run Manager
# ---------------------------------------------------------------------------

def page_run_manager() -> None:
    _header()

    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        max_workers = st.slider(
            "Parallel workers", 1, 6, 3,
            help="Max input files processed simultaneously. "
                 "Finished workers auto-pick the next queued file.",
        )
        st.divider()
        st.markdown(
            f"**Input folder**  \n"
            f"`RunQueue/`  \n\n"
            f"Drop `.xlsx` files here before clicking Start.  \n"
            f"Non-`.xlsx` files are automatically ignored.  \n\n"
            f"⚠️ **Maximum {MAX_FILES} files per batch.**",
        )

    col_btn, col_msg = st.columns([2, 10])
    with col_btn:
        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()

    xlsx_files, skipped, error = _scan_queue()

    with col_msg:
        if error:
            st.error(error)
        elif xlsx_files:
            st.success(
                f"📁 **{len(xlsx_files)} / {MAX_FILES}** .xlsx files ready in `RunQueue/`"
            )
        else:
            st.info(
                f"📁 **RunQueue/ is empty.** Drop `.xlsx` input files into  \n"
                f"`{RUN_QUEUE}`  \nthen click **Refresh**."
            )

    if skipped:
        with st.expander(f"⚠️ {len(skipped)} file(s) skipped (not .xlsx)"):
            for s in skipped:
                st.caption(f"• {s}")

    # Disk space warning — Chrome needs ~200 MB to launch; warn below 500 MB.
    _free_mb = _disk_free_mb()
    if _free_mb < 500:
        st.warning(
            f"⚠️ **Low disk space: {_free_mb:.0f} MB free on C:.**  "
            f"Runs will likely fail — Chrome alone needs ~200 MB.  "
            f"Free up disk space (empty Recycle Bin, clear Windows Temp, "
            f"remove large files from Downloads) before starting."
        )

    runs_active = st.session_state.get("_runs_active", False)
    # Also check the actual runs list — guards against _runs_active flag getting
    # out of sync with the real run state (e.g. flag reset by completion banner
    # while a run is still in progress).
    _live_runs  = st.session_state.get("_runs", [])
    runs_in_progress = any(r["status"] in ("running", "queued") for r in _live_runs)
    can_start   = bool(xlsx_files) and not error and not runs_active and not runs_in_progress
    sc, _ = st.columns([2, 10])
    with sc:
        clicked = st.button(
            "▶  Start", type="primary",
            disabled=not can_start,
            use_container_width=True,
        )

    if clicked and can_start:
        ts   = datetime.now().strftime("%Y%m%d_%H%M")
        runs = []
        for idx, f in enumerate(xlsx_files):
            state = _detect_state(f)
            # Include per-file index so two files with the same state/timestamp
            # never produce the same run_id (which would cause Streamlit key collisions).
            run_id = f"{ts}_{state}_{str(idx + 1).zfill(3)}"
            # Pre-read row count so the progress bar shows a real total immediately
            # rather than "X / ? rows" until the first log batch line appears.
            # max_row is unreliable in read_only mode (returns None for many xlsx files);
            # iterate column A from row 2 instead — fast even for 1000+ row files.
            try:
                import openpyxl as _oxl
                _wb = _oxl.load_workbook(str(f), read_only=True, data_only=True)
                _ws = _wb.active
                _total = sum(1 for _ in _ws.iter_rows(min_row=2, max_col=1))
                _wb.close()
            except Exception:
                _total = 0
            runs.append({
                "input_file": str(f),
                "filename":   f.name,
                "state":      state,
                "run_id":     run_id,
                "log_path":   "",
                "_proc":      None,
                "status":     "queued",
                "done": 0, "total": _total, "passes": 0, "fails": 0, "skips": 0,
                "started_at": None, "finished_at": None,
            })
        st.session_state._runs        = runs
        st.session_state._runs_active = True
        threading.Thread(
            target=_scheduler, args=(runs, max_workers), daemon=True
        ).start()
        st.rerun()

    runs: list[dict] = st.session_state.get("_runs", [])
    if not runs:
        return

    st.subheader("Run Queue")
    _seen_run_ids: set[str] = set()
    for run in runs:
        if run["run_id"] in _seen_run_ids:
            continue
        _seen_run_ids.add(run["run_id"])
        if run["status"] == "running" and run.get("log_path"):
            prog = _parse_log(run["log_path"])
            run["done"]   = prog["done"]
            run["passes"] = prog["passes"]
            run["fails"]  = prog["fails"]
            run["skips"]  = prog["skips"]
            # total is always the xlsx row count pre-read at queue time;
            # never overwrite it with the log's "done of M" which can be
            # smaller when rows are filtered by state.
        _run_card(run, runs)

    # Completion banner — shown when no run is still running or queued
    if runs and all(r["status"] in ("complete", "error", "stopped") for r in runs):
        total_p   = sum(r["passes"] for r in runs)
        total_f   = sum(r["fails"]  for r in runs)
        total_s   = sum(r["skips"]  for r in runs)
        processed = total_p + total_f
        rate      = total_p / processed * 100 if processed else 0
        n_complete = sum(1 for r in runs if r["status"] == "complete")
        n_stopped  = sum(1 for r in runs if r["status"] == "stopped")
        n_error    = sum(1 for r in runs if r["status"] == "error")
        if n_stopped or n_error:
            parts = [f"{n_complete} complete"]
            if n_stopped: parts.append(f"{n_stopped} stopped")
            if n_error:   parts.append(f"{n_error} error")
            banner_icon, banner_verb = "⚠️", f"All runs finished ({', '.join(parts)})"
        else:
            banner_icon, banner_verb = "✅", f"All {len(runs)} runs complete!"
        st.success(
            f"{banner_icon}  **{banner_verb}**  "
            f"&nbsp;&nbsp;**{total_p + total_f + total_s:,}** rows — "
            f"**{total_p:,}** Pass / **{total_f:,}** Fail / **{total_s:,}** Skip "
            f"({rate:.1f}% pass rate, skips excluded).  \n"
            f"Switch to **Results Viewer** in the sidebar to explore outputs."
        )
        st.session_state._runs_active = False

    if any(r["status"] in ("running", "queued") for r in runs):
        time.sleep(REFRESH_S)
        st.rerun()

# ---------------------------------------------------------------------------
# Page 2 — Results Viewer
# ---------------------------------------------------------------------------

def page_results() -> None:
    _header()

    run_options: dict[str, Path] = {}
    if OUTPUT_ROOT.exists():
        for yyyymm in sorted(OUTPUT_ROOT.iterdir(), reverse=True):
            if not yyyymm.is_dir():
                continue
            for run_dir in sorted(yyyymm.iterdir(), reverse=True):
                std_dir = run_dir / "Standard"
                if not std_dir.is_dir():
                    continue
                csvs = sorted(std_dir.glob("Standard_*.csv"), reverse=True)
                if csvs:
                    run_options[run_dir.name] = csvs[0]

    if not run_options:
        st.info(
            "No completed runs found. Run a batch from the **Run Manager** page first."
        )
        return

    selected = st.selectbox("Select Run", list(run_options.keys()))
    csv_path = run_options[selected]

    try:
        df = pd.read_csv(csv_path, dtype=str)
    except Exception as e:
        st.error(f"Could not load output CSV: {e}")
        return

    df["display_status"] = _derive_display_status(df)

    total     = len(df)
    passes    = int((df["display_status"] == "Pass").sum())
    fails     = int((df["display_status"] == "Fail").sum())
    skips     = int((df["display_status"] == "Skip").sum())
    processed = passes + fails
    rate      = passes / processed * 100 if processed else 0

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Records", f"{total:,}")
    m2.metric("✅ Pass",        f"{passes:,}")
    m3.metric("❌ Fail",        f"{fails:,}")
    m4.metric("⏭️ Skip",        f"{skips:,}",
              help="CAPTCHA-blocked or no routing configured for this provider type")
    m5.metric("Pass Rate",      f"{rate:.1f}%",
              help=f"Pass / (Pass + Fail) = {passes:,} / {processed:,} — Skips excluded")

    st.divider()

    f1, f2, f3, f4 = st.columns([2, 2, 2, 4])

    status_filter = f1.selectbox("Status", ["All", "Pass", "Fail", "Skip"])

    state_vals = (
        ["All"] + sorted(df["lic_state"].dropna().unique().tolist())
        if "lic_state" in df.columns else ["All"]
    )
    state_filter = f2.selectbox("State", state_vals)

    prov_vals = (
        ["All"] + sorted(df["prov_type"].dropna().unique().tolist())
        if "prov_type" in df.columns else ["All"]
    )
    prov_filter = f3.selectbox("Prov Type", prov_vals)
    search      = f4.text_input("🔍 Search", placeholder="Name or license number")

    filtered = df.copy()
    if status_filter != "All":
        filtered = filtered[filtered["display_status"] == status_filter]
    if state_filter != "All" and "lic_state" in df.columns:
        filtered = filtered[filtered["lic_state"] == state_filter]
    if prov_filter != "All" and "prov_type" in df.columns:
        filtered = filtered[filtered["prov_type"] == prov_filter]
    if search:
        mask = pd.Series(False, index=filtered.index)
        for col in ("first_name", "last_name", "license_id"):
            if col in filtered.columns:
                mask |= filtered[col].str.contains(search, case=False, na=False)
        filtered = filtered[mask]

    st.caption(
        f"Showing **{len(filtered):,}** of **{total:,}** records  ·  "
        f"Status: **{status_filter}**"
    )

    key_cols = [
        "display_status", "first_name", "last_name", "prov_type", "lic_state",
        "license_id", "reason", "board_name", "license_expiry", "match_method",
    ]
    display_cols = [c for c in key_cols if c in filtered.columns] + \
                   [c for c in filtered.columns if c not in key_cols and c != "status"]

    st.dataframe(
        filtered[display_cols],
        use_container_width=True,
        hide_index=True,
        height=500,
        column_config={
            "display_status": st.column_config.TextColumn("Status", width="small"),
            "reason":         st.column_config.TextColumn("Reason", width="large"),
            "license_expiry": st.column_config.TextColumn("Expiry", width="small"),
        },
    )

    export_df = filtered[display_cols].copy()
    export_df.insert(
        export_df.columns.get_loc("display_status") + 1,
        "original_status",
        filtered["status"] if "status" in filtered.columns else "",
    )
    export_tag = "_".join(filter(None, [
        status_filter if status_filter != "All" else "",
        state_filter  if state_filter  != "All" else "",
        prov_filter   if prov_filter   != "All" else "",
    ])) or "All"
    st.download_button(
        "📥 Export filtered CSV",
        data=export_df.to_csv(index=False).encode("utf-8"),
        file_name=f"psv_{selected}_{export_tag}.csv",
        mime="text/csv",
    )

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    st.session_state.setdefault("_runs",        [])
    st.session_state.setdefault("_runs_active", False)

    with st.sidebar:
        st.markdown(
            "<h3 style='color:#CC0000;margin-bottom:4px;'>🏥 PSV Runner</h3>",
            unsafe_allow_html=True,
        )
        page = st.radio(
            "Navigate",
            ["Run Manager", "Results Viewer"],
            label_visibility="collapsed",
        )
        st.divider()
        st.caption(f"Workspace: `{PSV_DEV.name}`")

    if page == "Run Manager":
        page_run_manager()
    else:
        page_results()


if __name__ == "__main__":
    main()
