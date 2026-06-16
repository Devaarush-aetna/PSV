"""
rebuild_inventory.py — Regenerate board_inventory.xlsx from current state.

Walks every sites/*/config.yaml, then pulls the most recent smoke status for
each board from the smoke_*.txt logs in output/, and writes a 10-column
spreadsheet.

Columns:
  # | State | Source ID | Board Name | Profession Codes | Source URL |
  Ingestion Type | Archetype | Smoke Test Status | Skip Reason / Notes
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent
SITES = ROOT / "sites"
OUTPUT_DIR = ROOT / "output"
INVENTORY_PATH = ROOT / "board_inventory.xlsx"

# Smoke logs in chronological order — later ones override earlier
SMOKE_LOGS = [
    "smoke_CO_NY_20260607.txt",
    "smoke_regression_20260607.txt",
    "smoke_MD_boards_20260608.txt",
    "smoke_regression_20260608.txt",
    "smoke_new_boards_20260611.txt",
    "smoke_fail_rerun_20260611.txt",
    "smoke_forceskip_20260611.txt",
    "smoke_skip_followup_20260611.txt",
    "smoke_wy_trio_final_20260611.txt",
]

ROW_RE = re.compile(
    r"^(?P<sid>[A-Z][A-Z0-9_]+)\s+"
    r"(?P<mode>\S+)\s+"
    r"(?P<query>.+?)\s+"
    r"(?P<status>PASS|FAIL|SKIP|MISSING)\s+"
    r"(?P<detail>.*)$"
)


def parse_smoke_log(path: Path) -> dict[str, tuple[str, str]]:
    """Return {source_id: (status, detail)} from a smoke run log.

    Scans every line that matches the result-row regex (any encoding tolerated).
    """
    if not path.exists():
        return {}
    out: dict[str, tuple[str, str]] = {}
    raw_bytes = path.read_bytes()
    text = raw_bytes.decode("utf-8", errors="replace")
    for line in text.splitlines():
        # Strip leading/trailing whitespace; only match lines that look like a result row
        stripped = line.strip()
        if not stripped:
            continue
        # Must end in a status keyword followed by detail (rough filter)
        if not re.search(r"\b(PASS|FAIL|SKIP|MISSING)\b", stripped):
            continue
        # Result rows start at column 0 with the source ID — exclude log lines
        if line[:1] not in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                            "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
                            "U", "V", "W", "X", "Y", "Z"):
            continue
        m = ROW_RE.match(stripped)
        if not m:
            continue
        sid = m.group("sid")
        status = m.group("status")
        detail = m.group("detail").strip()
        out[sid] = (status, detail)
    return out


def latest_smoke_status() -> dict[str, tuple[str, str]]:
    """Merge smoke logs in chronological order; latest wins."""
    merged: dict[str, tuple[str, str]] = {}
    for name in SMOKE_LOGS:
        for sid, val in parse_smoke_log(OUTPUT_DIR / name).items():
            merged[sid] = val
    return merged


def load_config(path: Path) -> dict:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def archetype_to_ingestion(arch: str) -> str:
    mapping = {
        "classic_html_form": "Web Scraping — HTML Form",
        "thentia_cloud": "Web Scraping — Thentia Cloud",
        "ag_grid_spa": "Web Scraping — AG Grid SPA",
        "csv_bulk": "CSV Bulk Roster",
        "pdf_bulk": "PDF Bulk Roster",
        "socrata_bulk_csv": "API — Socrata JSON (browser-based)",
        "socrata_api": "API — Socrata JSON (direct HTTP)",
        "aithent_portal_xls": "Web Scraping — Aithent Portal (XLS)",
        "playwright_form": "Web Scraping — Playwright Form",
    }
    return mapping.get(arch, f"Web Scraping — {arch}")


def smoke_status_from_config(cfg: dict, smoke_index: dict[str, tuple[str, str]]) -> tuple[str, str]:
    """Return (Status, Notes) for the inventory.

    Precedence:
      1. compliance.requires_captcha → SKIP (captcha_blocked)
      2. config smoke_test.skip == true → SKIP (use config skip_reason — current truth)
      3. live smoke log → use latest captured status/detail
      4. nothing captured → "—" / no run note
    """
    ident = cfg.get("identity", {}) or {}
    sid = ident.get("source_id", "")
    compliance = cfg.get("compliance", {}) or {}
    smoke = cfg.get("smoke_test", {}) or {}

    if compliance.get("requires_captcha"):
        cfg_reason = (smoke.get("skip_reason") or "").strip()
        return "SKIP", (f"requires_captcha: true — {cfg_reason}".strip(" —")) or "requires_captcha: true"

    cfg_skip = smoke.get("skip", False)
    cfg_reason = (smoke.get("skip_reason") or "").strip()

    # Config-declared skip beats any older PASS log entry — config is current truth.
    if cfg_skip:
        return "SKIP", cfg_reason or "skip flag set in config"

    live = smoke_index.get(sid)
    if live:
        return live  # (status, detail)

    return "—", "no smoke run captured for this board"


def main() -> int:
    smoke_index = latest_smoke_status()

    rows = []
    for cfg_path in sorted(SITES.glob("*/config.yaml")):
        cfg = load_config(cfg_path)
        ident = cfg.get("identity", {}) or {}
        sid = ident.get("source_id") or cfg_path.parent.name
        state = ident.get("state", "")
        board_name = ident.get("board_name", "")
        profession = ident.get("profession_codes") or []
        if isinstance(profession, list):
            profession_str = ", ".join(str(p) for p in profession)
        else:
            profession_str = str(profession)
        url = ident.get("base_url", "")
        arch = ident.get("archetype", "")
        ingestion = archetype_to_ingestion(arch)
        status, notes = smoke_status_from_config(cfg, smoke_index)

        rows.append({
            "state": state,
            "sid": sid,
            "board_name": board_name,
            "profession": profession_str,
            "url": url,
            "ingestion": ingestion,
            "archetype": arch,
            "status": status,
            "notes": notes,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "#", "State", "Source ID", "Board Name", "Profession Codes",
        "Source URL", "Ingestion Type", "Archetype",
        "Smoke Test Status", "Skip Reason / Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    status_fill = {
        "PASS": PatternFill("solid", fgColor="C6EFCE"),
        "FAIL": PatternFill("solid", fgColor="FFC7CE"),
        "SKIP": PatternFill("solid", fgColor="FFEB9C"),
        "MISSING": PatternFill("solid", fgColor="D9D9D9"),
    }

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r["state"],
            r["sid"],
            r["board_name"],
            r["profession"],
            r["url"],
            r["ingestion"],
            r["archetype"],
            r["status"],
            r["notes"],
        ])
        status_cell = ws.cell(row=ws.max_row, column=9)
        if r["status"] in status_fill:
            status_cell.fill = status_fill[r["status"]]
            status_cell.font = Font(bold=True)
        for col in range(1, 11):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    widths = [4, 6, 22, 50, 40, 60, 36, 22, 10, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    wb.save(INVENTORY_PATH)

    # Print a summary
    from collections import Counter
    status_counts = Counter(r["status"] for r in rows)
    print(f"Wrote {len(rows)} boards to {INVENTORY_PATH}")
    print("Smoke split:", dict(status_counts))
    return 0


if __name__ == "__main__":
    sys.exit(main())
